"""Tests for the MITRESaw evidence report module."""

import json
import os
import tempfile

import pytest
from openpyxl import load_workbook

from src.evidence_report import (
    generate_evidence_report,
    _clean_procedure_text,
)

# ---------------------------------------------------------------------------
# Column schema (13 columns, 1-indexed):
#  1  Evidential Element    2  Threat Group       3  Technique ID
#  4  Technique Name        5  Tactic             6  Platforms
#  7  Framework             8  Source Type        9  Procedure Example
# 10  Detection Guidance   11  Log Sources       12  Reference URL
# 13  Navigation Layer URL
# ---------------------------------------------------------------------------


def _make_row(
    group_id="G0049",
    group_name="OilRig",
    technique_id="T1059.001",
    technique_name="PowerShell",
    tactic="Execution",
    platforms="Windows",
    framework="Enterprise",
    procedure="OilRig has run `net user /domain`.",
    evidence=None,
    detectable_via="Process Creation",
):
    if evidence is None:
        evidence = {"cmd": ["net user /domain"]}
    return {
        "group_sw_id": group_id,
        "group_sw_name": group_name,
        "technique_id": technique_id,
        "technique_name": technique_name,
        "tactic": tactic,
        "platforms": platforms,
        "framework": framework,
        "procedure_example": procedure,
        "evidence": json.dumps(evidence),
        "detectable_via": detectable_via,
    }


# ---------------------------------------------------------------------------
# Atomisation / dedup
# ---------------------------------------------------------------------------


def test_atomise_cmd():
    """
    WHAT: Two cmd indicators in one row → two data rows in the XLSX.
    WHY:  Report must expand indicators into individual rows (atomise).
    PASS: Both indicator values appear as Evidential Element (col 1).
    """
    row = _make_row(evidence={"cmd": ["net user /domain", "net group /domain"]})
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        generate_evidence_report([row], path)
        wb = load_workbook(path)
        ws = wb.active
        data_rows = [ws.cell(row=r, column=1).value for r in range(4, ws.max_row + 1)]
        assert len(data_rows) == 2
        assert "net user /domain" in data_rows
        assert "net group /domain" in data_rows
    finally:
        os.unlink(path)


def test_atomise_empty_evidence():
    """
    WHAT: A row with no evidence produces a placeholder row.
    WHY:  Every technique must appear in the report even with no indicators.
    PASS: Col 1 of the single data row contains '(no extractable indicators)'.
    """
    row = _make_row(evidence={})
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        generate_evidence_report([row], path)
        wb = load_workbook(path)
        ws = wb.active
        val = ws.cell(row=4, column=1).value
        assert val == "(no extractable indicators)"
        assert ws.cell(row=5, column=1).value is None
    finally:
        os.unlink(path)


def test_dedup():
    """
    WHAT: Two rows with identical (group, technique, indicator) are merged.
    WHY:  Same indicator from two sources must not appear twice in the report.
    PASS: Exactly one data row written.
    """
    row1 = _make_row(evidence={"cmd": ["net user /domain"]})
    row2 = _make_row(evidence={"cmd": ["net user /domain"]})
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        generate_evidence_report([row1, row2], path)
        wb = load_workbook(path)
        ws = wb.active
        data_rows = [
            ws.cell(row=r, column=1).value
            for r in range(4, ws.max_row + 1)
            if ws.cell(row=r, column=1).value is not None
        ]
        assert len(data_rows) == 1
    finally:
        os.unlink(path)


def test_column_count():
    """
    WHAT: Every data row must have exactly 13 populated columns.
    WHY:  Sparse rows indicate a schema mismatch or missing field.
    PASS: All 13 columns are non-None for a standard indicator row.
    """
    row = _make_row()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        generate_evidence_report([row], path)
        wb = load_workbook(path)
        ws = wb.active
        for r in range(4, ws.max_row + 1):
            if ws.cell(row=r, column=1).value is None:
                break
            populated = sum(
                1 for c in range(1, 14) if ws.cell(row=r, column=c).value is not None
            )
            assert populated == 13, f"Row {r} has {populated} populated columns, expected 13"
    finally:
        os.unlink(path)


# ---------------------------------------------------------------------------
# Individual column content
# ---------------------------------------------------------------------------


def test_procedure_example_column():
    """
    WHAT: Col 9 contains the cleaned procedure text for each row.
    WHY:  Analysts need to see the original MITRE procedure sentence.
    PASS: Col 9 contains the command that was in the procedure text.
    """
    procedure = "OilRig has run `net user /domain` on target systems."
    row = _make_row(procedure=procedure, evidence={"cmd": ["net user /domain"]})
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        generate_evidence_report([row], path)
        wb = load_workbook(path)
        ws = wb.active
        val = ws.cell(row=4, column=9).value
        assert "net user" in val
    finally:
        os.unlink(path)


def test_cve_detection_context():
    """
    WHAT: CVE indicators produce detection guidance referencing CISA KEV.
    WHY:  CVEs are tracked in KEV; detection guidance should reflect that.
    PASS: Col 10 (Detection Guidance) contains 'CISA KEV'.
    """
    row = _make_row(
        evidence={"cve": [{"CVE-2020-0688": "Exchange vuln"}]},
        procedure="Exploited CVE-2020-0688.",
    )
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        generate_evidence_report([row], path)
        wb = load_workbook(path)
        ws = wb.active
        det = ws.cell(row=4, column=10).value
        assert "CISA KEV" in det
    finally:
        os.unlink(path)


def test_reg_detection_context():
    """
    WHAT: Registry path indicators produce detection guidance referencing Sysmon EID 12.
    WHY:  Registry key creation/modification is captured by Sysmon Event ID 12.
    PASS: Col 10 (Detection Guidance) contains 'Sysmon EID 12'.
    """
    row = _make_row(
        evidence={"reg": ["HKCU\\Software\\Microsoft\\Windows\\CurrentVersion\\Run"]},
        procedure="Added HKCU\\Software\\Microsoft\\Windows\\CurrentVersion\\Run key.",
    )
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        generate_evidence_report([row], path)
        wb = load_workbook(path)
        ws = wb.active
        det = ws.cell(row=4, column=10).value
        assert "Sysmon EID 12" in det
    finally:
        os.unlink(path)


def test_technique_url_construction():
    """
    WHAT: Reference URL (col 12) is constructed from the technique ID.
    WHY:  Analysts must be able to navigate directly to the ATT&CK page.
    PASS: Col 12 contains the canonical MITRE ATT&CK technique URL.
    """
    row = _make_row(
        technique_id="T1059.001",
        procedure="The group used PowerShell scripts.",
        evidence={"cmd": ["powershell"]},
    )
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        generate_evidence_report([row], path)
        wb = load_workbook(path)
        ws = wb.active
        url = ws.cell(row=4, column=12).value
        assert url == "https://attack.mitre.org/techniques/T1059/001/"
    finally:
        os.unlink(path)


def test_platforms_column():
    """
    WHAT: Platforms value (col 6) is written correctly from input row.
    WHY:  Filtering by platform requires this field to be accurate.
    PASS: Col 6 contains 'Windows'.
    """
    row = _make_row(platforms="Windows, Linux")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        generate_evidence_report([row], path)
        wb = load_workbook(path)
        ws = wb.active
        val = ws.cell(row=4, column=6).value  # Platforms is col 6
        assert "Windows" in val
    finally:
        os.unlink(path)


def test_framework_column():
    """
    WHAT: Framework value (col 7) is written correctly from input row.
    WHY:  Multi-framework runs produce mixed output; column must track source.
    PASS: Col 7 contains 'Enterprise'.
    """
    row = _make_row(framework="Enterprise")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        generate_evidence_report([row], path)
        wb = load_workbook(path)
        ws = wb.active
        val = ws.cell(row=4, column=7).value  # Framework is col 7
        assert val == "Enterprise"
    finally:
        os.unlink(path)


def test_log_sources_column():
    """
    WHAT: Log sources (col 11) are written from the detectable_via field.
    WHY:  Analysts need to know which log sources to query.
    PASS: Col 11 contains 'Sysmon'.
    """
    row = _make_row(detectable_via="Sysmon: 1; Security EventLog: 4688")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        generate_evidence_report([row], path)
        wb = load_workbook(path)
        ws = wb.active
        val = ws.cell(row=4, column=11).value  # Log Sources is col 11
        assert "Sysmon" in val
    finally:
        os.unlink(path)


def test_output_file_created():
    """
    WHAT: generate_evidence_report() writes a non-empty XLSX file.
    WHY:  Basic smoke test — if the file is empty/missing the pipeline broke.
    PASS: File exists and has size > 0.
    """
    row = _make_row()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        generate_evidence_report([row], path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0
    finally:
        os.unlink(path)


# ---------------------------------------------------------------------------
# Multi-sheet / Group Summary
# ---------------------------------------------------------------------------


def test_group_summary_sheet_headers():
    """
    WHAT: Group Summary sheet has the expected 5-column header row.
    WHY:  Column labels drive analyst workflow; wrong labels cause confusion.
    PASS: Cols 1-5 of row 3 match the canonical header names.
    """
    row = _make_row()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        generate_evidence_report([row], path)
        wb = load_workbook(path)
        ws2 = wb["Group Summary"]
        headers = [ws2.cell(row=3, column=c).value for c in range(1, 6)]
        assert headers == [
            "Group Name", "Technique Count", "Indicator Count",
            "Tactic Coverage", "Top Tactic",
        ]
        # Data row 4 must have a numeric indicator count
        assert isinstance(ws2.cell(row=4, column=3).value, (int, float))
    finally:
        os.unlink(path)


def test_technique_matrix_with_multiple_groups():
    """
    WHAT: Technique Matrix sheet is created and sorted by group count descending.
    WHY:  Cross-group technique frequency is key for prioritising detections.
    PASS: T1059.001 (2 groups) appears before T1078 (1 group).
    """
    row1 = _make_row(group_name="OilRig", technique_id="T1059.001")
    row2 = _make_row(group_name="APT33", technique_id="T1059.001",
                     evidence={"cmd": ["powershell -enc"]})
    row3 = _make_row(group_name="APT33", technique_id="T1078",
                     technique_name="Valid Accounts",
                     evidence={"cmd": ["net user"]})
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        generate_evidence_report([row1, row2, row3], path)
        wb = load_workbook(path)
        assert "Technique Matrix" in wb.sheetnames
        ws = wb["Technique Matrix"]
        assert ws.cell(row=3, column=1).value == "T1059.001"
        assert ws.cell(row=3, column=4).value == 2
        assert ws.cell(row=4, column=1).value == "T1078"
        assert ws.cell(row=4, column=4).value == 1
    finally:
        os.unlink(path)


def test_technique_matrix_not_created_single_group():
    """
    WHAT: Technique Matrix sheet is NOT created when only one group is present.
    WHY:  Cross-group comparison is meaningless for a single actor.
    PASS: 'Technique Matrix' is absent from workbook sheet names.
    """
    row = _make_row()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        generate_evidence_report([row], path)
        wb = load_workbook(path)
        assert "Technique Matrix" not in wb.sheetnames
    finally:
        os.unlink(path)


# ---------------------------------------------------------------------------
# Procedure text cleaning
# ---------------------------------------------------------------------------


def test_clean_procedure_markdown_links():
    """
    WHAT: _clean_procedure_text() resolves markdown links to 'Name (ID)' form.
    WHY:  MITRE procedure text uses markdown links; XLSX must show plain text.
    PASS: Links are replaced with 'Axiom (G0001)' / 'Mimikatz (S0002)'.
    """
    text = "[Axiom](https://attack.mitre.org/groups/G0001) used [Mimikatz](https://attack.mitre.org/software/S0002)."
    result = _clean_procedure_text(text)
    assert "Axiom (G0001)" in result
    assert "Mimikatz (S0002)" in result
    assert "[" not in result
    assert "](" not in result


def test_clean_procedure_citations_removed():
    """
    WHAT: _clean_procedure_text() strips (Citation: ...) tags from text.
    WHY:  Citation tags are MITRE markdown noise; they must not appear in XLSX.
    PASS: 'Citation' and 'FireEye' are absent from cleaned output.
    """
    text = "APT29 used PowerShell.(Citation: FireEye APT29 2020)(Citation: CISA Alert)"
    result = _clean_procedure_text(text)
    assert "APT29 used PowerShell" in result
    assert "Citation" not in result
    assert "FireEye" not in result


def test_clean_procedure_empty():
    """
    WHAT: _clean_procedure_text() handles empty string and None gracefully.
    WHY:  Null procedure fields must not crash the report generator.
    PASS: Empty string returns empty string; None returns None.
    """
    assert _clean_procedure_text("") == ""
    assert _clean_procedure_text(None) is None
