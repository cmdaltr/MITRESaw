#!/usr/bin/env python3
"""
mitre_all_groups_evidence.py
────────────────────────────
Generates a styled XLSX evidence report for ALL MITRE ATT&CK threat groups.

For each group it traverses all three STIX attribution paths:
  Path 1 — Direct:        group --uses--> technique
  Path 2 — Via Software:  group --uses--> software --uses--> technique
  Path 3 — Via Campaign:  campaign --attributed-to--> group
                          campaign --uses--> technique

For each (group, technique) pair it:
  • Extracts the procedure_example text from the STIX relationship
  • Mines that text for MITRE-documented invocation strings (backtick-wrapped
    commands, CLI patterns, CVEs, registry paths, file paths)
  • Atomises the evidence JSON into one row per indicator
  • Builds Col 7 (Contextual Evidence) from the extracted invocations

Usage:
    python3 mitre_all_groups_evidence.py

Options (edit the CONFIG block below):
    FRAMEWORK   — Enterprise | ICS | Mobile  (default: Enterprise)
    FORCE_FETCH — re-download even if cache exists (default: False)
    OUTPUT_DIR  — where to write the XLSX (default: current directory)
    NATION_FILTER — list of keywords to filter groups by description,
                    e.g. ["iran", "china", "russia"] or [] for all groups

Requirements:
    pip install openpyxl requests
"""

# ─── CONFIG ──────────────────────────────────────────────────────────────────
FRAMEWORK    = "Enterprise"   # Enterprise | ICS | Mobile
FORCE_FETCH  = False          # True = always re-download STIX bundle
OUTPUT_DIR   = "."            # directory for output XLSX
NATION_FILTER = []            # e.g. ["iran"] for Iranian only, [] for all groups
MAX_ROWS_PER_SHEET = 50000    # split into multiple sheets if exceeded
COLLECT_REFERENCES = False    # True = fetch citation URLs and extract content (-R flag)
# ─────────────────────────────────────────────────────────────────────────────

import json
import os
import re
import sys
import time
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")

# ─────────────────────────────────────────────────────────────────────────────
# STIX BUNDLE DOWNLOAD
# ─────────────────────────────────────────────────────────────────────────────

STIX_URLS = {
    "Enterprise": "https://raw.githubusercontent.com/mitre/cti/master/enterprise-attack/enterprise-attack.json",
    "ICS":        "https://raw.githubusercontent.com/mitre/cti/master/ics-attack/ics-attack.json",
    "Mobile":     "https://raw.githubusercontent.com/mitre/cti/master/mobile-attack/mobile-attack.json",
}

def download_stix(framework: str, force: bool = False) -> dict:
    cache_path = Path(f".cache_{framework.lower()}_attack.json")
    if cache_path.exists() and not force:
        age_days = (time.time() - cache_path.stat().st_mtime) / 86400
        if age_days < 7:
            print(f"[+] Using cached STIX data ({age_days:.1f} days old): {cache_path}")
            with open(cache_path) as f:
                return json.load(f)
        else:
            print(f"[!] Cache is {age_days:.1f} days old — refreshing")

    url = STIX_URLS.get(framework)
    if not url:
        sys.exit(f"Unknown framework: {framework}")

    print(f"[+] Downloading MITRE ATT&CK {framework} STIX bundle...")
    print(f"    URL: {url}")

    req = urllib.request.Request(url, headers={"User-Agent": "MITRESaw-EvidenceReport/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            raw = resp.read()
    except Exception as e:
        sys.exit(f"Download failed: {e}\nCheck your internet connection and try again.")

    print(f"[+] Downloaded {len(raw)/1024/1024:.1f} MB")
    data = json.loads(raw)

    with open(cache_path, "w") as f:
        json.dump(data, f)
    print(f"[+] Cached to {cache_path}")
    return data


# ─────────────────────────────────────────────────────────────────────────────
# STIX PARSING — BUILD LOOKUP INDEXES
# ─────────────────────────────────────────────────────────────────────────────

def build_indexes(stix_bundle: dict) -> dict:
    """
    Returns a dict of indexes needed for three-path attribution traversal:
      objects_by_id         — STIX ID → object
      groups                — list of intrusion-set objects (not revoked/deprecated)
      techniques            — attack-pattern ID → technique object
      relationships_by_src  — source_ref → list of relationship objects
      relationships_by_tgt  — target_ref → list of relationship objects
      software_by_id        — tool/malware ID → object
      campaigns_by_id       — campaign ID → object
    """
    objects = stix_bundle.get("objects", [])
    idx = defaultdict(list)
    by_id = {}
    groups = []
    techniques = {}
    software = {}
    campaigns = {}

    for obj in objects:
        oid  = obj.get("id", "")
        otype = obj.get("type", "")
        by_id[oid] = obj

        # Skip revoked and deprecated
        if obj.get("revoked", False) or obj.get("x_mitre_deprecated", False):
            continue

        if otype == "intrusion-set":
            groups.append(obj)
        elif otype == "attack-pattern":
            techniques[oid] = obj
        elif otype in ("tool", "malware"):
            software[oid] = obj
        elif otype == "campaign":
            campaigns[oid] = obj
        elif otype == "relationship":
            idx["src"].append(obj)  # will re-index below

    # Index relationships by source and target
    rels_by_src = defaultdict(list)
    rels_by_tgt = defaultdict(list)
    all_rels     = [o for o in objects if o.get("type") == "relationship"]
    for rel in all_rels:
        rels_by_src[rel.get("source_ref", "")].append(rel)
        rels_by_tgt[rel.get("target_ref", "")].append(rel)

    print(f"[+] Indexed: {len(groups)} groups, {len(techniques)} techniques, "
          f"{len(software)} software, {len(campaigns)} campaigns, "
          f"{len(all_rels)} relationships")

    return {
        "by_id":        by_id,
        "groups":       groups,
        "techniques":   techniques,
        "software":     software,
        "campaigns":    campaigns,
        "rels_by_src":  rels_by_src,
        "rels_by_tgt":  rels_by_tgt,
    }


# ─────────────────────────────────────────────────────────────────────────────
# THREE-PATH ATTRIBUTION
# ─────────────────────────────────────────────────────────────────────────────

def get_group_techniques(group_id: str, idx: dict) -> list[dict]:
    """
    Returns list of dicts:
      technique_id, technique_name, tactic, procedure_example,
      source_path (direct|software|campaign), source_name
    """
    techniques = idx["techniques"]
    rels_by_src = idx["rels_by_src"]
    rels_by_tgt = idx["rels_by_tgt"]
    by_id       = idx["by_id"]

    results = []
    seen_keys = set()  # (technique_id, procedure_example_hash) — dedup

    def tactic_from_technique(t_obj):
        phases = t_obj.get("kill_chain_phases", [])
        return ", ".join(
            p["phase_name"].replace("-", " ").title()
            for p in phases
            if p.get("kill_chain_name") == "mitre-attack"
        )

    def tech_id_from_obj(t_obj):
        for ref in t_obj.get("external_references", []):
            if ref.get("source_name") == "mitre-attack":
                return ref.get("external_id", "")
        return ""

    def add_result(t_obj, rel_obj, path_label, src_name=""):
        proc_text = rel_obj.get("description", "") if isinstance(rel_obj, dict) else rel_obj
        ext_refs = rel_obj.get("external_references", []) if isinstance(rel_obj, dict) else []
        tid = tech_id_from_obj(t_obj)
        if not tid:
            return
        tname  = t_obj.get("name", "")
        tactic = tactic_from_technique(t_obj)
        key    = (tid, hash(proc_text))
        if key in seen_keys:
            return
        seen_keys.add(key)
        results.append({
            "technique_id":       tid,
            "technique_name":     tname,
            "tactic":             tactic,
            "procedure_example":  proc_text,
            "source_path":        path_label,
            "source_name":        src_name,
            "technique_stix_id":  t_obj.get("id", ""),
            "external_references": ext_refs,
        })

    # PATH 1 — Direct: group → technique
    for rel in rels_by_src.get(group_id, []):
        if rel.get("relationship_type") != "uses":
            continue
        tgt = rel.get("target_ref", "")
        if tgt in techniques:
            add_result(techniques[tgt], rel, "direct")

    # PATH 2 — Via Software: group → software → technique
    for rel in rels_by_src.get(group_id, []):
        if rel.get("relationship_type") != "uses":
            continue
        sw_id = rel.get("target_ref", "")
        if sw_id not in idx["software"]:
            continue
        sw_name = idx["software"][sw_id].get("name", sw_id)
        for sw_rel in rels_by_src.get(sw_id, []):
            if sw_rel.get("relationship_type") != "uses":
                continue
            tgt = sw_rel.get("target_ref", "")
            if tgt in techniques:
                add_result(techniques[tgt], sw_rel, "software", sw_name)

    # PATH 3 — Via Campaign: campaign → group + campaign → technique
    for rel in rels_by_tgt.get(group_id, []):
        if rel.get("relationship_type") != "attributed-to":
            continue
        camp_id = rel.get("source_ref", "")
        if camp_id not in idx["campaigns"]:
            continue
        camp_name = idx["campaigns"][camp_id].get("name", camp_id)
        for camp_rel in rels_by_src.get(camp_id, []):
            if camp_rel.get("relationship_type") != "uses":
                continue
            tgt = camp_rel.get("target_ref", "")
            if tgt in techniques:
                add_result(techniques[tgt], camp_rel, "campaign", camp_name)

    return results


# ─────────────────────────────────────────────────────────────────────────────
# EVIDENCE EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def extract_evidence(procedure_text: str) -> dict:
    """
    Extracts typed indicators from procedure text.
    Returns dict with keys: cmd, reg, cve, ports, paths, software, event_ids
    """
    text = procedure_text

    evidence = {
        "cmd":       [],
        "reg":       [],
        "cve":       [],
        "ports":     [],
        "paths":     [],
        "software":  [],
        "event_ids": [],
    }

    # CVE IDs
    for m in re.findall(r'CVE-\d{4}-\d{4,7}', text):
        if m not in evidence["cve"]:
            evidence["cve"].append(m)

    # Registry paths
    for m in re.findall(r'HK(?:LM|CU|CR|U|CC)\\[^\s\n"\'`\]]{4,180}', text):
        m = m.rstrip(".,;)")
        if m not in evidence["reg"]:
            evidence["reg"].append(m)

    # Windows / UNC paths
    for m in re.findall(r'[A-Za-z]:\\[^\s\n"\'`\]<>]{4,180}', text):
        m = m.rstrip(".,;)")
        if m not in evidence["paths"]:
            evidence["paths"].append(m)
    for m in re.findall(r'\\\\[^\s\n"\'`\]<>]{4,120}', text):
        m = m.rstrip(".,;)")
        if m not in evidence["paths"]:
            evidence["paths"].append(m)

    # Unix paths
    for m in re.findall(r'/(?:etc|var|tmp|usr|home|opt|bin|sbin|proc|netscaler)/[^\s\n"\'`\]<>]{2,120}', text):
        m = m.rstrip(".,;)")
        if m not in evidence["paths"]:
            evidence["paths"].append(m)

    # Network ports (TCP/UDP nnnn)
    for m in re.findall(r'\b(?:TCP|UDP)\s+(?:port\s+)?(\d{1,5})\b', text, re.IGNORECASE):
        label = f"TCP/{m}"
        if label not in evidence["ports"]:
            evidence["ports"].append(label)
    for m in re.findall(r'\bport[s]?\s+(\d{1,5})\b', text, re.IGNORECASE):
        label = f"TCP/{m}"
        if label not in evidence["ports"]:
            evidence["ports"].append(label)

    # Windows Event IDs
    for m in re.findall(r'\b(?:Event\s+ID|EID|EventID|event\s+id)\s*[:#]?\s*(\d{3,5})\b',
                        text, re.IGNORECASE):
        label = f"Windows Event ID {m}"
        if label not in evidence["event_ids"]:
            evidence["event_ids"].append(label)

    # Commands from backtick-wrapped strings
    for m in re.findall(r'`([^`]{4,300})`', text):
        m = m.strip()
        if any(m in existing for existing in evidence["cmd"]):
            continue
        # Is it a command-like string?
        if re.search(r'(?:\.[a-zA-Z]{2,4}|\\|/|-\w|>\s*\w)', m):
            if m not in evidence["cmd"]:
                evidence["cmd"].append(m)

    # Commands from <code> tags (MITRE website uses these)
    for m in re.findall(r'<code>([^<]{4,300})</code>', text):
        m = m.strip()
        if re.search(r'(?:\.[a-zA-Z]{2,4}|\\|/|-\w)', m):
            if m not in evidence["cmd"]:
                evidence["cmd"].append(m)

    # Named command patterns
    cmd_patterns = [
        r'net\s+(?:user|group|localgroup|use|view|share|start|stop)\s+\S[^\n]{0,120}',
        r'reg\s+(?:add|delete|query|export)\s+\S[^\n]{0,120}',
        r'schtasks\s+/\S[^\n]{0,120}',
        r'powershell(?:\.exe)?\s+[-/][^\n]{0,180}',
        r'cmd(?:\.exe)?\s+/[cCkK]\s+[^\n]{0,120}',
        r'wmic\s+\S[^\n]{0,100}',
        r'certutil\s+[-/][^\n]{0,100}',
        r'bitsadmin\s+/[^\n]{0,100}',
        r'mshta(?:\.exe)?\s+\S[^\n]{0,100}',
        r'wscript(?:\.exe)?\s+\S[^\n]{0,100}',
        r'cscript(?:\.exe)?\s+\S[^\n]{0,100}',
        r'rundll32(?:\.exe)?\s+\S[^\n]{0,100}',
        r'sc\s+(?:create|start|stop|delete|config|query)\s+\S[^\n]{0,100}',
        r'vssadmin\s+\S[^\n]{0,80}',
        r'nltest\s+/[^\n]{0,80}',
        r'whoami\s*/[^\n]{0,60}',
        r'ipconfig\s*/[^\n]{0,60}',
        r'ssh(?:\.exe)?\s+[-\w][^\n]{0,100}',
        r'curl\s+[-\w][^\n]{0,150}',
        r'wget\s+[-\w][^\n]{0,150}',
    ]
    for pattern in cmd_patterns:
        for m in re.findall(pattern, text, re.IGNORECASE):
            m = m.strip()
            if m and m not in evidence["cmd"]:
                evidence["cmd"].append(m)

    # Software/tool names from [Name] MITRE markdown links
    for m in re.findall(r'\[([A-Za-z0-9][A-Za-z0-9\s\.\-_]{1,40})\]\(https://attack\.mitre\.org/software/', text):
        if m not in evidence["software"]:
            evidence["software"].append(m)

    # Remove empties
    return {k: v[:10] for k, v in evidence.items() if v}  # cap at 10 per category


# ─────────────────────────────────────────────────────────────────────────────
# INVOCATION EXTRACTOR (from v4 spreadsheet — same logic)
# ─────────────────────────────────────────────────────────────────────────────

def extract_procedure_invocations(procedure_text: str,
                                   indicator_type: str,
                                   indicator_value: str) -> list:
    if not procedure_text:
        return []

    results = []

    # Pattern 1 — Backtick-wrapped (highest confidence)
    for m in re.findall(r'`([^`]{2,500})`', procedure_text):
        results.append(("backtick", m.strip()))

    # Pattern 2 — Double-quoted with executable signals
    for m in re.findall(r'"([^"]{4,120})"', procedure_text):
        if re.search(r'\\|/|\.[a-zA-Z]{2,4}\b|\s[-/]\w', m):
            results.append(("quoted", m.strip()))

    # Pattern 3 — tool.exe invocations
    for m in re.findall(
            r'\b([A-Za-z0-9_\-]+\.(?:exe|ps1|bat|vbs|sh|py|pl|dll|cmd)'
            r'(?:\s+[^\.\n]{0,80})?)\b',
            procedure_text, re.IGNORECASE):
        results.append(("exe", m.strip()))

    # Pattern 4 — Named CLI command prefixes
    cmd_patterns = [
        r'net\s+(?:user|group|localgroup|use|view|share|session|start|stop)\s+\S[^\n]{0,100}',
        r'reg\s+(?:add|delete|query|export|import)\s+\S[^\n]{0,100}',
        r'schtasks\s+/[A-Za-z][^\n]{0,120}',
        r'powershell(?:\.exe)?\s+[-/][^\n]{0,150}',
        r'cmd(?:\.exe)?\s+/[cCkK]\s+[^\n]{0,120}',
        r'wmic\s+\S[^\n]{0,100}',
        r'certutil\s+[-/][^\n]{0,100}',
        r'bitsadmin\s+/[^\n]{0,100}',
        r'mshta(?:\.exe)?\s+\S[^\n]{0,100}',
        r'wscript(?:\.exe)?\s+\S[^\n]{0,100}',
        r'rundll32(?:\.exe)?\s+\S[^\n]{0,100}',
        r'sc\s+(?:create|start|stop|delete|config|query)\s+\S[^\n]{0,100}',
        r'vssadmin\s+\S[^\n]{0,80}',
        r'nltest\s+/[^\n]{0,100}',
        r'ssh(?:\.exe)?\s+[-\w][^\n]{0,100}',
        r'curl\s+[-\w][^\n]{0,150}',
        r'wget\s+[-\w][^\n]{0,150}',
    ]
    for pattern in cmd_patterns:
        for m in re.findall(pattern, procedure_text, re.IGNORECASE):
            results.append(("cmd_pattern", m.strip()))

    # Pattern 5 — CVE with context
    if indicator_type == "cve":
        for m in re.finditer(r'(CVE-\d{4}-\d{4,7}[^\.\n]{0,200})', procedure_text):
            results.append(("cve", m.group(1).strip()))

    # Pattern 6 — Registry paths
    if indicator_type == "reg":
        for m in re.findall(r'HK(?:LM|CU|CR|U|CC)\\[^\s\n"\'`]{6,200}', procedure_text):
            results.append(("reg", m.strip()))

    # Pattern 7 — File paths
    if indicator_type == "paths":
        for m in re.findall(r'[A-Za-z]:\\[^\s\n"\'`]{4,200}', procedure_text):
            results.append(("path", m.strip()))
        for m in re.findall(r'\\\\[^\s\n"\'`]{4,200}', procedure_text):
            results.append(("path", m.strip()))
        for m in re.findall(r'/(?:etc|var|tmp|usr|home|opt|bin|sbin|proc)/[^\s\n"\'`]{2,150}', procedure_text):
            results.append(("path", m.strip()))

    # Relevance filter
    iv_lower  = indicator_value.lower()
    iv_tokens = set(t for t in re.split(r'\W+', iv_lower) if len(t) >= 4)

    priority_order = {"backtick": 0, "cmd_pattern": 1, "exe": 2,
                      "quoted": 3, "cve": 4, "reg": 5, "path": 6}

    def is_relevant(match_str):
        ms = match_str.lower()
        if iv_lower in ms or ms in iv_lower:
            return True
        ms_tokens = set(t for t in re.split(r'\W+', ms) if len(t) >= 4)
        if iv_tokens & ms_tokens:
            return True
        # Substring token overlap (e.g. 'exchange' in 'exchangeservice')
        if any(it in mt or mt in it for it in iv_tokens for mt in ms_tokens):
            return True
        if indicator_type == "cmd" and iv_lower == ms:
            return True
        return False

    filtered = [(pt, m) for pt, m in results
                if is_relevant(m) and 4 <= len(m) <= 300]
    filtered.sort(key=lambda x: priority_order.get(x[0], 9))

    seen = set()
    deduped = []
    for _, m in filtered:
        key = m.lower()
        if key not in seen:
            seen.add(key)
            deduped.append(m)

    return deduped[:5]


DETECTION_CONTEXT = {
    "cmd":       "Process Creation — Sysmon EID 1 / Windows Security EID 4688",
    "reg":       "Registry modification — Sysmon EID 12/13/14 / Windows Security EID 4657",
    "cve":       "Exploit telemetry — check CISA KEV for active exploitation; review NVD for PoC",
    "ports":     "Network traffic — firewall/proxy logs, Zeek conn.log, Sysmon EID 3",
    "paths":     "File creation — Sysmon EID 11 (FileCreate) / EDR file telemetry",
    "software":  "Process name / image load — Sysmon EID 1, EID 7; check GitHub for CLI usage",
    "event_ids": "This IS a Windows Event ID — ensure log channel is enabled and ingested",
    "none":      "No extractable indicators — review procedure text manually",
}

_RE_MD_LINK = re.compile(r"\[([^\]]+)\]\((https?://[^\)]+)\)")
_RE_CITATION = re.compile(r"\(Citation:[^\)]*\)")

def _md_link_to_id(m):
    label = m.group(1)
    url = m.group(2).rstrip("/")
    identifier = url.rsplit("/", 1)[-1]
    return f"{label} ({identifier})"

def clean_text(text):
    """Convert markdown links to 'Name (ID)', remove citations, collapse whitespace."""
    if not text:
        return text
    text = _RE_MD_LINK.sub(_md_link_to_id, text)
    text = _RE_CITATION.sub("", text)
    text = re.sub(r"  +", " ", text).strip()
    return text

def build_invocations_and_detection(procedure_text, indicator_type, indicator_value):
    """Return (invocations_str, detection_str, had_invocations)."""
    invocations = extract_procedure_invocations(procedure_text, indicator_type, indicator_value)
    had_inv = len(invocations) > 0

    if invocations:
        inv_str = "\n".join(f"\u2022 {inv}" for inv in invocations)
    elif indicator_type == "none":
        inv_str = ""
    else:
        inv_str = "No specific invocation documented in MITRE procedure text for this indicator."

    det_str = DETECTION_CONTEXT.get(indicator_type, DETECTION_CONTEXT["none"])
    return inv_str, det_str, had_inv


# ─────────────────────────────────────────────────────────────────────────────
# ATOMISE ROWS
# ─────────────────────────────────────────────────────────────────────────────

ITYPE_SOURCE = {
    "cmd":       "Website",
    "reg":       "Website",
    "cve":       "Website",
    "ports":     "Website",
    "paths":     "Website",
    "software":  "GitHub | Website",
    "event_ids": "Website",
    "none":      "Website",
}

def get_group_meta(group_obj: dict) -> dict:
    gid   = ""
    gname = group_obj.get("name", "Unknown")
    gdesc = group_obj.get("description", "")
    for ref in group_obj.get("external_references", []):
        if ref.get("source_name") == "mitre-attack":
            gid = ref.get("external_id", "")
            break
    nav_url = f"https://attack.mitre.org/groups/{gid}/{gid}-enterprise-layer.json" if gid else ""
    stix_url = ""
    for ref in group_obj.get("external_references", []):
        if ref.get("source_name") == "mitre-attack" and ref.get("url"):
            stix_url = ref["url"]
            break
    return {"id": gid, "name": gname, "description": gdesc,
            "nav_url": nav_url, "stix_url": stix_url}


def atomise(group_meta: dict, technique_rows: list) -> list:
    """
    Returns list of flat dicts, one per atomic indicator.
    """
    out = []
    seen = set()

    for row in technique_rows:
        proc      = row["procedure_example"]
        proc_display = clean_text(proc)
        tid       = row["technique_id"]
        tname     = clean_text(row["technique_name"])
        tactic    = row["tactic"]
        src_path  = row["source_path"]
        src_name  = row["source_name"]
        ext_refs  = row.get("external_references", [])

        evidence = extract_evidence(proc)

        # Build reference URL — first URL in procedure text or ATT&CK URL
        ref_match = re.search(r'https?://\S{10,}', proc)
        if ref_match:
            ref_url = ref_match.group(0).rstrip(".,;)")
        else:
            tid_slash = tid.replace(".", "/")
            ref_url = f"https://attack.mitre.org/techniques/{tid_slash}/"

        nav_url  = group_meta["nav_url"]
        gname    = clean_text(group_meta["name"])

        if not evidence:
            dedup_key = (gname, tid, "(none)")
            if dedup_key not in seen:
                seen.add(dedup_key)
                out.append({
                    "evidential_element": "(no extractable indicators)",
                    "group":    gname,
                    "group_id": group_meta["id"],
                    "procedure_example": proc_display,
                    "technique_id": tid,
                    "technique_name": tname,
                    "tactic": tactic,
                    "indicator_type": "none",
                    "invocations": "",
                    "detection_guidance": DETECTION_CONTEXT["none"],
                    "ref_url":  ref_url,
                    "nav_url":  nav_url,
                    "source_type": "Website",
                    "source_path": src_path,
                    "source_name": src_name,
                    "_raw_procedure": proc,
                    "_ext_refs": ext_refs,
                })
            continue

        for itype, indicators in evidence.items():
            for indicator in indicators:
                dedup_key = (gname, tid, indicator.lower())
                if dedup_key in seen:
                    continue
                seen.add(dedup_key)

                inv_str, det_str, had_inv = build_invocations_and_detection(proc, itype, indicator)
                out.append({
                    "evidential_element": indicator,
                    "group":    gname,
                    "group_id": group_meta["id"],
                    "procedure_example": proc_display,
                    "technique_id": tid,
                    "technique_name": tname,
                    "tactic": tactic,
                    "indicator_type": itype,
                    "invocations": inv_str,
                    "detection_guidance": det_str,
                    "ref_url":  ref_url,
                    "nav_url":  nav_url,
                    "source_type": ITYPE_SOURCE.get(itype, "Website"),
                    "source_path": src_path,
                    "source_name": src_name,
                    "_raw_procedure": proc,
                    "_ext_refs": ext_refs,
                })

    return out


# ─────────────────────────────────────────────────────────────────────────────
# STYLING HELPERS
# ─────────────────────────────────────────────────────────────────────────────

C_NAVY   = "0D1B2A"
C_CYAN   = "0EA5E9"
C_WHITE  = "E0F2FE"
C_ORANGE = "F97316"
C_GREEN  = "22C55E"
C_YELLOW = "FACC15"
C_SLATE  = "CBD5E1"
C_TEAL   = "2DD4BF"
C_PURPLE = "A78BFA"
C_GREY   = "475569"

# Colour palette for nation-state clusters (keyword → (bg, fg))
NATION_COLOURS = {
    "iran":       ("0F2035", "38BDF8"),
    "russia":     ("200808", "F87171"),
    "china":      ("0A1A08", "4ADE80"),
    "north korea":("1A0A30", "C084FC"),
    "lazarus":    ("1A0A30", "C084FC"),
    "apt38":      ("1A0A30", "C084FC"),
    "india":      ("1A1200", "FCD34D"),
    "vietnam":    ("001A08", "6EE7B7"),
    "pakistan":   ("001215", "67E8F9"),
    "turkey":     ("1A0A00", "FB923C"),
    "lebanon":    ("001A00", "86EFAC"),
    "israel":     ("001520", "7DD3FC"),
    "ukraine":    ("00101A", "93C5FD"),
    "belarus":    ("10001A", "DDD6FE"),
    "syria":      ("0A0800", "FDE68A"),
}
DEFAULT_COLOUR = ("0F1C2E", "94A3B8")

def get_group_colour(group_name: str, group_desc: str) -> tuple:
    combined = (group_name + " " + group_desc).lower()
    for kw, colours in NATION_COLOURS.items():
        if kw in combined:
            return colours
    return DEFAULT_COLOUR

def thin_border(color="1E3A5F"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hcell(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Calibri", bold=True, color=C_WHITE, size=12)
    c.fill = PatternFill("solid", fgColor=C_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()

def dcell(ws, row, col, val, bg=None, fg=C_WHITE, bold=False, url=None, mono=False):
    c = ws.cell(row=row, column=col, value=val)
    bg = bg or DEFAULT_COLOUR[0]
    c.font = Font(name="Courier New" if mono else "Calibri",
                  color=fg, size=10, bold=bold,
                  underline="single" if url else None)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    c.border = thin_border()
    if url:
        c.hyperlink = url
    return c


# ─────────────────────────────────────────────────────────────────────────────
# WORKBOOK BUILDER
# ─────────────────────────────────────────────────────────────────────────────

HEADERS = [
    "Evidential Element\n(Atomic Indicator / Command / Artefact)",
    "Threat Group",
    "Group ID",
    "Procedure Example\n(MITRE ATT\u0026CK — verbatim)",
    "Technique ID",
    "Technique Name",
    "Tactic",
    "MITRE Invocations\n(Procedure Text Extractions)",
    "Detection Guidance",
    "Reference URL",
    "Navigation Layer URL\n(ATT\u0026CK Navigator JSON)",
    "Source Type",
    "Attribution Path",
]

COL_WIDTHS = [48, 22, 10, 55, 14, 28, 20, 48, 48, 45, 38, 16, 14]

def write_data_sheet(ws, all_rows: list, sheet_title: str, subtitle: str,
                     group_colour_map: dict):
    ws.sheet_view.showGridLines = False

    # Title
    num_cols = len(HEADERS)
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value = sheet_title
    t.font = Font(name="Calibri", bold=True, color=C_CYAN, size=16)
    t.fill = PatternFill("solid", fgColor=C_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{last_col}2")
    s = ws["A2"]
    s.value = subtitle
    s.font = Font(name="Calibri", italic=True, color="7FB3D3", size=10)
    s.fill = PatternFill("solid", fgColor=C_NAVY)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    for ci, h in enumerate(HEADERS, 1):
        hcell(ws, 3, ci, h)
    ws.row_dimensions[3].height = 40

    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(all_rows, 4):
        gname = row["group"]
        bg, gfg = group_colour_map.get(gname, DEFAULT_COLOUR)
        alt_bg = "0A1220" if ri % 2 == 0 else bg

        dcell(ws, ri,  1, row["evidential_element"],  alt_bg, C_TEAL,   bold=True, mono=True)
        dcell(ws, ri,  2, gname,                      alt_bg, gfg,      bold=True)
        dcell(ws, ri,  3, row["group_id"],             alt_bg, C_GREEN,  bold=True, mono=True)
        dcell(ws, ri,  4, row["procedure_example"],    alt_bg, C_SLATE)
        dcell(ws, ri,  5, row["technique_id"],         alt_bg, C_GREEN,  bold=True, mono=True)
        dcell(ws, ri,  6, row["technique_name"],       alt_bg, C_WHITE)
        dcell(ws, ri,  7, row["tactic"],               alt_bg, C_YELLOW)
        dcell(ws, ri,  8, row["invocations"],          alt_bg, C_SLATE,  mono=True)
        dcell(ws, ri,  9, row["detection_guidance"],   alt_bg, C_SLATE,  mono=True)
        dcell(ws, ri, 10, row["ref_url"],              alt_bg, C_CYAN,   url=row["ref_url"])
        dcell(ws, ri, 11, row["nav_url"],              alt_bg, C_PURPLE, url=row["nav_url"] or None)
        dcell(ws, ri, 12, row["source_type"],          alt_bg, C_ORANGE)
        dcell(ws, ri, 13, row["source_path"],          alt_bg, C_GREY)
        ws.row_dimensions[ri].height = 70

    ws.freeze_panes = "A4"
    last_data = 3 + len(all_rows)
    ws.auto_filter.ref = f"A3:{last_col}{last_data}"


def write_group_summary(wb, all_rows: list, group_meta_map: dict,
                        group_colour_map: dict):
    ws = wb.create_sheet("Group Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "MITRE ATT\u0026CK — All Groups Summary"
    t.font = Font(name="Calibri", bold=True, color=C_CYAN, size=16)
    t.fill = PatternFill("solid", fgColor=C_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    H = ["Group Name","Group ID","Technique Count","Indicator Count",
         "Invocation Coverage %","Top Tactic","Tactic Coverage","Description (excerpt)"]
    for ci, h in enumerate(H, 1):
        hcell(ws, 2, ci, h)
    ws.row_dimensions[2].height = 32

    widths = [28, 10, 16, 16, 20, 22, 55, 60]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Aggregate per group
    group_stats = defaultdict(lambda: {
        "techs": set(), "indicators": 0, "invocations": 0, "tactics": Counter()
    })
    for row in all_rows:
        g = row["group"]
        group_stats[g]["techs"].add(row["technique_id"])
        group_stats[g]["indicators"] += 1
        if row["invocations"] and not row["invocations"].startswith("No specific"):
            group_stats[g]["invocations"] += 1
        if row["tactic"]:
            group_stats[g]["tactics"][row["tactic"]] += 1

    all_groups = sorted(group_stats.keys())
    for ri, gname in enumerate(all_groups, 3):
        st   = group_stats[gname]
        meta = group_meta_map.get(gname, {})
        bg, gfg = group_colour_map.get(gname, DEFAULT_COLOUR)
        alt_bg = "0A1220" if ri % 2 == 0 else bg

        tech_count = len(st["techs"])
        ind_count  = st["indicators"]
        inv_pct    = f"{100*st['invocations']//ind_count}%" if ind_count else "0%"
        top_tactic = st["tactics"].most_common(1)[0][0] if st["tactics"] else ""
        tactic_cov = ", ".join(t for t, _ in st["tactics"].most_common(5))
        desc_raw   = clean_text(meta.get("description", ""))
        desc       = (desc_raw[:180] + "…") if desc_raw else ""

        dcell(ws, ri, 1, gname,       alt_bg, gfg,     bold=True)
        dcell(ws, ri, 2, meta.get("id",""),  alt_bg, C_GREEN, bold=True, mono=True)
        dcell(ws, ri, 3, tech_count,  alt_bg, C_WHITE)
        dcell(ws, ri, 4, ind_count,   alt_bg, C_WHITE)
        dcell(ws, ri, 5, inv_pct,     alt_bg, C_CYAN)
        dcell(ws, ri, 6, top_tactic,  alt_bg, C_YELLOW)
        dcell(ws, ri, 7, tactic_cov,  alt_bg, C_SLATE)
        dcell(ws, ri, 8, desc,        alt_bg, C_SLATE)
        ws.row_dimensions[ri].height = 40

    ws.freeze_panes = "A3"


def write_tactic_pivot(wb, all_rows: list):
    ws = wb.create_sheet("Tactic Pivot")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "ATT\u0026CK Tactic Distribution — All Groups"
    t.font = Font(name="Calibri", bold=True, color=C_CYAN, size=16)
    t.fill = PatternFill("solid", fgColor=C_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    H = ["Tactic","Indicator Count","% of Total","Invocations Found","Example Technique IDs"]
    for ci, h in enumerate(H, 1):
        hcell(ws, 2, ci, h)

    tactic_map   = defaultdict(list)
    inv_hits     = defaultdict(int)
    for row in all_rows:
        tac = row["tactic"] or "(none)"
        tactic_map[tac].append(row["technique_id"])
        if row["invocations"] and not row["invocations"].startswith("No specific"):
            inv_hits[tac] += 1

    for ci, w in enumerate([28, 16, 14, 18, 55], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    last_data = 2 + len(tactic_map)
    for ri, (tac, techs) in enumerate(
            sorted(tactic_map.items(), key=lambda x: -len(x[1])), 3):
        bg = "0F1C2E" if ri % 2 == 0 else "0A1220"
        dcell(ws, ri, 1, tac,  bg, C_YELLOW, bold=True)
        dcell(ws, ri, 2, len(techs), bg, C_WHITE)
        c3 = ws.cell(row=ri, column=3)
        c3.value = f"=B{ri}/SUM(B3:B{last_data})"
        c3.number_format = "0.0%"
        c3.font = Font(name="Calibri", color=C_GREEN, size=10)
        c3.fill = PatternFill("solid", fgColor=bg)
        c3.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        c3.border = thin_border()
        dcell(ws, ri, 4, inv_hits.get(tac, 0), bg, C_CYAN)
        dcell(ws, ri, 5, " | ".join(sorted(set(techs))[:6]), bg, C_GREEN, mono=True)
        ws.row_dimensions[ri].height = 22

    ws.freeze_panes = "A3"


def write_nation_pivot(wb, all_rows: list, group_meta_map: dict):
    """Summary sheet grouping groups by inferred nation-state."""
    ws = wb.create_sheet("Nation-State Pivot")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "Nation-State Attribution Pivot"
    t.font = Font(name="Calibri", bold=True, color=C_CYAN, size=16)
    t.fill = PatternFill("solid", fgColor=C_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    H = ["Nation/Cluster","Group Count","Indicator Count","Top Tactic","Groups"]
    for ci, h in enumerate(H, 1):
        hcell(ws, 2, ci, h)

    # Infer nation per group
    group_nation = {}
    for gname, meta in group_meta_map.items():
        combined = (gname + " " + meta.get("description","")).lower()
        assigned = "Unattributed / Unknown"
        for kw in ["iran","russia","china","north korea","vietnam",
                   "india","pakistan","turkey","lebanon","israel",
                   "ukraine","belarus","syria"]:
            if kw in combined:
                assigned = kw.title()
                break
        if "lazarus" in combined or "apt38" in combined or "hidden cobra" in combined:
            assigned = "North Korea"
        group_nation[gname] = assigned

    nation_stats = defaultdict(lambda: {
        "groups": set(), "indicators": 0, "tactics": Counter()
    })
    for row in all_rows:
        nat = group_nation.get(row["group"], "Unattributed / Unknown")
        nation_stats[nat]["groups"].add(row["group"])
        nation_stats[nat]["indicators"] += 1
        if row["tactic"]:
            nation_stats[nat]["tactics"][row["tactic"]] += 1

    widths = [24, 14, 16, 24, 80]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, (nat, st) in enumerate(
            sorted(nation_stats.items(), key=lambda x: -x[1]["indicators"]), 3):
        bg  = "0F1C2E" if ri % 2 == 0 else "0A1220"
        top_tactic = st["tactics"].most_common(1)[0][0] if st["tactics"] else ""
        groups_str = ", ".join(sorted(st["groups"]))
        dcell(ws, ri, 1, nat,                   bg, C_CYAN, bold=True)
        dcell(ws, ri, 2, len(st["groups"]),     bg, C_WHITE)
        dcell(ws, ri, 3, st["indicators"],       bg, C_WHITE)
        dcell(ws, ri, 4, top_tactic,             bg, C_YELLOW)
        dcell(ws, ri, 5, groups_str,             bg, C_SLATE)
        ws.row_dimensions[ri].height = 28

    ws.freeze_panes = "A3"


def write_reference_detail(wb, ref_results: list, group_colour_map: dict):
    """Write a 'Reference Detail' sheet with one row per citation."""
    ws = wb.create_sheet("Reference Detail")
    ws.sheet_view.showGridLines = False

    num_with_content = sum(1 for r in ref_results if r.get("extracted_content"))
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = (f"Citation Reference Detail  |  {len(ref_results)} Citations  |  "
               f"{num_with_content} with Extracted Content")
    t.font = Font(name="Calibri", bold=True, color=C_CYAN, size=16)
    t.fill = PatternFill("solid", fgColor=C_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    H = ["Threat Group", "Technique ID", "Technique Name", "Citation Name",
         "Source URL", "Source Description", "Extracted Content", "Status"]
    for ci, h in enumerate(H, 1):
        hcell(ws, 2, ci, h)
    ws.row_dimensions[2].height = 32

    widths = [22, 14, 28, 30, 55, 50, 80, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, ref in enumerate(ref_results, 3):
        gname = ref.get("group", "")
        bg, gfg = group_colour_map.get(gname, DEFAULT_COLOUR)
        alt_bg = "0A1220" if ri % 2 == 0 else bg

        dcell(ws, ri, 1, gname,                         alt_bg, gfg, bold=True)
        dcell(ws, ri, 2, ref.get("technique_id", ""),    alt_bg, C_GREEN, bold=True, mono=True)
        dcell(ws, ri, 3, ref.get("technique_name", ""),  alt_bg, C_WHITE)
        dcell(ws, ri, 4, ref.get("citation_name", ""),   alt_bg, C_CYAN)
        dcell(ws, ri, 5, ref.get("url", ""),             alt_bg, C_CYAN,
              url=ref.get("url") or None)
        dcell(ws, ri, 6, ref.get("description", ""),     alt_bg, C_SLATE)
        dcell(ws, ri, 7, ref.get("extracted_content",""),alt_bg, C_SLATE, mono=True)
        dcell(ws, ri, 8, ref.get("status", ""),          alt_bg, C_ORANGE)
        ws.row_dimensions[ri].height = 90

    ws.freeze_panes = "A3"
    last = 2 + len(ref_results)
    if ref_results:
        ws.auto_filter.ref = f"A2:H{last}"


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  MITRE ATT&CK All-Groups Evidence Report Generator")
    print(f"  Framework: {FRAMEWORK}")
    print(f"  Nation filter: {NATION_FILTER or 'None (all groups)'}")
    print("=" * 65)

    # 1. Download / load STIX
    stix = download_stix(FRAMEWORK, FORCE_FETCH)

    # 2. Build indexes
    idx = build_indexes(stix)

    # 3. Filter groups
    groups = idx["groups"]
    if NATION_FILTER:
        filtered = []
        for g in groups:
            combined = (g.get("name","") + " " + g.get("description","")).lower()
            if any(kw.lower() in combined for kw in NATION_FILTER):
                filtered.append(g)
        print(f"[+] Nation filter applied: {len(filtered)} of {len(groups)} groups match {NATION_FILTER}")
        groups = filtered
    else:
        print(f"[+] Processing all {len(groups)} groups")

    # 4. Process each group
    all_rows       = []
    group_meta_map = {}
    group_colour_map = {}

    for i, group_obj in enumerate(sorted(groups, key=lambda g: g.get("name",""))):
        meta = get_group_meta(group_obj)
        gname = meta["name"]
        group_meta_map[gname] = meta

        bg, fg = get_group_colour(gname, meta["description"])
        group_colour_map[gname] = (bg, fg)

        technique_rows = get_group_techniques(group_obj["id"], idx)
        group_rows     = atomise(meta, technique_rows)
        all_rows.extend(group_rows)

        if (i + 1) % 10 == 0 or (i + 1) == len(groups):
            print(f"  [{i+1:3d}/{len(groups)}] {gname:35s}  "
                  f"{len(technique_rows):3d} techniques  {len(group_rows):4d} indicators  "
                  f"total: {len(all_rows):,}")

    print(f"\n[+] Total: {len(all_rows):,} indicator rows across {len(groups)} groups")

    # Invocation coverage
    inv_count = sum(1 for r in all_rows
                    if r["invocations"] and not r["invocations"].startswith("No specific"))
    pct = 100 * inv_count // len(all_rows) if all_rows else 0
    print(f"[+] Invocation coverage: {inv_count:,}/{len(all_rows):,} rows ({pct}%) "
          f"have MITRE-documented invocations")

    # 5. Build workbook
    print("\n[+] Building workbook...")
    wb = Workbook()

    ts  = datetime.now().strftime("%Y-%m-%d %H:%M")
    nat = ", ".join(NATION_FILTER) if NATION_FILTER else "All Groups"

    # If too many rows, split into sheets by letter range
    if len(all_rows) <= MAX_ROWS_PER_SHEET:
        ws = wb.active
        ws.title = "Evidence Table"
        title    = (f"MITRE ATT\u0026CK Evidence Report  |  {FRAMEWORK}  |  "
                    f"{len(groups)} Groups  |  {len(all_rows):,} Indicators  |  {ts}")
        subtitle = (f"Filter: {nat}  |  "
                    f"Invocation Coverage: {pct}%  |  Source: MITRE ATT\u0026CK STIX via CTI repo")
        write_data_sheet(ws, all_rows, title, subtitle, group_colour_map)
        print(f"    Sheet 'Evidence Table': {len(all_rows):,} rows")
    else:
        # Split alphabetically by group name
        buckets = defaultdict(list)
        for row in all_rows:
            letter = row["group"][0].upper()
            if letter < "G":
                buckets["A-F"].append(row)
            elif letter < "L":
                buckets["G-K"].append(row)
            elif letter < "R":
                buckets["L-Q"].append(row)
            else:
                buckets["R-Z"].append(row)

        first = True
        for bucket_name, bucket_rows in sorted(buckets.items()):
            if first:
                ws = wb.active
                ws.title = f"Evidence {bucket_name}"
                first = False
            else:
                ws = wb.create_sheet(f"Evidence {bucket_name}")
            title = (f"MITRE ATT\u0026CK Evidence  |  Groups {bucket_name}  |  "
                     f"{len(bucket_rows):,} Indicators  |  {ts}")
            subtitle = f"Filter: {nat}  |  Source: MITRE ATT\u0026CK STIX"
            write_data_sheet(ws, bucket_rows, title, subtitle, group_colour_map)
            print(f"    Sheet 'Evidence {bucket_name}': {len(bucket_rows):,} rows")

    write_group_summary(wb, all_rows, group_meta_map, group_colour_map)
    print(f"    Sheet 'Group Summary': {len(groups)} groups")

    write_tactic_pivot(wb, all_rows)
    print(f"    Sheet 'Tactic Pivot'")

    write_nation_pivot(wb, all_rows, group_meta_map)
    print(f"    Sheet 'Nation-State Pivot'")

    # 5b. Reference collection (if enabled)
    if COLLECT_REFERENCES:
        # Add parent path so we can import reference_collector
        _parent = str(Path(__file__).resolve().parent.parent)
        if _parent not in sys.path:
            sys.path.insert(0, _parent)
        from reference_collector import (
            resolve_citations, collect_reference_content
        )

        print("\n[+] Collecting citation references...")

        # Build stix_ref_map and collect unique procedures
        seen_ref_procs = set()
        ref_results = []
        total_citations = 0
        total_fetched = 0

        for row in all_rows:
            proc_key = (row["group"], row["technique_id"],
                        hash(row.get("_raw_procedure", row["procedure_example"])))
            if proc_key in seen_ref_procs:
                continue
            seen_ref_procs.add(proc_key)

            ext_refs = row.get("_ext_refs", [])
            raw_proc = row.get("_raw_procedure", row["procedure_example"])

            citations = resolve_citations(raw_proc, ext_refs)
            if not citations:
                continue
            total_citations += len(citations)

            indicators = []
            if row["evidential_element"] != "(no extractable indicators)":
                indicators = [row["evidential_element"]]

            fetched = collect_reference_content(
                citations,
                group_name=row["group"],
                technique_name=row["technique_name"],
                technique_id=row["technique_id"],
                indicators=indicators,
                verbose=True,
            )
            for r in fetched:
                r["group"] = row["group"]
                r["group_id"] = row["group_id"]
                r["technique_id"] = row["technique_id"]
                r["technique_name"] = row["technique_name"]
                r["tactic"] = row["tactic"]
                ref_results.append(r)
                if r["status"] in ("fetched", "cached"):
                    total_fetched += 1

        print(f"[+] Citations resolved: {total_citations}, "
              f"URLs fetched/cached: {total_fetched}, "
              f"with content: {sum(1 for r in ref_results if r.get('extracted_content'))}")

        write_reference_detail(wb, ref_results, group_colour_map)
        print(f"    Sheet 'Reference Detail': {len(ref_results)} citations")

    # 6. Save
    stamp    = datetime.now().strftime("%Y%m%d_%H%M%S")
    nat_slug = "_".join(NATION_FILTER) if NATION_FILTER else "all_groups"
    filename = f"Evidence_{FRAMEWORK}_{nat_slug}_{stamp}.xlsx"
    out_path = Path(OUTPUT_DIR) / filename
    wb.save(out_path)

    size_mb = out_path.stat().st_size / 1024 / 1024
    print(f"\n[+] Saved: {out_path}  ({size_mb:.1f} MB)")
    print(f"[+] Completed at {datetime.now().strftime('%H:%M:%S')}")


if __name__ == "__main__":
    import argparse
    _parser = argparse.ArgumentParser(description="Generate MITRE ATT&CK evidence report")
    _parser.add_argument("-R", "--references", action="store_true",
                         help="Fetch citation URLs and extract pertinent content")
    _parser.add_argument("-F", "--force-fetch", action="store_true",
                         help="Force re-download of STIX data")
    _parser.add_argument("--framework", default=None,
                         help=f"ATT&CK framework (default: {FRAMEWORK})")
    _parser.add_argument("--nation", nargs="*", default=None,
                         help="Filter groups by nation keyword (e.g. iran china)")
    _args = _parser.parse_args()

    if _args.references:
        COLLECT_REFERENCES = True
    if _args.force_fetch:
        FORCE_FETCH = True
    if _args.framework:
        FRAMEWORK = _args.framework
    if _args.nation is not None:
        NATION_FILTER = _args.nation

    main()
