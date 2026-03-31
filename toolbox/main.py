#!/usr/bin/env python3 -tt
import os
import json
import random
import re
import requests
import subprocess
import sys
import time
from collections import Counter
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, List, Tuple, Set

from mitreattack.stix20 import MitreAttackData


class _ProgressBar:
    """Inline progress bar that overwrites a single line during progress,
    then prints a permanent 100% line when done."""

    def __init__(self, label=""):
        self._start = None
        self._label = label

    def update(self, current, total, detail="", bar_width=30):
        """Overwrite the current line with progress."""
        if total == 0:
            return
        try:
            tw = os.get_terminal_size().columns
        except OSError:
            tw = 120

        pct = current / total
        filled = int(bar_width * pct)
        bar = "\033[36m" + "█" * filled + "\033[90m" + "░" * (bar_width - filled) + "\033[0m"

        now = time.time()
        if self._start is None or current <= 1:
            self._start = now
        secs = now - self._start
        if current > 0 and secs > 0:
            eta = (secs / current) * (total - current)
            eta_str = f"{int(eta // 60)}m{int(eta % 60):02d}s" if eta >= 60 else f"{int(eta)}s"
        else:
            eta_str = "..."

        line = f"     {self._label} {bar} {current}/{total} ({pct:.0%}) ETA: {eta_str}  {detail}"
        sys.stdout.write(f"\r{line[:tw].ljust(tw)}")
        sys.stdout.flush()

    def done(self, total, detail="Complete", bar_width=30):
        """Print a permanent green 100% line and move to next line."""
        try:
            tw = os.get_terminal_size().columns
        except OSError:
            tw = 120

        bar = "\033[32m" + "█" * bar_width + "\033[0m"
        line = f"     {self._label} {bar} {total}/{total} (100%) {detail}"
        sys.stdout.write(f"\r{line[:tw].ljust(tw)}\n")
        sys.stdout.flush()
from stix2 import TAXIICollectionSource, Filter
from taxii2client.v20 import Server, Collection

import urllib3
import warnings

import pandas

from toolbox.extract import extract_indicators


def _fetch(url: str, **kwargs) -> requests.Response:
    """GET with automatic SSL-verify fallback for corporate VPN/proxy environments."""
    try:
        return requests.get(url, **kwargs)
    except requests.exceptions.SSLError:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        return requests.get(url, verify=False, **kwargs)
from toolbox.tools.write_csv import write_csv_summary
from toolbox.tools.write_csv import write_csv_techniques_mapped_to_logsources
from toolbox.output.matrix import build_matrix
from toolbox.output.query import build_queries
from toolbox.tools.keywords import match_keywords
from toolbox.tools.read_files import collect_files
from toolbox.tools.print_saw import print_saw


def get_latest_attack_version() -> str:
    """Fetch the latest MITRE ATT&CK version from STIX data."""
    try:
        # Use the TAXII server to get the latest version
        try:
            server = Server("https://cti-taxii.mitre.org/taxii/")
            api_root = server.api_roots[0]
            collections = api_root.collections
        except requests.exceptions.SSLError:
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
            server = Server("https://cti-taxii.mitre.org/taxii/", verify=False)
            api_root = server.api_roots[0]
            collections = api_root.collections

        for collection in collections:
            if "Enterprise ATT&CK" in collection.title:
                # Get collection source
                collection_source = TAXIICollectionSource(collection)
                # Get the x-mitre-collection object to find version
                filters = [Filter("type", "=", "x-mitre-collection")]
                collections_data = collection_source.query(filters)
                if collections_data:
                    version = collections_data[0].get("x_mitre_version", "Unknown")
                    return version

        # Fallback to web scraping if STIX method fails
        version_history = _fetch("https://attack.mitre.org/resources/versions/", timeout=10)
        latest_version = re.findall(
            r"<span><strong>([^<]+)",
            str(version_history.content)
            .split("<h5><strong>Current Version</strong></h5>")[1]
            .split("<h5><strong>Most Recent Versions</strong></h5>")[0],
        )[0].split("ATT&amp;CK v")[1]
        return latest_version
    except Exception as e:
        print(f"\n\n\tWarning: Could not determine latest version: {e}")
        return "Unknown"


def build_technique_datasource_map(stix_filepath: str) -> Dict[str, str]:
    """Build a mapping from technique external ID to data source strings.

    Returns dict like {"T1001": "Network Traffic: Network Connection Creation, Process: Process Creation, ..."}
    """
    with open(stix_filepath, encoding="utf-8") as f:
        data = json.load(f)

    id_to_obj = {o["id"]: o for o in data["objects"]}

    # Build data component id -> name
    dc_names = {o["id"]: o.get("name", "") for o in data["objects"] if o.get("type") == "x-mitre-data-component"}

    # Build full "DataSource: DataComponent" names by matching DC names to DS names
    ds_names = sorted(
        [o.get("name", "") for o in data["objects"] if o.get("type") == "x-mitre-data-source"],
        key=len, reverse=True,
    )
    # Manual fallback for DCs whose names don't start with their parent DS name
    dc_parent_override = {
        "Response Content": "Internet Scan", "Response Metadata": "Internet Scan",
        "Malware Content": "Malware Repository", "Malware Metadata": "Malware Repository",
        "Network Connection Creation": "Network Traffic",
        "Active DNS": "Domain Name", "Passive DNS": "Domain Name",
        "Domain Registration": "Domain Name",
        "Host Status": "Sensor Health", "Social Media": "Persona",
        "OS API Execution": "Process",
    }
    dc_full_names = {}
    for dc_id, dc_name in dc_names.items():
        if dc_name in dc_parent_override:
            dc_full_names[dc_id] = f"{dc_parent_override[dc_name]}: {dc_name}"
        else:
            for ds_name in ds_names:
                if dc_name.startswith(ds_name):
                    dc_full_names[dc_id] = f"{ds_name}: {dc_name}"
                    break
            else:
                dc_full_names[dc_id] = dc_name

    # Build detection_strategy -> data component full names (via analytics)
    ds_to_dcs: Dict[str, Set] = {}
    for ds_obj in [o for o in data["objects"] if o.get("type") == "x-mitre-detection-strategy"]:
        dcs: Set[str] = set()
        for aref in ds_obj.get("x_mitre_analytic_refs", []):
            analytic = id_to_obj.get(aref)
            if analytic:
                for lref in analytic.get("x_mitre_log_source_references", []):
                    dc_ref = lref.get("x_mitre_data_component_ref", "")
                    if dc_ref in dc_full_names:
                        dcs.add(dc_full_names[dc_ref])
        ds_to_dcs[ds_obj["id"]] = dcs

    # Build technique ext_id -> data source string from detects relationships
    tech_ext_ids = {}
    for o in data["objects"]:
        if o.get("type") == "attack-pattern":
            ext_id = o.get("external_references", [{}])[0].get("external_id", "")
            tech_ext_ids[o["id"]] = ext_id

    tech_to_ds: Dict[str, Set] = {}
    for rel in data["objects"]:
        if rel.get("type") == "relationship" and rel.get("relationship_type") == "detects":
            src = rel["source_ref"]
            tgt = rel["target_ref"]
            if src in ds_to_dcs and tgt in tech_ext_ids:
                ext_id = tech_ext_ids[tgt]
                if ext_id not in tech_to_ds:
                    tech_to_ds[ext_id] = set()
                tech_to_ds[ext_id].update(ds_to_dcs[src])

    return {tid: ", ".join(sorted(dcs)) for tid, dcs in tech_to_ds.items()}


def load_attack_data(framework: str = "enterprise", force_fetch: bool = False) -> MitreAttackData:
    """Load MITRE ATT&CK data using STIX via the mitreattack-python library.

    Re-downloads if the cached file is older than 7 days or if force_fetch is True (--fetch).
    """
    print(f"    -> Loading {framework} ATT&CK data from STIX...")

    framework_map = {
        "enterprise": "enterprise-attack",
        "mobile": "mobile-attack",
        "ics": "ics-attack"
    }

    stix_source = framework_map.get(framework.lower(), "enterprise-attack")
    stix_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "stix_data")
    os.makedirs(stix_dir, exist_ok=True)
    stix_filepath = os.path.join(stix_dir, f"{stix_source}.json")

    need_download = False
    if not os.path.exists(stix_filepath):
        need_download = True
    elif force_fetch:
        print(f"    -> --fetch flag set, forcing fresh download...")
        need_download = True
    else:
        file_age_days = (time.time() - os.path.getmtime(stix_filepath)) / 86400
        if file_age_days > 7:
            print(f"    -> STIX data is {int(file_age_days)} days old, re-downloading...")
            need_download = True
        else:
            print(f"    -> Using cached STIX data ({int(file_age_days)} day(s) old)")

    if need_download:
        stix_url = f"https://raw.githubusercontent.com/mitre/cti/master/{stix_source}/{stix_source}.json"
        print(f"    -> Downloading {stix_source} STIX data...")
        resp = _fetch(stix_url, timeout=60)
        resp.raise_for_status()
        with open(stix_filepath, "w", encoding="utf-8") as f:
            f.write(resp.text)
        print(f"    -> Saved to {stix_filepath}")

    attack_data = MitreAttackData(stix_filepath)
    return attack_data, stix_filepath


def process_technique_parallel(args: Tuple) -> List[Dict]:
    """Process a single technique in parallel."""
    technique_entry, groups, platforms, attack_data = args
    results = []

    try:
        # get_techniques_used_by_group returns {'object': ..., 'relationships': [...]}
        technique = technique_entry.get("object", technique_entry)
        relationships = technique_entry.get("relationships", [])

        technique_id = technique.get("external_references", [{}])[0].get("external_id", "")
        technique_name = technique.get("name", "")
        technique_description = technique.get("description", "")

        # Get usage description from relationship (how the group uses the technique)
        usage = ""
        if relationships:
            usage = relationships[0].get("description", "")

        # Get platforms
        tech_platforms = technique.get("x_mitre_platforms", [])

        # Check if platform matches
        if platforms != ["."] and not any(p in tech_platforms for p in platforms):
            return results

        # Get tactics
        kill_chain_phases = technique.get("kill_chain_phases", [])
        tactics = [phase.get("phase_name", "").replace("-", " ").title() for phase in kill_chain_phases]

        # Get data sources
        data_sources = []
        data_components = technique.get("x_mitre_data_sources", [])
        if isinstance(data_components, list):
            data_sources = data_components

        # Get detection info
        detection = technique.get("x_mitre_detection", "")

        result = {
            "id": technique_id,
            "name": technique_name,
            "description": technique_description,
            "usage": usage,
            "platforms": tech_platforms,
            "tactics": tactics,
            "data_sources": data_sources,
            "detection": detection,
            "stix_id": technique.get("id", "")
        }
        results.append(result)

    except Exception as e:
        print(f"    Warning: Error processing technique: {e}")

    return results


def get_group_techniques_parallel(attack_data: MitreAttackData, groups: List[str],
                                   platforms: List[str], max_workers: int = 10) -> Tuple[Dict, Dict, List]:
    """Get techniques used by groups using parallel processing."""

    group_techniques = {}
    group_info = {}
    all_techniques = []

    # Get all groups
    all_groups = attack_data.get_groups(remove_revoked_deprecated=True)

    # Filter groups if specified
    if groups != ["."]:
        filtered_groups = []
        for group in all_groups:
            group_name = group.get("name", "")
            group_aliases = group.get("aliases", [])
            # Check if group matches any provided group names
            if any(g.replace("_", " ").lower() in [group_name.lower()] + [a.lower() for a in group_aliases]
                   for g in groups):
                filtered_groups.append(group)
        all_groups = filtered_groups

    # Process each group
    for group in all_groups:
        try:
            group_name = group.get("name", "")
            group_id = group.get("external_references", [{}])[0].get("external_id", "")
            group_description = group.get("description", "")

            group_info[group_id] = {
                "name": group_name,
                "description": group_description,
                "aliases": group.get("aliases", [])
            }

            # Get techniques used by this group
            techniques = attack_data.get_techniques_used_by_group(group.get("id"))

            # Also get techniques from campaigns attributed to this group
            campaign_techniques = []
            try:
                campaigns = attack_data.get_campaigns_attributed_to_group(group.get("id"))
                if campaigns:
                    seen_technique_ids = set()
                    for t in techniques:
                        tobj = t.get("object", t)
                        tid = tobj.get("external_references", [{}])[0].get("external_id", "")
                        seen_technique_ids.add(tid)
                    for campaign_entry in campaigns:
                        campaign_obj = campaign_entry.get("object", campaign_entry)
                        campaign_name = campaign_obj.get("name", "")
                        campaign_id = campaign_obj.get("id", "")
                        try:
                            camp_techs = attack_data.get_techniques_used_by_campaign(campaign_id)
                            for ct in camp_techs:
                                ct_obj = ct.get("object", ct)
                                ct_id = ct_obj.get("external_references", [{}])[0].get("external_id", "")
                                if ct_id not in seen_technique_ids:
                                    seen_technique_ids.add(ct_id)
                                    # Prepend campaign name to usage for context
                                    ct_rels = ct.get("relationships", [])
                                    if ct_rels:
                                        orig_desc = ct_rels[0].get("description", "")
                                        ct_rels[0]["description"] = f"[Campaign: {campaign_name}] {orig_desc}"
                                    campaign_techniques.append(ct)
                        except Exception:
                            pass
            except Exception:
                pass

            all_group_techniques = list(techniques) + campaign_techniques
            group_techniques[group_id] = []

            # Process techniques in parallel
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = []
                for technique in all_group_techniques:
                    futures.append(executor.submit(process_technique_parallel,
                                                  (technique, groups, platforms, attack_data)))

                for future in as_completed(futures):
                    try:
                        results = future.result()
                        for result in results:
                            group_techniques[group_id].append(result)
                            all_techniques.append(result)
                    except Exception as e:
                        print(f"    Warning: Thread error: {e}")

        except Exception as e:
            print(f"    Warning: Error processing group {group.get('name', 'Unknown')}: {e}")
            continue

    return group_techniques, group_info, all_techniques


def replace_commas_in_group_desc(csv_line):
    return re.sub(
        r"^((?:[^,]+,){6}.*,relationship--[^,]+,\d{2} [A-Za-z]{3,20} \d{4},\d{2} [A-Za-z]{3,20} \d{4}[^\[]+[^\(]+\([^\)]+\)[^,]+), ",
        r"\1%2C ",
        csv_line,
    )


def _write_reference_sheet(xlsx_path, all_refs):
    """Append a Reference Detail sheet to an existing XLSX."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(xlsx_path)
    ws = wb.create_sheet("Reference Detail")
    ws.sheet_view.showGridLines = False

    border = Border(
        left=Side(style="thin", color="1E3A5F"),
        right=Side(style="thin", color="1E3A5F"),
        top=Side(style="thin", color="1E3A5F"),
        bottom=Side(style="thin", color="1E3A5F"),
    )
    fill_navy = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    font_header = Font(name="Calibri", size=12, bold=True, color="E0F2FE")
    font_data = Font(name="Calibri", size=10, color="CBD5E1")
    font_url = Font(name="Calibri", size=10, color="0EA5E9")
    align_wrap = Alignment(wrap_text=True, vertical="center", horizontal="left")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["Citation Name", "Source URL", "Source Description",
               "Extracted Content", "Collection Method", "Attempts"]
    widths = [30, 55, 45, 80, 18, 40]

    font_method = Font(name="Courier New", size=9, color="4ADE80")
    font_attempts = Font(name="Courier New", size=9, color="94A3B8")

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 32

    for ri, ref in enumerate(all_refs, 2):
        bg = "0F1C2E" if ri % 2 == 0 else "0A1220"
        row_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        attempts_list = ref.get("attempts", [])
        attempts_str = " → ".join(attempts_list) if attempts_list else ref.get("status", "")
        values = [
            ref.get("citation_name", ""),
            ref.get("url", ""),
            ref.get("description", ""),
            ref.get("extracted_content", ""),
            ref.get("method", ""),
            attempts_str,
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci == 2:
                cell.font = font_url
            elif ci == 5:
                cell.font = font_method
            elif ci == 6:
                cell.font = font_attempts
            else:
                cell.font = font_data
            cell.fill = row_fill
            cell.alignment = align_wrap
            cell.border = border
            if ci == 2 and val and val.startswith("http"):
                cell.hyperlink = val
        ws.row_dimensions[ri].height = 90

    ws.freeze_panes = "A2"
    if all_refs:
        ws.auto_filter.ref = f"A1:F{1 + len(all_refs)}"

    wb.save(xlsx_path)


def mainsaw(
    operating_platforms,
    search_terms,
    provided_groups,
    show_others,
    art,
    navigationlayers,
    queries,
    truncate,
    attack_frameworks,
    attack_version,
    sheet_tabs,
    columns=None,
    preset=False,
    export_format="csv",
    quiet=False,
    fetch=False,
    evidence_report=False,
    collect_citations=False,
):

    # checking latest version and loading STIX data
    try:
        print("    -> Checking for latest ATT&CK version...")
        latest_version = get_latest_attack_version()

        if latest_version != "Unknown" and latest_version != attack_version:
            print(
                "\n\n\tNote: ATT&CK version in \033[1;36mMITRESaw.py\033[1;m (\033[1;31m{}\033[1;m) differs from \n\tlatest published version (\033[1;31m{}\033[1;m). \n\tUsing latest STIX data from TAXII server...\n".format(
                    attack_version, latest_version
                )
            )
            attack_version = latest_version

        # Load STIX data for all requested frameworks
        all_attack_data = {}
        technique_datasource_map = {}
        for fw in attack_frameworks:
            attack_data, stix_filepath = load_attack_data(fw, force_fetch=fetch)
            all_attack_data[fw] = attack_data
            technique_datasource_map.update(build_technique_datasource_map(stix_filepath))

    except requests.exceptions.ConnectionError:
        print("\n\n\tUnable to connect to the Internet. Please try again.\n\n\n")
        sys.exit()
    except Exception as e:
        print(f"\n\n\tError loading ATT&CK data: {e}\n\n\n")
        sys.exit()

    # Setup output directories
    frameworks_label = ", ".join(attack_frameworks)
    frameworks_slug = "-".join(fw.lower() for fw in attack_frameworks)
    mitresaw_root_date = os.path.join(".", str(datetime.now())[0:10])
    if not os.path.exists(mitresaw_root_date):
        os.makedirs(mitresaw_root_date)
    mitre_files = os.path.join(
        mitresaw_root_date, "{}-{}-stix".format(frameworks_slug, attack_version)
    )
    if not os.path.exists(mitre_files):
        os.makedirs(mitre_files)

    print(f"    -> Using STIX data from TAXII server (cached locally)...")

    time.sleep(0.1)
    saw = """
@                                                         ,
@                 ╓╗╗,                          ,╓▄▄▄Φ▓▓██▌╫D
@                ║▌ `▓L            ,,, ╓▄▄▄Φ▓▓▀▀▀╫╫╫╫╫╫╫▀▀╫▓▓▄
@                 ▓▄▓▓▓        ,▄▄B░▀╫Ñ╬░░╫╫▓▓▓▓╫╫╫╫▓▓▓╫╫╫╫╣▓▓▓▄
@                 ║████L   ,╓#▀▀▀╨╫ÑÑ╦▄▒▀╣▓▄▄▀╣▌╫▀    ██╫╫╫╫▓▓╫▓▓φ
@                  ▓╫╫╫▀]Ñ░░░░ÑÑÑÑ░░░░░╠▀W▄╠▀▓▒░╫Ñ╖   ╙└"╜▀▓▓▓▓▓█▓▓
@                  ║░░░╦╬╫╫╫╫╫╫╫╫╫╫╫╫╫ÑÑ░░░╠Ñ░╨╫Ñ░╫╫╫╫N     ▀▓▓▓╫██▓╕
@                ,]░╦╬╫╫╫╫╫╫╫▓▓▓▓▓▓╫╫╫╫╫╫╫Ñ░░╠░░╫M░╠╫╫╫╫╦,    ▀▓▓▓▓▓▓⌐
@       ╗▄╦     ]░░╬╫╫╫╫╫▓▓██████████▓▓▒╫╫╫╫Ñ░░╟▒╟▓▒ñ▓▓▓▓░N    ╙▓▓▓▓▓▓
@   ║███╫█╫    ]░░╫╫╫╫╫▓███▓▓▓▓▓▓▓▓▓▓███▓╫╫╫╫╫░░╟▒╟▓Ü╟▓▓▓▓░H    ╟▓▓▓▓▓L
@   ║███╫█╫   ]░░╫╫╫╫▓██▓╫▓▓▓▀▀╠╠╬▀▓▓▓╫▓██▓╫╫╫╫░░ÑÑ╠▄░╠▓▓▓▄▄▄▄▄▓▓▓╫╫╫╫
@    ╓▄▄╫█╫╖╖╖╦░╫╫╫╫╫██▓▓▓▓▀░╬Ñ╣╬╫Ñ░╟▓▓▓▓██╫╫╫╫Ñ░╦]░░░║████▀▀╫╫╫▓╩╨╟╫
@    ╟▓▓╫█╫▀▀▀╩╬╩╫╫▓██▓▓▓▓▌░╫░╟▓▓K╫Ñ░▓▓▓▓╫██▓▒╩╩╩╩ ╙╩╨▀▓M╨╩╨╙╝╣N╦╗Φ╝
@       ╫█╫     ▀███▀╣▓▓▓▓▓░╫Ñ░╠▀░╫Ü░▓▓▓▓▓▀▀███╕      ▐▓▌╖
@   ▄▄▄▄▓█▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄╛
@                ▀╩╫╫╫╠╣▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▀░╫╫╫╫▌
@                 ╗▄╫╫Ñ░╠▀▓▓▓▓▓▓▓▓▓▓▓▓▀░╦╬╫╫∩
@                   `⌠╫╫╫Ñ░░Å╣▀▀▀▀▀▒░╦╬╫╫╫`█
@                    ╙╙""╫╫╫½╫╫╫╬╫╫╫╫╫M"▓╛
@                       └╙└ ▄▓╩`║▓╩ Å▀\n\n
    """
    titles = [
        """
       ███▄ ▄███▓ ██▓▄▄▄█████▓ ██▀███  ▓█████   ██████  ▄▄▄       █     █░
      ▓██▒▀█▀ ██▒▓██▒▓  ██▒ ▓▒▓██ ▒ ██▒▓█   ▀ ▒██    ▒ ▒████▄    ▓█░ █ ░█░
      ▓██    ▓██░▒██▒▒ ▓██░ ▒░▓██ ░▄█ ▒▒███   ░ ▓██▄   ▒██  ▀█▄  ▒█░ █ ░█ 
      ▒██    ▒██ ░██░░ ▓██▓ ░ ▒██▀▀█▄  ▒▓█  ▄   ▒   ██▒░██▄▄▄▄██ ░█░ █ ░█ 
      ▒██▒   ░██▒░██░  ▒██▒ ░ ░██▓ ▒██▒░▒████▒▒██████▒▒ ▓█   ▓██▒░░██▒██▓ 
      ░ ▒░   ░  ░░▓    ▒ ░░   ░ ▒▓ ░▒▓░░░ ▒░ ░▒ ▒▓▒ ▒ ░ ▒▒   ▓▒█░░ ▓░▒ ▒  
      ░  ░      ░ ▒ ░    ░      ░▒ ░ ▒░ ░ ░  ░░ ░▒  ░ ░  ▒   ▒▒ ░  ▒ ░ ░  
      ░      ░    ▒ ░  ░        ░░   ░    ░   ░  ░  ░    ░   ▒     ░   ░  
             ░    ░              ░        ░  ░      ░        ░  ░    ░    
""",
        """
      ______  ___________________________ __________________                   
      ___   |/  /____  _/___  __/___  __ \\___  ____/__  ___/______ ____      __
      __  /|_/ /  __  /  __  /   __  /_/ /__  __/   _____ \\ _  __ `/__ | /| / /
      _  /  / /  __/ /   _  /    _  _, _/ _  /___   ____/ / / /_/ / __ |/ |/ / 
      /_/  /_/   /___/   /_/     /_/ |_|  /_____/   /____/  \\__,_/  ____/|__/  
""",
        """
                   ______      ______    ____        ____       ____                              
       /'\\_/`\\    /\\__  _\\    /\\__  _\\  /\\  _`\\     /\\  _`\\    /\\  _`\\                            
      /\\      \\   \\/_/\\ \\/    \\/_/\\ \\/  \\ \\ \\L\\ \\   \\ \\ \\L\\_\\  \\ \\,\\L\\_\\      __      __  __  __  
      \\ \\ \\__\\ \\     \\ \\ \\       \\ \\ \\   \\ \\ ,  /    \\ \\  _\\L   \\/_\\__ \\    /'__`\\   /\\ \\/\\ \\/\\ \\ 
       \\ \\ \\_/\\ \\     \\_\\ \\__     \\ \\ \\   \\ \\ \\\\ \\    \\ \\ \\L\\ \\   /\\ \\L\\ \\ /\\ \\L\\.\\_ \\ \\ \\_/ \\_/ \\
        \\ \\_\\\\ \\_\\    /\\_____\\     \\ \\_\\   \\ \\_\\ \\_\\   \\ \\____/   \\ `\\____\\\\ \\__/.\\_\\ \\ \\___x___/'
         \\/_/ \\/_/    \\/_____/      \\/_/    \\/_/\\/ /    \\/___/     \\/_____/ \\/__/\\/_/  \\/__//__/  
""",
        """
         _____   .___ _____________________ ___________  _________                 
        /     \\  |   |\\__    ___/\\______   \\\\_   _____/ /   _____/_____   __  _  __
       /  \\ /  \\ |   |  |    |    |       _/ |    __)_  \\_____  \\ \\__  \\  \\ \\/ \\/ /
      /    Y    \\|   |  |    |    |    |   \\ |        \\ /        \\ / __ \\_ \\     / 
      \\____|__  /|___|  |____|    |____|_  //_______  //_______  /(____  /  \\/\\_/  
              \\/                         \\/         \\/         \\/      \\/          
""",
        """
        ___ ___   ___   _______   _______   _______   _______                    
       |   Y   | |   | |       | |   _   \\ |   _   | |   _   | .---.-. .--.--.--.
       |.      | |.  | |.|   | | |.  l   / |.  1___| |   1___| |  _  | |  |  |  |
       |. \\_/  | |.  | `-|.  |-' |.  _   1 |.  __)_  |____   | |___._| |________|
       |:  |   | |:  |   |:  |   |:  |   | |:  1   | |:  1   |                   
       |::.|:. | |::.|   |::.|   |::.|:. | |::.. . | |::.. . |                   
       `--- ---' `---'   `---'   `--- ---' `-------' `-------'                   
""",
    ]
    chosen_title = random.choice(titles)
    tagline = "{}        *ATT&CK for {} v{}\n".format(
        chosen_title, frameworks_label, attack_version
    )
    time.sleep(1)
    subprocess.Popen(["clear"]).communicate()
    if art:
        if saw:
            print_saw(saw, tagline, "                                        ")
            print_saw(saw, tagline, "                                      ")
            print_saw(saw, tagline, "                                    ")
            print_saw(saw, tagline, "                                  ")
            print_saw(saw, tagline, "                                ")
            print_saw(saw, tagline, "                              ")
            print_saw(saw, tagline, "                            ")
    platforms = str(operating_platforms)[2:-2].split(",")
    platforms = list(filter(None, platforms))
    if art:
        print_saw(saw, tagline, "                          ")
    terms = str(search_terms)[2:-2].split(",")
    terms = list(filter(None, terms))
    if art:
        print_saw(saw, tagline, "                        ")
    groups = str(provided_groups)[2:-2].split(",")
    groups = list(filter(None, groups))
    if art:
        print_saw(saw, tagline, "                      ")
    else:
        print(tagline)
    if True:  # creating MITRESaw output file names
        if str(platforms) == "['.']":
            platforms_filename_insert = ""
        else:
            platforms_filename_insert = "{}".format(
                str(platforms)[2:-2].replace("', '", "-")
            )
        if str(terms) == "['.']":
            terms_filename_insert = ""
        else:
            terms_filename_insert = "{}".format(str(terms)[2:-2].replace("', '", "-"))
        if str(groups) == "['.']":
            groups_filename_insert = ""
            groups_insert = "Threat Actors"
            all_insert = "all "
        else:
            groups_filename_insert = "{}".format(str(groups)[2:-2].replace("', '", "-"))
            groups_insert = "{}".format(str(groups)[2:-2])
            all_insert = ""
        mitresaw_output_directory = os.path.join(
            mitresaw_root_date,
            "{}_{}_{}".format(
                platforms_filename_insert.replace("_", ""),
                terms_filename_insert.replace("_", ""),
                groups_filename_insert.replace("_", ""),
            ),
        )
        if not os.path.exists(os.path.join(mitresaw_output_directory)):
            os.makedirs(os.path.join(mitresaw_output_directory))
    (
        additional_terms,
        evidence_found,
        valid_procedures,
        all_evidence,
        log_sources,
        logsources,
        groups_in_scope,
        techniques_in_scope,
        groups_techniques_in_scope,
    ) = ([] for _ in range(9))
    (
        group_procedures,
        group_descriptions,
        contextual_information,
        previous_findings,
    ) = ({} for _ in range(4))
    if art:
        if saw:
            print_saw(saw, tagline, "                    ")
            print_saw(saw, tagline, "                  ")
            print_saw(saw, tagline, "                ")
            print_saw(saw, tagline, "              ")
            print_saw(saw, tagline, "            ")
            print_saw(saw, tagline, "          ")
            print_saw(saw, tagline, "        ")
            print_saw(saw, tagline, "      ")
            print_saw(saw, tagline, "    ")
            print_saw(saw, tagline, "  ")
            print_saw(saw, tagline, "partial")
            print_saw(saw, tagline, "-")  # remove saw
            print()
    if str(terms) != "['.']":
        terms_insert = " associated with '\033[1;36m{}\033[1;m'".format(
            str(terms)[2:-2].replace("_", " ").replace("', '", "\033[1;m', '\033[1;36m")
        )
    else:
        terms_insert = ""
    # Use STIX-based parallel processing instead of CSV files
    print()
    print(
        "    -> Extracting \033[1;33mIdentifiers\033[1;m from \033[1;32mTechniques\033[1;m using STIX data based on {}\033[1;31m{}\033[1;m{}".format(
            all_insert,
            groups_insert.replace("', '", "\033[1;m, \033[1;31m"),
            terms_insert,
        )
    )

    # Get group techniques using parallel processing across all frameworks
    print("    -> Extracting techniques with 10 parallel workers...")
    all_group_techniques_data = {}
    all_group_info_data = {}
    for fw, attack_data in all_attack_data.items():
        group_techniques_data, group_info_data, _ = get_group_techniques_parallel(
            attack_data, groups, platforms, max_workers=10
        )
        for gid, techs in group_techniques_data.items():
            for tech in techs:
                tech["framework"] = fw
            all_group_techniques_data.setdefault(gid, []).extend(techs)
        all_group_info_data.update(group_info_data)
    group_techniques_data = all_group_techniques_data
    group_info_data = all_group_info_data

    # Process the STIX data into the format expected by the rest of the tool
    contextual_information = []
    for group_id, techniques in group_techniques_data.items():
        group_name = group_info_data[group_id]["name"]
        group_description = group_info_data[group_id]["description"]

        for technique in techniques:
            technique_id = technique["id"]
            technique_name = technique["name"]
            technique_description = technique["description"]
            technique_usage = technique.get("usage", "")
            technique_platforms = ", ".join(technique["platforms"])
            technique_tactics = ", ".join(technique["tactics"])
            technique_detection = technique.get("detection", "") or ""
            technique_data_sources = technique_datasource_map.get(technique_id, "")
            technique_framework = technique.get("framework", "")

            # Build context string in the format expected by the rest of the tool
            # Format: group_id||group_name||technique_id||technique_name
            context = f"{group_id}||{group_name}||{technique_id}||{technique_name}"
            contextual_information.append(context)

            # obtaining navigation layers for all identified threat groups
            if navigationlayers:
                navlayer_output_directory = os.path.join(
                    mitresaw_root_date,
                    "{}_navigationlayers".format(str(datetime.now())[0:10]),
                )
                for fw in attack_frameworks:
                    domain = f"{fw.lower()}-attack"
                    navlayer_json = os.path.join(
                        navlayer_output_directory,
                        f"{group_id}_{group_name}-{domain}-layer.json",
                    )
                    if not os.path.exists(navlayer_json):
                        if not os.path.exists(navlayer_output_directory):
                            os.makedirs(navlayer_output_directory)
                            print(
                                "     -> Obtaining ATT&CK Navigator Layers for \033[1;33mThreat Actors\033[1;m related to identified \033[1;32mTechniques\033[1;m..."
                            )
                        try:
                            group_navlayer = _fetch(
                                f"https://attack.mitre.org/groups/{group_id}/{group_id}-{domain}-layer.json",
                                timeout=10
                            )
                            if group_navlayer.status_code == 200:
                                with open(navlayer_json, "wb") as navlayer_file:
                                    navlayer_file.write(group_navlayer.content)
                        except Exception as e:
                            print(f"    Warning: Could not download nav layer for {group_name} ({fw}): {e}")

            # Build valid procedure
            # Format expected by extract.py:
            # [0]group_id || [1]group_name || [2]technique_id || [3]technique_name ||
            # [4]usage(relationship desc) || [5]- || [6]group_description(terms) ||
            # [7]technique_description || [8]technique_detection ||
            # [9]technique_platforms || [10]technique_data_sources ||
            # [11]technique_tactics || [12]framework
            # extract.py appends [13]evidence_dict (JSON)
            # Sanitize free-text fields to avoid corrupting the || delimiter
            _usage = technique_usage.replace("||", " ")
            _gdesc = group_description.replace("||", " ")
            _tdesc = technique_description.replace("||", " ")
            _tdet = technique_detection.replace("||", " ")
            _ttactics = technique_tactics.replace("||", " ")
            valid_procedure = f"{group_id}||{group_name}||{technique_id}||{technique_name}||{_usage}||-||{_gdesc}||{_tdesc}||{_tdet}||{technique_platforms}||{technique_data_sources}||{_ttactics}||{technique_framework}"
            valid_procedures.append(valid_procedure)

            # Track techniques
            techniques_in_scope.append(f"{technique_id}||{technique_name}")
            groups_techniques_in_scope.append(f"{group_name}||{technique_id}||{technique_name}||{technique_tactics}")
            groups_in_scope.append(group_name)
    print()
    consolidated_procedures = sorted(list(set(valid_procedures)))
    counted_techniques = Counter(techniques_in_scope)
    sorted_techniques = sorted(
        counted_techniques.items(), key=lambda x: x[1], reverse=True
    )
    sorted_threat_actors_techniques_in_scope = list(set(groups_techniques_in_scope))
    technique_combos = []
    for technique in counted_techniques.most_common():
        technique_count = technique[1]
        if ": " in technique[0]:
            parent_technique = technique[0].split(": ")[0]
            sub_technique = technique[0].split(": ")[1]
        else:
            parent_technique = technique[0]
            sub_technique = "-"
        technique_combo = [parent_technique, sub_technique, technique_count]
        technique_combos.append(technique_combo)
    # Build STIX citation lookup if -R is enabled (before extraction loop)
    # Build citation lookup: collect ALL external_references from STIX relationships
    # Keyed by citation source_name → {url, description} for direct lookup
    _citation_url_lookup = {}  # source_name → {"url": ..., "description": ...}
    _mitre_ref_numbers = {}   # (group_name_lower, source_name) → MITRE [N] number
    _all_citation_refs = []
    _seen_citations = set()
    if collect_citations:
        for _fw, _ad in all_attack_data.items():
            _sp = getattr(_ad, 'stix_filepath', None) or getattr(_ad, 'src', None)
            if not _sp:
                continue
            try:
                import json as _json
                with open(_sp) as _f:
                    _bundle = _json.load(_f)
                # Build citation URL lookup from relationships
                for _obj in _bundle.get("objects", []):
                    if _obj.get("type") != "relationship":
                        continue
                    for _ref in _obj.get("external_references", []):
                        _sn = _ref.get("source_name", "")
                        if _sn and _sn != "mitre-attack" and _sn not in _citation_url_lookup:
                            _citation_url_lookup[_sn] = {
                                "url": _ref.get("url", ""),
                                "description": _ref.get("description", ""),
                            }
                # Build MITRE reference numbers from group objects
                for _obj in _bundle.get("objects", []):
                    if _obj.get("type") != "intrusion-set":
                        continue
                    _gname = _obj.get("name", "").strip().lower()
                    _num = 0
                    for _ref in _obj.get("external_references", []):
                        _sn = _ref.get("source_name", "")
                        if _sn == "mitre-attack":
                            continue
                        _num += 1
                        _mitre_ref_numbers[(_gname, _sn)] = _num
            except Exception:
                continue
        if _citation_url_lookup:
            print(f"    -> {len(_citation_url_lookup)} unique citation sources indexed for collection\n")

    # Sort by (group, technique_name) so procedures for the same group+technique are contiguous
    consolidated_procedures = sorted(consolidated_procedures, key=lambda p: (
        p.split("||")[1].strip().lower(),
        p.split("||")[3].strip().lower() if len(p.split("||")) > 3 else "",
    ))

    last_group_name = None
    _total_procedures = len(consolidated_procedures)
    _pb_extract = _ProgressBar("Processing:")
    _cit_num = 0  # running citation counter, resets per group
    _deferred_cits = []  # citations from procedures with no technique output

    for _proc_idx, each_procedure in enumerate(consolidated_procedures, 1):
        _proc_parts = each_procedure.split("||")
        current_group_name = _proc_parts[1]
        if last_group_name and current_group_name.strip().lower() != last_group_name.strip().lower():
            _cit_num = 0
            _deferred_cits = []  # don't carry citations across groups
        last_group_name = current_group_name
        if quiet:
            _cit_label = f"{current_group_name} ({len(_all_citation_refs)} refs)" if collect_citations else current_group_name
            _pb_extract.update(_proc_idx, _total_procedures, _cit_label)
        (
            technique_findings,
            previous_findings,
        ) = extract_indicators(
            each_procedure,
            terms,
            evidence_found,
            "",
            previous_findings,
            truncate,
            quiet,
        )
        threat_actor_technique_id_name_findings = []

        # constructing sub-technique pairing
        for technique_found in technique_findings:
            threat_actor_found = technique_found.split("||")[1]
            technique_id_found = technique_found.split("||")[2]
            technique_name_found = technique_found.split("||")[3]
            if "." in technique_id_found:
                parent_technique_found = "{}||{}".format(
                    technique_id_found,
                    str(sorted_techniques)
                    .split("{}||".format(technique_id_found))[1]
                    .split("{}".format(technique_name_found))[0][0:-2],
                )
                technique_id_name_found = "{}: {}".format(
                    parent_technique_found, technique_name_found
                )
            else:
                technique_id_name_found = "{}||{}".format(
                    technique_id_found, technique_name_found
                )
            threat_actor_technique_id_name_found = "{}||{}".format(
                threat_actor_found, technique_id_name_found
            )
            threat_actor_technique_id_name_findings.append(
                threat_actor_technique_id_name_found
            )

        # Collect citations silently into buffer (if -C enabled)
        if collect_citations and _citation_url_lookup:
            _parts = each_procedure.split("||")
            _raw_proc = _parts[4] if len(_parts) > 4 else ""
            _group = _parts[1] if len(_parts) > 1 else ""
            _tid = _parts[2] if len(_parts) > 2 else ""
            _tname = _parts[3] if len(_parts) > 3 else ""

            _cit_names = re.findall(r"\(Citation:\s*([^)]+)\)", _raw_proc)
            _new_cits = []
            if _cit_names:
                from toolbox.citation_collector import collect_reference_content
                for _cn in _cit_names:
                    _cn = _cn.strip()
                    # Dedup display per (group, citation) — same citation shows under each group
                    _display_key = (_group.strip().lower(), _cn)
                    if _display_key in _seen_citations:
                        continue
                    _seen_citations.add(_display_key)

                    _ref_data = _citation_url_lookup.get(_cn, {})
                    _cit = {
                        "citation_name": _cn,
                        "url": _ref_data.get("url", ""),
                        "description": _ref_data.get("description", ""),
                    }
                    # collect_reference_content uses disk cache, so already-fetched URLs are instant
                    _fetched = collect_reference_content(
                        [_cit], _group, _tname, _tid, verbose=False,
                    )
                    for _ref in _fetched:
                        _ref["group"] = _group
                        _ref["technique_id"] = _tid
                        _ref["technique_name"] = _tname
                        _all_citation_refs.append(_ref)
                        _new_cits.append(_ref)

            # If no technique output, defer citations to print with next procedure that has output
            if _new_cits and not technique_findings:
                _deferred_cits.extend(_new_cits)
                _new_cits = []

            # Print deferred + new citations when technique output was produced
            if technique_findings:
                _print_cits = _deferred_cits + _new_cits
                _deferred_cits = []
            else:
                _print_cits = []
            if _print_cits:
                _indent = "          "
                for _ref in _print_cits:
                    _cit_num += 1
                    _cn = _ref.get("citation_name", "")
                    _num_str = f"[{_cit_num}]"
                    _method = _ref.get("method", "unknown")
                    _icon = "\033[32m\u2705\033[0m" if _ref.get("extracted_content") else "\033[31m\u274c\033[0m"
                    _name = _cn[:28].ljust(28)
                    _method_short = _method[:14].ljust(14)
                    _url = _ref.get("url", "")
                    try:
                        _tw = os.get_terminal_size().columns
                    except OSError:
                        _tw = 120
                    _used = 10 + 5 + 1 + 28 + 4 + 14 + 3 + 5
                    _url_max = max(30, _tw - _used)
                    _url_part = f" - {_url[:_url_max]}" if _url else ""
                    print(f"{_indent}\033[90m{_num_str:>5}\033[0m \033[36m{_name}\033[0m \033[90m\u2192\033[0m \033[33m{_method_short}\033[0m {_icon}{_url_part}")


    threat_actor_technique_id_name_findings = list(
        set(threat_actor_technique_id_name_findings)
    )
    if quiet:
        _done_label = "Extraction complete"
        if collect_citations:
            _with_content = sum(1 for r in _all_citation_refs if r.get("extracted_content"))
            _done_label = f"Complete — {len(_all_citation_refs)} citations, {_with_content} with content"
        _pb_extract.done(_total_procedures, _done_label)
    elif collect_citations and _all_citation_refs:
        _with_content = sum(1 for r in _all_citation_refs if r.get("extracted_content"))
        print(f"\n     {len(_all_citation_refs)} citations collected, {_with_content} with content")
    all_evidence.append(technique_findings)
    consolidated_techniques = all_evidence[0]

    # Report CVEs with no actionable intelligence
    from toolbox.tools.map_bespoke_logs import report_cve_summary
    report_cve_summary()

    if len(consolidated_techniques) > 0:

        # outputting relevant queries
        query_pairings, mapped_log_sources = build_matrix(
            mitresaw_output_directory,
            consolidated_techniques,
            sorted_threat_actors_techniques_in_scope,
            threat_actor_technique_id_name_findings,
        )
        """if queries:
            print()
            print(
                "    -> Compiling queries based on \033[1;31midentifiers\033[1;m based on {}".format(
                    terms_insert
                )
            )"""
        # outputting csv file for ingestion into other tools
        write_csv_summary(
            consolidated_techniques,
            mitresaw_output_directory,
            mitre_files,
            queries,
            query_pairings,
            log_sources,
        )
        # Export main CSV files to JSON/XML if requested
        if export_format != "csv":
            for csv_name in ["ThreatActors_Techniques"]:
                csv_path = os.path.join(mitresaw_output_directory, f"{csv_name}.csv")
                if os.path.exists(csv_path):
                    df = pandas.read_csv(csv_path, on_bad_lines="warn")
                    if export_format == "json":
                        out_path = os.path.join(mitresaw_output_directory, f"{csv_name}.json")
                        df.to_json(out_path, orient="records", indent=2)
                    elif export_format == "xml":
                        out_path = os.path.join(mitresaw_output_directory, f"{csv_name}.xml")
                        df.to_xml(out_path)
                    print(f"      {export_format.upper()} written to {out_path}")
        # Generate filtered export if --columns is specified
        if columns:
            valid_columns = [
                "group_sw_id", "group_sw_name", "group_sw_description",
                "technique_id", "technique_name", "technique_description",
                "tactic", "platforms", "framework",
                "procedure_example", "evidence", "detectable_via",
                "keywords",
            ]
            requested_columns = [c.strip() for c in columns.split(",")]
            invalid = [c for c in requested_columns if c not in valid_columns]
            if invalid:
                print(f"\n    Error: Invalid column(s): {', '.join(invalid)}")
                print(f"    Valid columns: {', '.join(valid_columns)}")
            else:
                csv_path = os.path.join(mitresaw_output_directory, "ThreatActors_Techniques.csv")
                df = pandas.read_csv(csv_path, on_bad_lines="warn")

                if "keywords" in requested_columns:
                    keyword_map = {}
                    for gid, info in group_info_data.items():
                        keyword_map[gid] = match_keywords(info["description"])
                    df["keywords"] = df["group_sw_id"].map(keyword_map).fillna("")

                df = df[requested_columns].drop_duplicates()
                if preset:
                    filtered_base_name = "mitre_procedures"
                    filtered_dir = mitresaw_root_date
                else:
                    filtered_base_name = "ThreatActors_Keywords"
                    filtered_dir = mitresaw_output_directory
                if export_format == "json":
                    filtered_path = os.path.join(filtered_dir, f"{filtered_base_name}.json")
                    df.to_json(filtered_path, orient="records", indent=2)
                elif export_format == "xml":
                    filtered_path = os.path.join(filtered_dir, f"{filtered_base_name}.xml")
                    df.to_xml(filtered_path)
                else:
                    filtered_path = os.path.join(filtered_dir, f"{filtered_base_name}.csv")
                    df.to_csv(filtered_path, index=False)
                if not evidence_report:
                    print(f"      Filtered export written to {filtered_path}")

        mitresaw_techniques = re.findall(
            r"\|\|(T\d{3}[\d\.]+)\|\|", str(consolidated_techniques)
        )
        mitresaw_techniques = list(set(mitresaw_techniques))
        mitresaw_techniques_insert = str(mitresaw_techniques)[2:-2].replace(
            "', '",
            '", "comment": "", "score": 1, "color": "#66b1ff", "showSubtechniques": false}}, {{"techniqueID": "',
        )

        # Generate ATT&CK Navigator layer per framework
        for fw in attack_frameworks:
            domain = f"{fw.lower()}-attack"
            mitresaw_navlayer = '{{"description": "{} techniques used by various Threat Actors, produced by MITRESaw", "name": "{}", "domain": "{}", "versions": {{"layer": "4.4", "attack": "15", "navigator": "4.8.1"}}, "techniques": [{{"techniqueID": "{}", "comment": "", "score": 1, "color": "#66b1ff", "showSubtechniques": false}}], "gradient": {{"colors": ["#ffffff", "#66b1ff"], "minValue": 0, "maxValue": 1}}, "legendItems": [{{"label": "identified from MITRESaw analysis", "color": "#66b1ff"}}]}}\n'.format(
                fw, mitresaw_output_directory.split("/")[2][11:], domain, mitresaw_techniques_insert
            )
            with open(
                os.path.join(mitresaw_output_directory, f"{domain}-layer.json"), "w"
            ) as mitresaw_navlayer_json:
                mitresaw_navlayer_json.write(
                    mitresaw_navlayer.replace("{{", "{").replace("}}", "}")
                )
        build_queries(queries, mitresaw_output_directory, query_pairings)

        # Generate evidence report if requested
        if evidence_report:
            csv_path = os.path.join(mitresaw_output_directory, "ThreatActors_Techniques.csv")
            if os.path.exists(csv_path):
                df = pandas.read_csv(csv_path, on_bad_lines="warn")
                result_rows = df.to_dict(orient="records")
                from toolbox.evidence_report import generate_evidence_report
                from datetime import datetime as _dt
                # If no filters provided, put alongside mitre_procedures.csv
                no_filters = (
                    str(operating_platforms) == "['.']"
                    and str(search_terms) == "['.']"
                    and str(provided_groups) == "['.']"
                )
                _er_dir = mitresaw_root_date if no_filters else mitresaw_output_directory
                _er_path = os.path.join(
                    _er_dir,
                    "mitre_procedures.xlsx",
                )
                generate_evidence_report(
                    rows=result_rows,
                    output_path=_er_path,
                    framework=",".join(attack_frameworks),
                    platforms_arg=",".join(str(p) for p in operating_platforms),
                    searchterms_arg=",".join(str(s) for s in search_terms),
                    threatgroups_arg=",".join(str(g) for g in provided_groups),
                )

                # Append Reference Detail sheet if citations were collected
                if collect_citations and _all_citation_refs:
                    _write_reference_sheet(_er_path, _all_citation_refs)

                # Move CSV alongside evidence report with matching name
                import shutil
                _csv_dest = os.path.join(_er_dir, "mitre_procedures.csv")
                shutil.move(csv_path, _csv_dest)
                print(f"\n\n     Outputs written to: {_er_dir}/")
                print(f"                              mitre_procedures.csv")
                print(f"                              mitre_procedures.xlsx")

    else:
        print("\n    -> No evidence could be found which match the provided criteria.")
    print("\n\n")
