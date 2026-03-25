"""
MITRESaw Evidence Report Generator

Generates a styled XLSX evidence report with one row per atomic indicator
extracted from MITRE ATT&CK procedure examples.
"""

import json
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# PATTERN 1-7: Extraction regexes for extract_procedure_invocations()
# ---------------------------------------------------------------------------

_RE_BACKTICK = re.compile(r"`([^`]+)`")

_RE_DOUBLE_QUOTE = re.compile(r'"([^"]{4,120})"')
_RE_DQUOTE_EXEC = re.compile(
    r"(?:[\\\/]|\.[\w]{2,4}\b|(?:[-\/])\s)", re.IGNORECASE
)

_RE_EXE_INVOKE = re.compile(
    r"\b([A-Za-z0-9_\-]+\.(?:exe|ps1|bat|vbs|sh|py|pl|dll|cmd)"
    r"(?:\s+[^\.\n]{0,80})?)\b",
    re.IGNORECASE,
)

_CLI_PATTERNS = [
    re.compile(p, re.IGNORECASE)
    for p in [
        r"net\s+(?:user|group|localgroup|use|view|share|session|start|stop)\s+\S[^\n]{0,100}",
        r"reg\s+(?:add|delete|query|export|import)\s+\S[^\n]{0,100}",
        r"schtasks\s+/[A-Za-z][^\n]{0,120}",
        r"powershell(?:\.exe)?\s+[-/][^\n]{0,150}",
        r"cmd(?:\.exe)?\s+/[cCkK]\s+[^\n]{0,120}",
        r"wmic\s+\S[^\n]{0,100}",
        r"certutil\s+[-/][^\n]{0,100}",
        r"bitsadmin\s+/[^\n]{0,100}",
        r"mshta(?:\.exe)?\s+\S[^\n]{0,100}",
        r"wscript(?:\.exe)?\s+\S[^\n]{0,100}",
        r"cscript(?:\.exe)?\s+\S[^\n]{0,100}",
        r"rundll32(?:\.exe)?\s+\S[^\n]{0,100}",
        r"regsvr32(?:\.exe)?\s+\S[^\n]{0,100}",
        r"sc\s+(?:create|start|stop|delete|config|query)\s+\S[^\n]{0,100}",
        r"vssadmin\s+\S[^\n]{0,80}",
        r"nltest\s+/[^\n]{0,100}",
        r"dsquery\s+\S[^\n]{0,100}",
        r"ipconfig\s*[^\n]{0,60}",
        r"whoami\s*[^\n]{0,60}",
        r"ssh(?:\.exe)?\s+[-\w][^\n]{0,100}",
        r"curl\s+[-\w][^\n]{0,150}",
        r"wget\s+[-\w][^\n]{0,150}",
    ]
]

_RE_CVE = re.compile(r"(CVE-\d{4}-\d{4,7}[^.;\n]{0,200})")
_RE_CVE_ID_ONLY = re.compile(r"CVE-\d{4}-\d{4,7}")

_RE_REG_PATH = re.compile(r"HK(?:LM|CU|CR|U|CC)\\[^\s\n\"'`]{6,200}")

_RE_WIN_PATH = re.compile(r"[A-Za-z]:\\[^\s\n\"'`]{4,200}")
_RE_UNC_PATH = re.compile(r"\\\\[^\s\n\"'`]{4,200}")
_RE_UNIX_PATH = re.compile(
    r"/(?:etc|var|tmp|usr|home|opt|bin|sbin|proc)/[^\s\n\"'`]{2,150}"
)

# Sentence boundary for capping exe invocations
_RE_SENTENCE_END = re.compile(r"[.;`]\s|[`]\b|\n")


def _cap_at_sentence(text: str) -> str:
    """Cap text at the first sentence boundary."""
    m = _RE_SENTENCE_END.search(text)
    if m:
        return text[: m.start()].rstrip(" `")
    return text.rstrip(" `")


def _is_relevant(match: str, indicator_value: str, indicator_type: str) -> bool:
    """Check if a match is relevant to the indicator_value."""
    match_lower = match.lower()
    iv_lower = indicator_value.lower()

    # Exception: cmd type where the indicator_value IS the match
    if indicator_type == "cmd" and match_lower.strip() == iv_lower.strip():
        return True

    # indicator_value appears within the match as a substring
    if iv_lower in match_lower:
        return True

    # match shares at least one significant token (length >= 4) with indicator_value
    iv_tokens = set(re.findall(r"\w{4,}", iv_lower))
    match_tokens = set(re.findall(r"\w{4,}", match_lower))
    if iv_tokens & match_tokens:
        return True

    return False


def extract_procedure_invocations(
    procedure_text: str,
    indicator_type: str,
    indicator_value: str,
) -> list:
    """Extract invocation strings from MITRE procedure text for a specific indicator.

    Returns a list of strings found in the procedure text that are relevant
    to the given indicator_type and indicator_value.
    """
    if not procedure_text or not indicator_value:
        return []

    backtick_matches = []
    cli_matches = []
    other_matches = []

    # PATTERN 1 — Backtick-quoted strings
    for m in _RE_BACKTICK.finditer(procedure_text):
        backtick_matches.append(m.group(1))

    # PATTERN 2 — Double-quoted strings with executable-looking content
    for m in _RE_DOUBLE_QUOTE.finditer(procedure_text):
        content = m.group(1)
        if _RE_DQUOTE_EXEC.search(content):
            other_matches.append(content)

    # PATTERN 3 — Executable invocations (unquoted)
    for m in _RE_EXE_INVOKE.finditer(procedure_text):
        capped = _cap_at_sentence(m.group(1))
        other_matches.append(capped)

    # PATTERN 4 — Common CLI command prefixes
    for pat in _CLI_PATTERNS:
        for m in pat.finditer(procedure_text):
            capped = _cap_at_sentence(m.group(0))
            cli_matches.append(capped)

    # PATTERN 5 — CVE IDs (when indicator_type is "cve")
    if indicator_type == "cve":
        for m in _RE_CVE.finditer(procedure_text):
            text = m.group(1).rstrip(".,;) ")
            # Trim at next full stop
            dot_pos = text.find(".")
            if dot_pos > 15:  # after CVE-YYYY-NNNNN
                text = text[:dot_pos]
            other_matches.append(text.strip())

    # PATTERN 6 — Windows Registry paths
    if indicator_type == "reg":
        for m in _RE_REG_PATH.finditer(procedure_text):
            other_matches.append(m.group(0))

    # PATTERN 7 — File/directory paths
    if indicator_type in ("paths", "filepath"):
        for m in _RE_WIN_PATH.finditer(procedure_text):
            other_matches.append(m.group(0))
        for m in _RE_UNC_PATH.finditer(procedure_text):
            other_matches.append(m.group(0))
        for m in _RE_UNIX_PATH.finditer(procedure_text):
            other_matches.append(m.group(0))

    # --- Filtering and relevance scoring ---

    # Step 1 — Relevance filter
    def keep(match, source):
        if indicator_type == "cmd" and match.strip().lower() == indicator_value.strip().lower():
            return True
        return _is_relevant(match, indicator_value, indicator_type)

    backtick_matches = [m for m in backtick_matches if keep(m, "backtick")]
    cli_matches = [m for m in cli_matches if keep(m, "cli")]
    other_matches = [m for m in other_matches if keep(m, "other")]

    # Step 2 — Deduplication (case-insensitive, preserving first occurrence)
    seen = set()
    ordered = []
    # Return order: backtick first, then CLI, then others
    for m in backtick_matches + cli_matches + other_matches:
        key = m.strip().lower()
        if key not in seen:
            seen.add(key)
            ordered.append(m.strip())

    # Step 3 — Length filter
    ordered = [m for m in ordered if 4 <= len(m) <= 300]

    # Step 4 — Max 5 items
    return ordered[:5]


# ---------------------------------------------------------------------------
# Port protocol mapping
# ---------------------------------------------------------------------------

_PORT_PROTOCOL = {
    "80": "TCP/{} (HTTP)",
    "8080": "TCP/{} (HTTP)",
    "8443": "TCP/{} (HTTPS)",
    "8888": "TCP/{} (HTTP)",
    "443": "TCP/{} (TLS)",
    "53": "UDP/TCP/{} (DNS)",
    "22": "TCP/{} (SSH)",
    "3389": "TCP/{} (RDP)",
    "445": "TCP/{} (SMB)",
    "139": "TCP/{} (NetBIOS)",
}


def _format_port(port_str: str) -> str:
    """Format a port string with protocol prefix."""
    port = port_str.strip()
    fmt = _PORT_PROTOCOL.get(port)
    if fmt:
        return fmt.format(port)
    return f"TCP/{port}"


# ---------------------------------------------------------------------------
# Detection context strings
# ---------------------------------------------------------------------------

_DETECTION_CONTEXT = {
    "cmd": (
        "\nDetection: Process Creation \u2014 Sysmon EID 1 / Windows Security EID 4688"
        " (requires command-line auditing enabled)"
    ),
    "reg": (
        "\nDetection: Registry modification \u2014 Sysmon EID 12/13/14 / Windows"
        " Security EID 4657 (requires object access auditing)"
    ),
    "cve": (
        "\nDetection: Exploit telemetry \u2014 check CISA KEV for active exploitation"
        " status; review NVD for PoC availability; patch status is primary control"
    ),
    "ports": (
        "\nDetection: Network traffic \u2014 firewall/proxy logs, Zeek conn.log,"
        " Sysmon EID 3 (network connection)"
    ),
    "paths": (
        "\nDetection: File creation/modification \u2014 Sysmon EID 11 (FileCreate),"
        " EID 23 (FileDelete) / EDR file telemetry"
    ),
    "filepath": (
        "\nDetection: File creation/modification \u2014 Sysmon EID 11 (FileCreate),"
        " EID 23 (FileDelete) / EDR file telemetry"
    ),
    "software": (
        "\nDetection: Process name / image load \u2014 Sysmon EID 1 (process),"
        " EID 7 (image load); check GitHub for tool-specific CLI usage"
    ),
    "event_ids": (
        "\nDetection: This IS a Windows event ID \u2014 ensure the corresponding log"
        " channel is enabled and ingested into your SIEM"
    ),
    "evt": (
        "\nDetection: This IS a Windows event ID \u2014 ensure the corresponding log"
        " channel is enabled and ingested into your SIEM"
    ),
    "none": "(no extractable indicators \u2014 review procedure text manually)",
}

# Source type per indicator type
_SOURCE_TYPE = {
    "cmd": "Website",
    "reg": "Website",
    "cve": "Website",
    "ports": "Website",
    "paths": "Website",
    "filepath": "Website",
    "software": "GitHub | Website",
    "event_ids": "Website",
    "evt": "Website",
    "none": "Website",
}


# ---------------------------------------------------------------------------
# Group colour/background mapping
# ---------------------------------------------------------------------------

def _group_bg(name: str) -> str:
    """Return hex background colour for a group row."""
    n = name.lower()
    if "oilrig" in n or "apt34" in n:
        return "0F2035"
    if "apt33" in n or "peach sandstorm" in n:
        return "0F250F"
    if "muddywater" in n or "seedworm" in n:
        return "1F0A2A"
    if "magic hound" in n or "apt35" in n:
        return "2A150A"
    if "apt39" in n or "chafer" in n:
        return "0A1535"
    if "fox kitten" in n or "pioneer kitten" in n:
        return "1A0A2A"
    return "0F1C2E"


def _group_accent(name: str) -> str:
    """Return hex accent colour for group name font."""
    n = name.lower()
    if "oilrig" in n or "apt34" in n:
        return "38BDF8"
    if "apt33" in n or "peach sandstorm" in n:
        return "4ADE80"
    if "muddywater" in n or "seedworm" in n:
        return "C084FC"
    if "magic hound" in n or "apt35" in n:
        return "FB923C"
    if "apt39" in n or "chafer" in n:
        return "60A5FA"
    if "fox kitten" in n or "pioneer kitten" in n:
        return "E879F9"
    return "E0F2FE"


# ---------------------------------------------------------------------------
# URL helpers
# ---------------------------------------------------------------------------

_RE_URL = re.compile(r"https?://\S{10,}")


def _extract_url(procedure_text: str, technique_id: str) -> str:
    """Extract URL from procedure text or construct ATT&CK technique URL."""
    m = _RE_URL.search(procedure_text or "")
    if m:
        url = m.group(0).rstrip(".,;)")
        return url
    # Construct from technique_id: T1059.001 → T1059/001
    tid = technique_id.replace(".", "/")
    return f"https://attack.mitre.org/techniques/{tid}/"


def _nav_layer_url(group_id: str) -> str:
    """Construct ATT&CK Navigator layer URL from group ID."""
    if not group_id or not group_id.startswith("G"):
        return "N/A"
    return f"https://attack.mitre.org/groups/{group_id}/{group_id}-enterprise-layer.json"


# ---------------------------------------------------------------------------
# Contextual evidence builder
# ---------------------------------------------------------------------------

def _build_contextual_evidence(
    procedure_text: str,
    indicator_type: str,
    indicator_value: str,
    detectable_via: str,
) -> tuple:
    """Build contextual evidence string and return (text, had_invocations)."""
    parts = []

    # Step A — extract invocations
    invocations = extract_procedure_invocations(
        procedure_text, indicator_type, indicator_value
    )

    had_invocations = len(invocations) > 0

    # Step B — primary evidence from invocations
    if invocations:
        lines = "\n".join(f"  \u2022 {inv}" for inv in invocations)
        parts.append(f"MITRE documented invocation(s):\n{lines}")

    # Step D — no invocations fallback
    if not invocations and indicator_type != "none":
        parts.append(
            "No specific invocation documented in MITRE procedure text for this indicator."
        )

    # Step C — detection context
    det = _DETECTION_CONTEXT.get(indicator_type, "")
    if det:
        parts.append(det)

    # Step E — ATT&CK data sources
    if detectable_via and str(detectable_via).strip() and str(detectable_via) != "nan":
        parts.append(f"\nATT&CK Data Source(s): {detectable_via}")

    return "\n".join(parts), had_invocations


# ---------------------------------------------------------------------------
# XLSX styling constants
# ---------------------------------------------------------------------------

_BG_NAVY = "0D1B2A"
_BG_DARK = "0F1C2E"
_BG_ALT = "0A1220"
_BORDER_CLR = "1E3A5F"

_THIN_BORDER = Border(
    left=Side(style="thin", color=_BORDER_CLR),
    right=Side(style="thin", color=_BORDER_CLR),
    top=Side(style="thin", color=_BORDER_CLR),
    bottom=Side(style="thin", color=_BORDER_CLR),
)

_COL_WIDTHS = [50, 16, 55, 14, 28, 18, 72, 45, 38, 16]

_HEADER_LABELS = [
    "Evidential Element\n(Atomic Indicator / Command / Artefact)",
    "Threat Group",
    "Procedure Example\n(MITRE ATT&CK \u2014 verbatim)",
    "Technique ID",
    "Technique Name",
    "Tactic",
    "Contextual Evidence\n(MITRE Invocations + Detection Guidance)",
    "Reference URL",
    "Navigation Layer URL\n(ATT&CK Navigator JSON)",
    "Source Type",
]


# ---------------------------------------------------------------------------
# generate_evidence_report()
# ---------------------------------------------------------------------------

def generate_evidence_report(
    rows: list,
    output_path: str,
    framework: str = "Enterprise",
    platforms_arg: str = ".",
    searchterms_arg: str = ".",
    threatgroups_arg: str = ".",
) -> None:
    """Generate a styled XLSX evidence report from MITRESaw result rows.

    Parameters
    ----------
    rows : list[dict]
        Result dicts with keys matching ThreatActors_Techniques.csv columns:
        group_sw_id, group_sw_name, technique_id, technique_name, tactic,
        procedure_example, evidence, detectable_via.
    output_path : str
        Path for the output XLSX file.
    framework, platforms_arg, searchterms_arg, threatgroups_arg : str
        Metadata for the report title/subtitle.
    """

    # Map from evidence dict keys to canonical indicator types
    # The evidence JSON uses "evt" and "filepath" internally
    evidence_key_map = {
        "cmd": "cmd",
        "reg": "reg",
        "cve": "cve",
        "ports": "ports",
        "filepath": "paths",
        "software": "software",
        "evt": "event_ids",
    }

    # ------------------------------------------------------------------
    # ATOMISE rows → one output row per indicator
    # ------------------------------------------------------------------
    atomised = []
    seen = set()

    for row in rows:
        group_id = str(row.get("group_sw_id", "") or "")
        group_name = str(row.get("group_sw_name", "") or "")
        technique_id = str(row.get("technique_id", "") or "")
        technique_name = str(row.get("technique_name", "") or "")
        tactic = str(row.get("tactic", "") or "")
        procedure_text = str(row.get("procedure_example", "") or "")
        detectable_via = str(row.get("detectable_via", "") or "")
        evidence_raw = row.get("evidence", "{}")

        # Parse evidence JSON
        if isinstance(evidence_raw, str):
            try:
                evidence = json.loads(evidence_raw)
            except (json.JSONDecodeError, TypeError):
                evidence = {}
        elif isinstance(evidence_raw, dict):
            evidence = evidence_raw
        else:
            evidence = {}

        has_any = False

        for ev_key, indicator_type in evidence_key_map.items():
            values = evidence.get(ev_key, [])
            if not values:
                continue

            has_any = True
            for val in values:
                # CVE entries are dicts: {"CVE-ID": "description|..."}
                if isinstance(val, dict):
                    indicator_value = list(val.keys())[0] if val else ""
                else:
                    indicator_value = str(val)

                if not indicator_value:
                    continue

                # Format display value
                if indicator_type == "ports":
                    display_value = _format_port(indicator_value)
                elif indicator_type == "event_ids":
                    display_value = f"Windows Event ID {indicator_value}"
                else:
                    display_value = indicator_value

                # Dedup on (group_name, technique_id, display_value)
                dedup_key = (group_name.lower(), technique_id.lower(), display_value.lower())
                if dedup_key in seen:
                    continue
                seen.add(dedup_key)

                # Build contextual evidence
                ctx, had_inv = _build_contextual_evidence(
                    procedure_text, indicator_type, indicator_value, detectable_via
                )

                # Reference URL
                ref_url = _extract_url(procedure_text, technique_id)

                # Navigation layer URL
                nav_url = _nav_layer_url(group_id)

                # Source type
                source_type = _SOURCE_TYPE.get(indicator_type, "Website")

                atomised.append({
                    "evidential_element": display_value,
                    "threat_group": group_name,
                    "procedure_example": procedure_text,
                    "technique_id": technique_id,
                    "technique_name": technique_name,
                    "tactic": tactic,
                    "contextual_evidence": ctx,
                    "reference_url": ref_url,
                    "nav_layer_url": nav_url,
                    "source_type": source_type,
                    "had_invocations": had_inv,
                })

        # If no evidence at all, emit placeholder row
        if not has_any:
            dedup_key = (group_name.lower(), technique_id.lower(), "(no extractable indicators)")
            if dedup_key not in seen:
                seen.add(dedup_key)
                ctx = _DETECTION_CONTEXT["none"]
                atomised.append({
                    "evidential_element": "(no extractable indicators)",
                    "threat_group": group_name,
                    "procedure_example": procedure_text,
                    "technique_id": technique_id,
                    "technique_name": technique_name,
                    "tactic": tactic,
                    "contextual_evidence": ctx,
                    "reference_url": _extract_url(procedure_text, technique_id),
                    "nav_layer_url": _nav_layer_url(group_id),
                    "source_type": "Website",
                    "had_invocations": False,
                })

    # ------------------------------------------------------------------
    # Build XLSX workbook
    # ------------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Evidence Report"
    ws.sheet_view.showGridLines = False

    # Fonts
    font_title = Font(name="Calibri", size=16, bold=True, color="0EA5E9")
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="7FB3D3")
    font_header = Font(name="Calibri", size=12, bold=True, color="E0F2FE")
    font_col1 = Font(name="Courier New", size=10, bold=True, color="2DD4BF")
    font_col4 = Font(name="Courier New", size=10, bold=True, color="22C55E")
    font_col5 = Font(name="Calibri", size=10, color="E0F2FE")
    font_col3 = Font(name="Calibri", size=10, color="CBD5E1")
    font_col6 = Font(name="Calibri", size=10, color="FACC15")
    font_col7 = Font(name="Courier New", size=10, color="CBD5E1")
    font_col8 = Font(name="Calibri", size=10, color="0EA5E9")
    font_col9 = Font(name="Calibri", size=10, color="A78BFA")
    font_col10 = Font(name="Calibri", size=10, color="F97316")

    fill_navy = PatternFill(start_color=_BG_NAVY, end_color=_BG_NAVY, fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Column widths
    for i, w in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ------------------------------------------------------------------
    # Row 1 — Title banner
    # ------------------------------------------------------------------
    unique_groups = len(set(r["threat_group"] for r in atomised)) if atomised else 0
    total_indicators = len(atomised)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    title_text = (
        f"MITRESaw Evidence Report  |  {framework}  |  "
        f"Groups: {unique_groups}  |  Indicators: {total_indicators}  |  "
        f"Generated: {now_str}"
    )
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = title_text
    cell.font = font_title
    cell.fill = fill_navy
    cell.alignment = align_center
    cell.border = _THIN_BORDER
    ws.row_dimensions[1].height = 28

    # ------------------------------------------------------------------
    # Row 2 — Subtitle
    # ------------------------------------------------------------------
    subtitle_text = (
        f"Platforms: {platforms_arg}  |  Search Terms: {searchterms_arg}  |  "
        f"Threat Groups: {threatgroups_arg}  |  Source: MITRE ATT&CK STIX via MITRESaw"
    )
    ws.merge_cells("A2:J2")
    cell = ws["A2"]
    cell.value = subtitle_text
    cell.font = font_subtitle
    cell.fill = fill_navy
    cell.alignment = align_center
    cell.border = _THIN_BORDER
    ws.row_dimensions[2].height = 16

    # ------------------------------------------------------------------
    # Row 3 — Headers
    # ------------------------------------------------------------------
    for col_idx, label in enumerate(_HEADER_LABELS, 1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = _THIN_BORDER
    ws.row_dimensions[3].height = 40

    # ------------------------------------------------------------------
    # Data rows (starting row 4)
    # ------------------------------------------------------------------
    col_fonts = {
        1: font_col1,
        2: None,  # per-group accent
        3: font_col3,
        4: font_col4,
        5: font_col5,
        6: font_col6,
        7: font_col7,
        8: font_col8,
        9: font_col9,
        10: font_col10,
    }

    for row_idx, item in enumerate(atomised, 4):
        group_name = item["threat_group"]
        bg_hex = _group_bg(group_name)

        # Alternate tint for even data rows (0-indexed: row_idx 4 is first data row)
        data_row_num = row_idx - 4
        if data_row_num % 2 == 1:
            bg_hex = _BG_ALT

        row_fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")

        values = [
            item["evidential_element"],
            group_name,
            item["procedure_example"],
            item["technique_id"],
            item["technique_name"],
            item["tactic"],
            item["contextual_evidence"],
            item["reference_url"],
            item["nav_layer_url"],
            item["source_type"],
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.alignment = align_left
            cell.border = _THIN_BORDER

            if col_idx == 2:
                accent = _group_accent(group_name)
                cell.font = Font(name="Calibri", size=10, bold=True, color=accent)
            else:
                cell.font = col_fonts[col_idx]

            # Hyperlinks for URL columns
            if col_idx == 8 and val and val.startswith("http"):
                cell.hyperlink = val
            if col_idx == 9 and val and val != "N/A" and val.startswith("http"):
                cell.hyperlink = val

        ws.row_dimensions[row_idx].height = 70

    # Freeze panes at A4
    ws.freeze_panes = "A4"

    # Auto-filter on row 3
    last_data_row = 3 + len(atomised)
    if atomised:
        ws.auto_filter.ref = f"A3:J{last_data_row}"

    # Apply border to merged cells in rows 1-2
    for col_idx in range(2, 11):
        for r in (1, 2):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = _THIN_BORDER

    # ------------------------------------------------------------------
    # Sheet 2 — Group Summary
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Group Summary")
    ws2.sheet_view.showGridLines = False

    gs_headers = [
        "Group Name", "Technique Count", "Indicator Count",
        "Tactic Coverage", "Top Tactic", "Invocation Coverage (%)",
    ]
    for col_idx, label in enumerate(gs_headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=label)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = _THIN_BORDER
    ws2.row_dimensions[3].height = 40

    # Column widths for Group Summary
    gs_widths = [30, 18, 18, 50, 25, 25]
    for i, w in enumerate(gs_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Compute group stats
    group_stats = {}
    for item in atomised:
        g = item["threat_group"]
        if g not in group_stats:
            group_stats[g] = {
                "techniques": set(),
                "indicators": 0,
                "tactics": {},
                "invocations_found": 0,
            }
        group_stats[g]["techniques"].add(item["technique_id"])
        group_stats[g]["indicators"] += 1
        tactic = item["tactic"]
        group_stats[g]["tactics"][tactic] = group_stats[g]["tactics"].get(tactic, 0) + 1
        if item.get("had_invocations", False):
            group_stats[g]["invocations_found"] += 1

    gs_row = 4
    for g_name, stats in sorted(group_stats.items()):
        tactics_sorted = sorted(stats["tactics"].items(), key=lambda x: x[1], reverse=True)
        top_tactic = tactics_sorted[0][0] if tactics_sorted else ""
        tactic_coverage = ", ".join(sorted(stats["tactics"].keys()))
        inv_pct = (stats["invocations_found"] / stats["indicators"] * 100) if stats["indicators"] else 0

        values = [
            g_name,
            len(stats["techniques"]),
            stats["indicators"],
            tactic_coverage,
            top_tactic,
            round(inv_pct, 1),
        ]
        bg = _group_bg(g_name)
        row_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=gs_row, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=10, color="E0F2FE")
            cell.fill = row_fill
            cell.alignment = align_left
            cell.border = _THIN_BORDER
            if col_idx == 6:
                cell.number_format = "0.0%"
                cell.value = inv_pct / 100  # Store as fraction for % format
        gs_row += 1

    # ------------------------------------------------------------------
    # Sheet 3 — Tactic Pivot
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("Tactic Pivot")
    ws3.sheet_view.showGridLines = False

    tp_headers = [
        "Tactic", "Indicator Count", "% of Total",
        "Invocations Found", "Example Technique IDs",
    ]
    for col_idx, label in enumerate(tp_headers, 1):
        cell = ws3.cell(row=3, column=col_idx, value=label)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = _THIN_BORDER
    ws3.row_dimensions[3].height = 40

    tp_widths = [30, 18, 14, 20, 55]
    for i, w in enumerate(tp_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Compute tactic stats
    tactic_stats = {}
    for item in atomised:
        t = item["tactic"]
        if t not in tactic_stats:
            tactic_stats[t] = {
                "count": 0,
                "invocations": 0,
                "techniques": set(),
            }
        tactic_stats[t]["count"] += 1
        tactic_stats[t]["techniques"].add(item["technique_id"])
        if item.get("had_invocations", False):
            tactic_stats[t]["invocations"] += 1

    total_count = sum(s["count"] for s in tactic_stats.values())
    tp_row = 4
    last_tp_row = tp_row + len(tactic_stats) - 1

    for tactic, stats in sorted(tactic_stats.items(), key=lambda x: x[1]["count"], reverse=True):
        example_tids = " | ".join(sorted(stats["techniques"])[:5])
        pct = stats["count"] / total_count if total_count else 0

        values = [
            tactic,
            stats["count"],
            pct,
            stats["invocations"],
            example_tids,
        ]
        bg = _BG_DARK if (tp_row - 4) % 2 == 0 else _BG_ALT
        row_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        for col_idx, val in enumerate(values, 1):
            cell = ws3.cell(row=tp_row, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=10, color="E0F2FE")
            cell.fill = row_fill
            cell.alignment = align_left
            cell.border = _THIN_BORDER
            if col_idx == 3:
                cell.number_format = "0.0%"
        tp_row += 1

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------
    wb.save(output_path)
