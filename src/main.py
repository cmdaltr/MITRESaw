#!/usr/bin/env python3 -tt
import os
import json
import random
import re
import requests
import subprocess
import sys
import time
from collections import Counter
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, List, Tuple, Set

from mitreattack.stix20 import MitreAttackData


class _ScrollWriter:
    """Stdout wrapper that erases the progress bar before any content is written,
    so the bar never gets duplicated in the scroll buffer."""

    def __init__(self, real_stdout, pb):
        self._real = real_stdout
        self._pb = pb

    def write(self, text):
        if not text:
            return 0
        # If the bar is currently drawn, erase it before writing content.
        # "\033[0J" erases from the cursor (which sits at the first bar row)
        # to the end of the screen, removing all bar lines in one shot.
        if self._pb._bar_drawn:
            self._real.write("\033[0J")
            self._pb._bar_drawn = False
        return self._real.write(text)

    def flush(self):
        self._real.flush()

    def fileno(self):
        return self._real.fileno()

    def isatty(self):
        return self._real.isatty()

    def __getattr__(self, name):
        return getattr(self._real, name)


class _ProgressBar:
    """Dual progress bar that follows the end of output.

    No scroll regions are used — the bar is drawn after the last content
    line, and _ScrollWriter erases it before the next content write.
    This avoids the multi-bar duplication caused by unreliable DECSTBM
    support across terminal emulators.
    """

    _ROWS = 6  # procedures + citations + separator + eta + elapsed + blank

    def __init__(self):
        self._start = None
        self._active = False
        self._bar_drawn = False   # True when the bar is currently on screen
        self._total_procs = 0
        self._total_cits = 0
        self._recent_times = []
        self._real_stdout = None

    def _bar(self, current, total, bar_width=60, color="\033[36m"):
        if total == 0:
            return "\033[90m" + "░" * bar_width + "\033[0m", "0.0%"
        pct = current / total
        filled = int(bar_width * pct)
        bar = color + "█" * filled + "\033[90m" + "░" * (bar_width - filled) + "\033[0m"
        return bar, f"{pct:.1%}"

    def _bar_done(self, total, bar_width=60, color="\033[32m"):
        return color + "█" * bar_width + "\033[0m"

    def _format_time(self, secs):
        if secs >= 3600:
            return f"{int(secs // 3600)}h {int((secs % 3600) // 60)}m {int(secs % 60)}s"
        elif secs >= 60:
            return f"{int(secs // 60)}m {int(secs % 60):02d}s"
        return f"{int(secs)}s"

    def _get_out(self):
        return self._real_stdout if self._real_stdout else sys.stdout

    def _setup(self):
        out = sys.stdout
        self._real_stdout = out
        sys.stdout = _ScrollWriter(out, self)
        self._active = True

    def _draw_bar(self, lines):
        """Draw bar lines at the current cursor position, then move cursor
        back up to the first bar row so _ScrollWriter knows where to erase."""
        out = self._get_out()
        try:
            tw = os.get_terminal_size().columns
        except OSError:
            tw = 120
        for line in lines:
            out.write(f"\r\033[2K{line[:tw]}\n")
        # Move cursor back up to the first bar row.
        # _ScrollWriter will erase from here with \033[0J before next write.
        out.write(f"\033[{len(lines)}A\r")
        out.flush()
        self._bar_drawn = True

    def update(
        self,
        proc_current,
        proc_total,
        cit_current,
        cit_total,
        group_name="",
        rate_limited=0,
        workers=0,
    ):
        if not self._active:
            self._total_procs = proc_total
            self._total_cits = cit_total
            self._setup()

        now = time.time()
        if self._start is None or proc_current <= 1:
            self._start = now
            self._recent_times = []
        secs = now - self._start

        self._recent_times.append(now)
        if len(self._recent_times) > 51:
            self._recent_times = self._recent_times[-51:]

        try:
            tw = os.get_terminal_size().columns
        except OSError:
            tw = 120

        bw = min(60, tw - 35)

        remaining = proc_total - proc_current
        if len(self._recent_times) >= 2:
            window = self._recent_times[-1] - self._recent_times[0]
            window_count = len(self._recent_times) - 1
            avg_per_proc = window / window_count if window_count > 0 else 0
            eta = avg_per_proc * remaining
            eta_str = self._format_time(eta)
        else:
            eta_str = "..."

        p_bar, p_pct = self._bar(proc_current, proc_total, bw, "\033[36m")
        c_bar, c_pct = (
            self._bar(cit_current, cit_total, bw, "\033[35m")
            if cit_total > 0
            else ("\033[90m" + "░" * bw + "\033[0m", "—")
        )

        _p_raw = f"{proc_current}/{proc_total}"
        _c_raw = f"{cit_current}/{cit_total}" if cit_total > 0 else f"{cit_current}"
        _count_w = max(len(_p_raw), len(_c_raw))
        _p_count = f"{_p_raw:>{_count_w}}"
        _c_count = f"{_c_raw:>{_count_w}}"

        _rl_str = (
            f"  \033[31m({rate_limited} rate-limited)\033[0m" if rate_limited else ""
        )
        _w_str = f"  \033[90m[{workers}w]\033[0m" if workers else ""

        _sep = "\033[90m" + "─" * (bw + _count_w + 23) + "\033[0m"
        self._draw_bar(
            [
                f"   Procedures: {p_bar}  {_p_count}  ({p_pct:>5})",
                (
                    f"   Citations:  {c_bar}  {_c_count}  ({c_pct:>5})"
                    if cit_total > 0
                    else f"   Citations:  {cit_current} collected"
                ),
                f"   {_sep}",
                f"   \033[1mETA:        {eta_str}\033[0m{_w_str}{_rl_str}",
                f"   \033[90mElapsed:    {self._format_time(secs)}\033[0m",
                "",
            ]
        )

    def done(self, proc_total, cit_total):
        out = self._get_out()
        try:
            tw = os.get_terminal_size().columns
        except OSError:
            tw = 120

        bw = min(60, tw - 35)
        p_bar = self._bar_done(proc_total, bw, "\033[32m")
        c_bar = self._bar_done(cit_total, bw, "\033[32m") if cit_total > 0 else None

        _p_raw = f"{proc_total}/{proc_total}"
        _c_raw = f"{cit_total}/{cit_total}" if cit_total > 0 else f"{cit_total}"
        _count_w = max(len(_p_raw), len(_c_raw))
        _p_count = f"{_p_raw:>{_count_w}}"
        _c_count = f"{_c_raw:>{_count_w}}"

        secs = time.time() - self._start if self._start else 0
        _cit_line = (
            f"   Citations:  {c_bar}  {_c_count}  (100.0%)"
            if c_bar
            else f"   Citations:  {cit_total} collected"
        )

        _sep = "\033[90m" + "─" * (bw + _count_w + 23) + "\033[0m"
        elapsed_str = self._format_time(secs)
        self._draw_bar(
            [
                f"   Procedures: {p_bar}  {_p_count}  (100.0%)",
                _cit_line,
                f"   {_sep}",
                f"   \033[1mCompleted in {elapsed_str}\033[0m",
                "",
                "",
            ]
        )
        time.sleep(0.5)

        # Move cursor to end of bar, then restore sys.stdout
        out.write(f"\033[{self._ROWS}B\r")
        out.flush()

        if self._real_stdout:
            sys.stdout = self._real_stdout
            self._real_stdout = None

        self._active = False
        self._bar_drawn = False

        # Permanent summary printed normally after bar
        print(f"\n   Procedures: {p_bar}  {_p_count}  (100.0%)")
        print(f"   {_cit_line.strip()}")
        print(f"   {_sep}")
        print(f"   \033[1mCompleted in {elapsed_str}\033[0m")


from stix2 import TAXIICollectionSource, Filter
from taxii2client.v20 import Server, Collection

import urllib3
import warnings
import yaml

import pandas

from src.extract import extract_indicators


_ssl_verify_failed = False  # Set True if SSL verification fails during STIX loading


def _fetch(url: str, **kwargs) -> requests.Response:
    """GET with automatic SSL-verify fallback for corporate VPN/proxy environments."""
    global _ssl_verify_failed
    if _ssl_verify_failed:
        kwargs.setdefault("verify", False)
        return requests.get(url, **kwargs)
    try:
        return requests.get(url, **kwargs)
    except requests.exceptions.SSLError:
        _ssl_verify_failed = True
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        print(
            "    -> SSL verification failed, switching to unverified mode for all requests"
        )
        return requests.get(url, verify=False, **kwargs)


from src.tools.write_csv import write_csv_summary
from src.tools.write_csv import write_csv_techniques_mapped_to_logsources
from src.output.matrix import build_matrix
from src.output.query import build_queries
from src.tools.keywords import match_keywords
from src.tools.read_files import collect_files
from src.tools.print_saw import print_saw


def get_latest_attack_version() -> str:
    """Fetch the latest MITRE ATT&CK version from STIX data."""
    try:
        # Use the TAXII server to get the latest version
        try:
            server = Server("https://cti-taxii.mitre.org/taxii/")
            api_root = server.api_roots[0]
            collections = api_root.collections
        except requests.exceptions.SSLError:
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
            server = Server("https://cti-taxii.mitre.org/taxii/", verify=False)
            api_root = server.api_roots[0]
            collections = api_root.collections

        for collection in collections:
            if "Enterprise ATT&CK" in collection.title:
                # Get collection source
                collection_source = TAXIICollectionSource(collection)
                # Get the x-mitre-collection object to find version
                filters = [Filter("type", "=", "x-mitre-collection")]
                collections_data = collection_source.query(filters)
                if collections_data:
                    version = collections_data[0].get("x_mitre_version", "Unknown")
                    return version

        # Fallback to web scraping if STIX method fails
        version_history = _fetch(
            "https://attack.mitre.org/resources/versions/", timeout=10
        )
        latest_version = re.findall(
            r"<span><strong>([^<]+)",
            str(version_history.content)
            .split("<h5><strong>Current Version</strong></h5>")[1]
            .split("<h5><strong>Most Recent Versions</strong></h5>")[0],
        )[0].split("ATT&amp;CK v")[1]
        return latest_version
    except Exception as e:
        print(f"\n\n\tWarning: Could not determine latest version: {e}")
        return "Unknown"


def build_technique_datasource_map(stix_filepath: str) -> Dict[str, str]:
    """Build a mapping from technique external ID to data source strings.

    Returns dict like {"T1001": "Network Traffic: Network Connection Creation, Process: Process Creation, ..."}
    """
    with open(stix_filepath, encoding="utf-8") as f:
        data = json.load(f)

    id_to_obj = {o["id"]: o for o in data["objects"]}

    # Build data component id -> name
    dc_names = {
        o["id"]: o.get("name", "")
        for o in data["objects"]
        if o.get("type") == "x-mitre-data-component"
    }

    # Build full "DataSource: DataComponent" names by matching DC names to DS names
    ds_names = sorted(
        [
            o.get("name", "")
            for o in data["objects"]
            if o.get("type") == "x-mitre-data-source"
        ],
        key=len,
        reverse=True,
    )
    # Manual fallback for DCs whose names don't start with their parent DS name
    dc_parent_override = {
        "Response Content": "Internet Scan",
        "Response Metadata": "Internet Scan",
        "Malware Content": "Malware Repository",
        "Malware Metadata": "Malware Repository",
        "Network Connection Creation": "Network Traffic",
        "Active DNS": "Domain Name",
        "Passive DNS": "Domain Name",
        "Domain Registration": "Domain Name",
        "Host Status": "Sensor Health",
        "Social Media": "Persona",
        "OS API Execution": "Process",
    }
    dc_full_names = {}
    for dc_id, dc_name in dc_names.items():
        if dc_name in dc_parent_override:
            dc_full_names[dc_id] = f"{dc_parent_override[dc_name]}: {dc_name}"
        else:
            for ds_name in ds_names:
                if dc_name.startswith(ds_name):
                    dc_full_names[dc_id] = f"{ds_name}: {dc_name}"
                    break
            else:
                dc_full_names[dc_id] = dc_name

    # Build detection_strategy -> data component full names (via analytics)
    ds_to_dcs: Dict[str, Set] = {}
    for ds_obj in [
        o for o in data["objects"] if o.get("type") == "x-mitre-detection-strategy"
    ]:
        dcs: Set[str] = set()
        for aref in ds_obj.get("x_mitre_analytic_refs", []):
            analytic = id_to_obj.get(aref)
            if analytic:
                for lref in analytic.get("x_mitre_log_source_references", []):
                    dc_ref = lref.get("x_mitre_data_component_ref", "")
                    if dc_ref in dc_full_names:
                        dcs.add(dc_full_names[dc_ref])
        ds_to_dcs[ds_obj["id"]] = dcs

    # Build technique ext_id -> data source string from detects relationships
    tech_ext_ids = {}
    for o in data["objects"]:
        if o.get("type") == "attack-pattern":
            ext_id = o.get("external_references", [{}])[0].get("external_id", "")
            tech_ext_ids[o["id"]] = ext_id

    tech_to_ds: Dict[str, Set] = {}
    for rel in data["objects"]:
        if (
            rel.get("type") == "relationship"
            and rel.get("relationship_type") == "detects"
        ):
            src = rel["source_ref"]
            tgt = rel["target_ref"]
            if src in ds_to_dcs and tgt in tech_ext_ids:
                ext_id = tech_ext_ids[tgt]
                if ext_id not in tech_to_ds:
                    tech_to_ds[ext_id] = set()
                tech_to_ds[ext_id].update(ds_to_dcs[src])

    return {tid: ", ".join(sorted(dcs)) for tid, dcs in tech_to_ds.items()}


def load_attack_data(
    framework: str = "enterprise", force_fetch: bool = False
) -> MitreAttackData:
    """Load MITRE ATT&CK data using STIX via the mitreattack-python library.

    Re-downloads if the cached file is older than 7 days or if force_fetch is True (--fetch).
    """
    print(f"    -> Loading {framework} ATT&CK data from STIX...")

    framework_map = {
        "enterprise": "enterprise-attack",
        "mobile": "mobile-attack",
        "ics": "ics-attack",
    }

    stix_source = framework_map.get(framework.lower(), "enterprise-attack")
    stix_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        "data",
        "stix",
    )
    os.makedirs(stix_dir, exist_ok=True)
    stix_filepath = os.path.join(stix_dir, f"{stix_source}.json")

    need_download = False
    if not os.path.exists(stix_filepath):
        need_download = True
    elif force_fetch:
        print(f"    -> --fetch flag set, forcing fresh download...")
        need_download = True
    else:
        file_age_days = (time.time() - os.path.getmtime(stix_filepath)) / 86400
        if file_age_days > 7:
            print(
                f"    -> STIX data is {int(file_age_days)} days old, re-downloading..."
            )
            need_download = True
        else:
            print(f"    -> Using cached STIX data ({int(file_age_days)} day(s) old)")

    if need_download:
        stix_url = f"https://raw.githubusercontent.com/mitre/cti/master/{stix_source}/{stix_source}.json"
        print(f"    -> Downloading {stix_source} STIX data...")
        resp = _fetch(stix_url, timeout=60)
        resp.raise_for_status()
        with open(stix_filepath, "w", encoding="utf-8") as f:
            f.write(resp.text)
        print(f"    -> Saved to {stix_filepath}")

    attack_data = MitreAttackData(stix_filepath)
    return attack_data, stix_filepath


def process_technique_parallel(args: Tuple) -> List[Dict]:
    """Process a single technique in parallel."""
    technique_entry, groups, platforms, attack_data = args
    results = []

    try:
        # get_techniques_used_by_group returns {'object': ..., 'relationships': [...]}
        technique = technique_entry.get("object", technique_entry)
        relationships = technique_entry.get("relationships", [])

        technique_id = technique.get("external_references", [{}])[0].get(
            "external_id", ""
        )
        technique_name = technique.get("name", "")
        technique_description = technique.get("description", "")

        # Get usage description from relationship (how the group uses the technique)
        usage = ""
        if relationships:
            usage = relationships[0].get("description", "")

        # Get platforms
        tech_platforms = technique.get("x_mitre_platforms", [])

        # Check if platform matches
        if platforms != ["."] and not any(p in tech_platforms for p in platforms):
            return results

        # Get tactics
        kill_chain_phases = technique.get("kill_chain_phases", [])
        tactics = [
            phase.get("phase_name", "").replace("-", " ").title()
            for phase in kill_chain_phases
        ]

        # Get data sources
        data_sources = []
        data_components = technique.get("x_mitre_data_sources", [])
        if isinstance(data_components, list):
            data_sources = data_components

        # Get detection info
        detection = technique.get("x_mitre_detection", "")

        result = {
            "id": technique_id,
            "name": technique_name,
            "description": technique_description,
            "usage": usage,
            "platforms": tech_platforms,
            "tactics": tactics,
            "data_sources": data_sources,
            "detection": detection,
            "stix_id": technique.get("id", ""),
        }
        results.append(result)

    except Exception as e:
        print(f"    Warning: Error processing technique: {e}")

    return results


def get_group_techniques_parallel(
    attack_data: MitreAttackData,
    groups: List[str],
    platforms: List[str],
    max_workers: int = 10,
) -> Tuple[Dict, Dict, List]:
    """Get techniques used by groups using parallel processing."""

    group_techniques = {}
    group_info = {}
    all_techniques = []

    # Get all groups
    all_groups = attack_data.get_groups(remove_revoked_deprecated=True)

    # Filter groups if specified
    if groups != ["."]:
        filtered_groups = []
        for group in all_groups:
            group_name = group.get("name", "")
            # Guard against None — STIX objects may return None for optional fields
            group_aliases = list(group.get("aliases") or [])
            # All candidate strings to match against (name + all aliases), lowercased
            _candidates = [n.strip().lower() for n in [group_name] + group_aliases if n]
            # Match if any user-supplied term is a substring of any candidate name/alias.
            # Substring matching lets users enter partial names ("Cozy Bear" → APT29,
            # "APT" → all APT groups) without requiring exact full-name spelling.
            if any(
                g.replace("_", " ").strip().lower() in candidate
                for g in groups
                for candidate in _candidates
            ):
                filtered_groups.append(group)
        all_groups = filtered_groups

    # Process each group
    for group in all_groups:
        try:
            group_name = group.get("name", "")
            group_id = group.get("external_references", [{}])[0].get("external_id", "")
            group_description = group.get("description", "")

            group_info[group_id] = {
                "name": group_name,
                "description": group_description,
                "aliases": group.get("aliases", []),
            }

            # Get techniques used by this group
            techniques = attack_data.get_techniques_used_by_group(group.get("id"))

            # Also get techniques from campaigns attributed to this group
            campaign_techniques = []
            try:
                campaigns = attack_data.get_campaigns_attributed_to_group(
                    group.get("id")
                )
                if campaigns:
                    seen_technique_ids = set()
                    for t in techniques:
                        tobj = t.get("object", t)
                        tid = tobj.get("external_references", [{}])[0].get(
                            "external_id", ""
                        )
                        seen_technique_ids.add(tid)
                    for campaign_entry in campaigns:
                        campaign_obj = campaign_entry.get("object", campaign_entry)
                        campaign_name = campaign_obj.get("name", "")
                        campaign_id = campaign_obj.get("id", "")
                        try:
                            camp_techs = attack_data.get_techniques_used_by_campaign(
                                campaign_id
                            )
                            for ct in camp_techs:
                                ct_obj = ct.get("object", ct)
                                ct_id = ct_obj.get("external_references", [{}])[0].get(
                                    "external_id", ""
                                )
                                if ct_id not in seen_technique_ids:
                                    seen_technique_ids.add(ct_id)
                                    # Prepend campaign name to usage for context
                                    ct_rels = ct.get("relationships", [])
                                    if ct_rels:
                                        orig_desc = ct_rels[0].get("description", "")
                                        ct_rels[0][
                                            "description"
                                        ] = f"[Campaign: {campaign_name}] {orig_desc}"
                                    campaign_techniques.append(ct)
                        except Exception:
                            pass
            except Exception:
                pass

            all_group_techniques = list(techniques) + campaign_techniques
            group_techniques[group_id] = []

            # Process techniques in parallel
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = []
                for technique in all_group_techniques:
                    futures.append(
                        executor.submit(
                            process_technique_parallel,
                            (technique, groups, platforms, attack_data),
                        )
                    )

                for future in as_completed(futures):
                    try:
                        results = future.result()
                        for result in results:
                            group_techniques[group_id].append(result)
                            all_techniques.append(result)
                    except Exception as e:
                        print(f"    Warning: Thread error: {e}")

        except Exception as e:
            print(
                f"    Warning: Error processing group {group.get('name', 'Unknown')}: {e}"
            )
            continue

    return group_techniques, group_info, all_techniques


def replace_commas_in_group_desc(csv_line):
    return re.sub(
        r"^((?:[^,]+,){6}.*,relationship--[^,]+,\d{2} [A-Za-z]{3,20} \d{4},\d{2} [A-Za-z]{3,20} \d{4}[^\[]+[^\(]+\([^\)]+\)[^,]+), ",
        r"\1%2C ",
        csv_line,
    )


def _write_reference_sheet(xlsx_path, all_refs):
    """Append a Reference Detail sheet to an existing XLSX."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(xlsx_path)
    ws = wb.create_sheet("Reference Detail")
    ws.sheet_view.showGridLines = False

    border = Border(
        left=Side(style="thin", color="1E3A5F"),
        right=Side(style="thin", color="1E3A5F"),
        top=Side(style="thin", color="1E3A5F"),
        bottom=Side(style="thin", color="1E3A5F"),
    )
    fill_navy = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    font_header = Font(name="Calibri", size=12, bold=True, color="E0F2FE")
    font_data = Font(name="Calibri", size=10, color="CBD5E1")
    font_url = Font(name="Calibri", size=10, color="0EA5E9")
    align_wrap = Alignment(wrap_text=True, vertical="center", horizontal="left")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Citation Name",
        "Source URL",
        "Source Description",
        "Extracted Content",
        "Collection Method",
        "Attempts",
    ]
    widths = [30, 55, 45, 80, 18, 40]

    font_method = Font(name="Courier New", size=9, color="4ADE80")
    font_attempts = Font(name="Courier New", size=9, color="94A3B8")

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 32

    for ri, ref in enumerate(all_refs, 2):
        bg = "0F1C2E" if ri % 2 == 0 else "0A1220"
        row_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        attempts_list = ref.get("attempts", [])
        attempts_str = (
            " → ".join(attempts_list) if attempts_list else ref.get("status", "")
        )
        values = [
            ref.get("citation_name", ""),
            ref.get("url", ""),
            ref.get("description", ""),
            ref.get("extracted_content", ""),
            ref.get("method", ""),
            attempts_str,
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ci == 2:
                cell.font = font_url
            elif ci == 5:
                cell.font = font_method
            elif ci == 6:
                cell.font = font_attempts
            else:
                cell.font = font_data
            cell.fill = row_fill
            cell.alignment = align_wrap
            cell.border = border
            if ci == 2 and val and val.startswith("http"):
                cell.hyperlink = val
        ws.row_dimensions[ri].height = 90

    ws.freeze_panes = "A2"
    if all_refs:
        ws.auto_filter.ref = f"A1:F{1 + len(all_refs)}"

    wb.save(xlsx_path)


def mainsaw(
    operating_platforms,
    search_terms,
    provided_groups,
    show_others,
    art,
    navigationlayers,
    queries,
    truncate,
    attack_frameworks,
    attack_version,
    sheet_tabs,
    columns=None,
    preset=False,
    export_format="csv",
    quiet=False,
    fetch=False,
    evidence_report=False,
    collect_citations=False,
    citation_workers=10,
    auto_confirm=False,
    dry_run=False,
):

    # Load STIX data (only check version if -F is used)
    try:
        if fetch:
            print("    -> Checking for latest ATT&CK version...")
            latest_version = get_latest_attack_version()
            if latest_version != "Unknown" and latest_version != attack_version:
                print(f"    -> Updating from v{attack_version} to v{latest_version}")
                attack_version = latest_version

        # Load STIX data for all requested frameworks
        all_attack_data = {}
        technique_datasource_map = {}
        for fw in attack_frameworks:
            attack_data, stix_filepath = load_attack_data(fw, force_fetch=fetch)
            all_attack_data[fw] = attack_data
            technique_datasource_map.update(
                build_technique_datasource_map(stix_filepath)
            )

    except requests.exceptions.ConnectionError:
        print("\n\n\tUnable to connect to the Internet. Please try again.\n\n\n")
        sys.exit()
    except Exception as e:
        print(f"\n\n\tError loading ATT&CK data: {e}\n\n\n")
        sys.exit()

    # Setup output directories
    frameworks_label = ", ".join(attack_frameworks)
    frameworks_slug = "-".join(fw.lower() for fw in attack_frameworks)
    mitresaw_root_date = os.path.join("data", str(datetime.now())[0:10])
    if not os.path.exists(mitresaw_root_date):
        os.makedirs(mitresaw_root_date)
    mitre_files = os.path.join(
        mitresaw_root_date, "{}-{}-stix".format(frameworks_slug, attack_version)
    )
    if not os.path.exists(mitre_files):
        os.makedirs(mitre_files)

    print(f"    -> Using STIX data from TAXII server (cached locally)...")

    time.sleep(0.1)
    saw = """
@                                                         ,
@                 ╓╗╗,                          ,╓▄▄▄Φ▓▓██▌╫D
@                ║▌ `▓L            ,,, ╓▄▄▄Φ▓▓▀▀▀╫╫╫╫╫╫╫▀▀╫▓▓▄
@                 ▓▄▓▓▓        ,▄▄B░▀╫Ñ╬░░╫╫▓▓▓▓╫╫╫╫▓▓▓╫╫╫╫╣▓▓▓▄
@                 ║████L   ,╓#▀▀▀╨╫ÑÑ╦▄▒▀╣▓▄▄▀╣▌╫▀    ██╫╫╫╫▓▓╫▓▓φ
@                  ▓╫╫╫▀]Ñ░░░░ÑÑÑÑ░░░░░╠▀W▄╠▀▓▒░╫Ñ╖   ╙└"╜▀▓▓▓▓▓█▓▓
@                  ║░░░╦╬╫╫╫╫╫╫╫╫╫╫╫╫╫ÑÑ░░░╠Ñ░╨╫Ñ░╫╫╫╫N     ▀▓▓▓╫██▓╕
@                ,]░╦╬╫╫╫╫╫╫╫▓▓▓▓▓▓╫╫╫╫╫╫╫Ñ░░╠░░╫M░╠╫╫╫╫╦,    ▀▓▓▓▓▓▓⌐
@       ╗▄╦     ]░░╬╫╫╫╫╫▓▓██████████▓▓▒╫╫╫╫Ñ░░╟▒╟▓▒ñ▓▓▓▓░N    ╙▓▓▓▓▓▓
@   ║███╫█╫    ]░░╫╫╫╫╫▓███▓▓▓▓▓▓▓▓▓▓███▓╫╫╫╫╫░░╟▒╟▓Ü╟▓▓▓▓░H    ╟▓▓▓▓▓L
@   ║███╫█╫   ]░░╫╫╫╫▓██▓╫▓▓▓▀▀╠╠╬▀▓▓▓╫▓██▓╫╫╫╫░░ÑÑ╠▄░╠▓▓▓▄▄▄▄▄▓▓▓╫╫╫╫
@    ╓▄▄╫█╫╖╖╖╦░╫╫╫╫╫██▓▓▓▓▀░╬Ñ╣╬╫Ñ░╟▓▓▓▓██╫╫╫╫Ñ░╦]░░░║████▀▀╫╫╫▓╩╨╟╫
@    ╟▓▓╫█╫▀▀▀╩╬╩╫╫▓██▓▓▓▓▌░╫░╟▓▓K╫Ñ░▓▓▓▓╫██▓▒╩╩╩╩ ╙╩╨▀▓M╨╩╨╙╝╣N╦╗Φ╝
@       ╫█╫     ▀███▀╣▓▓▓▓▓░╫Ñ░╠▀░╫Ü░▓▓▓▓▓▀▀███╕      ▐▓▌╖
@   ▄▄▄▄▓█▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄▄╛
@                ▀╩╫╫╫╠╣▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▀░╫╫╫╫▌
@                 ╗▄╫╫Ñ░╠▀▓▓▓▓▓▓▓▓▓▓▓▓▀░╦╬╫╫∩
@                   `⌠╫╫╫Ñ░░Å╣▀▀▀▀▀▒░╦╬╫╫╫`█
@                    ╙╙""╫╫╫½╫╫╫╬╫╫╫╫╫M"▓╛
@                       └╙└ ▄▓╩`║▓╩ Å▀\n\n
    """
    titles = [
        """
       ███▄ ▄███▓ ██▓▄▄▄█████▓ ██▀███  ▓█████   ██████  ▄▄▄       █     █░
      ▓██▒▀█▀ ██▒▓██▒▓  ██▒ ▓▒▓██ ▒ ██▒▓█   ▀ ▒██    ▒ ▒████▄    ▓█░ █ ░█░
      ▓██    ▓██░▒██▒▒ ▓██░ ▒░▓██ ░▄█ ▒▒███   ░ ▓██▄   ▒██  ▀█▄  ▒█░ █ ░█ 
      ▒██    ▒██ ░██░░ ▓██▓ ░ ▒██▀▀█▄  ▒▓█  ▄   ▒   ██▒░██▄▄▄▄██ ░█░ █ ░█ 
      ▒██▒   ░██▒░██░  ▒██▒ ░ ░██▓ ▒██▒░▒████▒▒██████▒▒ ▓█   ▓██▒░░██▒██▓ 
      ░ ▒░   ░  ░░▓    ▒ ░░   ░ ▒▓ ░▒▓░░░ ▒░ ░▒ ▒▓▒ ▒ ░ ▒▒   ▓▒█░░ ▓░▒ ▒  
      ░  ░      ░ ▒ ░    ░      ░▒ ░ ▒░ ░ ░  ░░ ░▒  ░ ░  ▒   ▒▒ ░  ▒ ░ ░  
      ░      ░    ▒ ░  ░        ░░   ░    ░   ░  ░  ░    ░   ▒     ░   ░  
             ░    ░              ░        ░  ░      ░        ░  ░    ░    
""",
        """
      ______  ___________________________ __________________                   
      ___   |/  /____  _/___  __/___  __ \\___  ____/__  ___/______ ____      __
      __  /|_/ /  __  /  __  /   __  /_/ /__  __/   _____ \\ _  __ `/__ | /| / /
      _  /  / /  __/ /   _  /    _  _, _/ _  /___   ____/ / / /_/ / __ |/ |/ / 
      /_/  /_/   /___/   /_/     /_/ |_|  /_____/   /____/  \\__,_/  ____/|__/  
""",
        """
                   ______      ______    ____        ____       ____                              
       /'\\_/`\\    /\\__  _\\    /\\__  _\\  /\\  _`\\     /\\  _`\\    /\\  _`\\                            
      /\\      \\   \\/_/\\ \\/    \\/_/\\ \\/  \\ \\ \\L\\ \\   \\ \\ \\L\\_\\  \\ \\,\\L\\_\\      __      __  __  __  
      \\ \\ \\__\\ \\     \\ \\ \\       \\ \\ \\   \\ \\ ,  /    \\ \\  _\\L   \\/_\\__ \\    /'__`\\   /\\ \\/\\ \\/\\ \\ 
       \\ \\ \\_/\\ \\     \\_\\ \\__     \\ \\ \\   \\ \\ \\\\ \\    \\ \\ \\L\\ \\   /\\ \\L\\ \\ /\\ \\L\\.\\_ \\ \\ \\_/ \\_/ \\
        \\ \\_\\\\ \\_\\    /\\_____\\     \\ \\_\\   \\ \\_\\ \\_\\   \\ \\____/   \\ `\\____\\\\ \\__/.\\_\\ \\ \\___x___/'
         \\/_/ \\/_/    \\/_____/      \\/_/    \\/_/\\/ /    \\/___/     \\/_____/ \\/__/\\/_/  \\/__//__/  
""",
        """
         _____   .___ _____________________ ___________  _________                 
        /     \\  |   |\\__    ___/\\______   \\\\_   _____/ /   _____/_____   __  _  __
       /  \\ /  \\ |   |  |    |    |       _/ |    __)_  \\_____  \\ \\__  \\  \\ \\/ \\/ /
      /    Y    \\|   |  |    |    |    |   \\ |        \\ /        \\ / __ \\_ \\     / 
      \\____|__  /|___|  |____|    |____|_  //_______  //_______  /(____  /  \\/\\_/  
              \\/                         \\/         \\/         \\/      \\/          
""",
        """
        ___ ___   ___   _______   _______   _______   _______                    
       |   Y   | |   | |       | |   _   \\ |   _   | |   _   | .---.-. .--.--.--.
       |.      | |.  | |.|   | | |.  l   / |.  1___| |   1___| |  _  | |  |  |  |
       |. \\_/  | |.  | `-|.  |-' |.  _   1 |.  __)_  |____   | |___._| |________|
       |:  |   | |:  |   |:  |   |:  |   | |:  1   | |:  1   |                   
       |::.|:. | |::.|   |::.|   |::.|:. | |::.. . | |::.. . |                   
       `--- ---' `---'   `---'   `--- ---' `-------' `-------'                   
""",
    ]
    chosen_title = random.choice(titles)
    tagline = "\n\n\n\n{}        *ATT&CK for {} v{}\n".format(
        chosen_title, frameworks_label, attack_version
    )
    time.sleep(1)
    subprocess.Popen(["clear"]).communicate()
    if art:
        if saw:
            print_saw(saw, tagline, "                                        ")
            print_saw(saw, tagline, "                                      ")
            print_saw(saw, tagline, "                                    ")
            print_saw(saw, tagline, "                                  ")
            print_saw(saw, tagline, "                                ")
            print_saw(saw, tagline, "                              ")
            print_saw(saw, tagline, "                            ")
    platforms = str(operating_platforms)[2:-2].split(",")
    platforms = list(filter(None, platforms))
    if art:
        print_saw(saw, tagline, "                          ")
    terms = str(search_terms)[2:-2].split(",")
    terms = list(filter(None, terms))
    if art:
        print_saw(saw, tagline, "                        ")
    groups = str(provided_groups)[2:-2].split(",")
    groups = list(filter(None, groups))
    if art:
        print_saw(saw, tagline, "                      ")
    else:
        print(tagline)
    if True:  # creating MITRESaw output file names
        if str(platforms) == "['.']":
            platforms_filename_insert = ""
        else:
            platforms_filename_insert = "{}".format(
                str(platforms)[2:-2].replace("', '", "-")
            )
        if str(terms) == "['.']":
            terms_filename_insert = ""
        else:
            terms_filename_insert = "{}".format(str(terms)[2:-2].replace("', '", "-"))
        if str(groups) == "['.']":
            groups_filename_insert = ""
            groups_insert = "Threat Actors"
            all_insert = "all "
        else:
            groups_filename_insert = "{}".format(str(groups)[2:-2].replace("', '", "-"))
            groups_insert = "{}".format(str(groups)[2:-2])
            all_insert = ""
        mitresaw_output_directory = os.path.join(
            mitresaw_root_date,
            "{}_{}_{}".format(
                platforms_filename_insert.replace("_", ""),
                terms_filename_insert.replace("_", ""),
                groups_filename_insert.replace("_", ""),
            ),
        )
        if not os.path.exists(os.path.join(mitresaw_output_directory)):
            os.makedirs(os.path.join(mitresaw_output_directory))
    (
        additional_terms,
        evidence_found,
        valid_procedures,
        all_evidence,
        log_sources,
        logsources,
        groups_in_scope,
        techniques_in_scope,
        groups_techniques_in_scope,
    ) = ([] for _ in range(9))
    (
        group_procedures,
        group_descriptions,
        contextual_information,
        previous_findings,
    ) = ({} for _ in range(4))
    if art:
        if saw:
            print_saw(saw, tagline, "                    ")
            print_saw(saw, tagline, "                  ")
            print_saw(saw, tagline, "                ")
            print_saw(saw, tagline, "              ")
            print_saw(saw, tagline, "            ")
            print_saw(saw, tagline, "          ")
            print_saw(saw, tagline, "        ")
            print_saw(saw, tagline, "      ")
            print_saw(saw, tagline, "    ")
            print_saw(saw, tagline, "  ")
            print_saw(saw, tagline, "partial")
            print_saw(saw, tagline, "-")  # remove saw
            print()
    if str(terms) != "['.']":
        terms_insert = " associated with '\033[1;36m{}\033[1;m'".format(
            str(terms)[2:-2].replace("_", " ").replace("', '", "\033[1;m', '\033[1;36m")
        )
    else:
        terms_insert = ""
    # Use STIX-based parallel processing instead of CSV files
    print()
    print(
        "    -> Extracting \033[1;33mIdentifiers\033[1;m from \033[1;32mTechniques\033[1;m using STIX data based on {}\033[1;31m{}\033[1;m{}".format(
            all_insert,
            groups_insert.replace("', '", "\033[1;m, \033[1;31m"),
            terms_insert,
        )
    )

    # Get group techniques using parallel processing across all frameworks
    all_group_techniques_data = {}
    all_group_info_data = {}
    for fw, attack_data in all_attack_data.items():
        group_techniques_data, group_info_data, _ = get_group_techniques_parallel(
            attack_data, groups, platforms, max_workers=citation_workers
        )
        for gid, techs in group_techniques_data.items():
            for tech in techs:
                tech["framework"] = fw
            all_group_techniques_data.setdefault(gid, []).extend(techs)
        all_group_info_data.update(group_info_data)
    group_techniques_data = all_group_techniques_data
    group_info_data = all_group_info_data

    # Process the STIX data into the format expected by the rest of the tool
    contextual_information = []
    for group_id, techniques in group_techniques_data.items():
        group_name = group_info_data[group_id]["name"]
        group_description = group_info_data[group_id]["description"]

        for technique in techniques:
            technique_id = technique["id"]
            technique_name = technique["name"]
            technique_description = technique["description"]
            technique_usage = technique.get("usage", "")
            technique_platforms = ", ".join(technique["platforms"])
            technique_tactics = ", ".join(technique["tactics"])
            technique_detection = technique.get("detection", "") or ""
            technique_data_sources = technique_datasource_map.get(technique_id, "")
            technique_framework = technique.get("framework", "")

            # Build context string in the format expected by the rest of the tool
            # Format: group_id||group_name||technique_id||technique_name
            context = f"{group_id}||{group_name}||{technique_id}||{technique_name}"
            contextual_information.append(context)

            # obtaining navigation layers for all identified threat groups
            if navigationlayers:
                navlayer_output_directory = os.path.join(
                    mitresaw_root_date,
                    "{}_navigationlayers".format(str(datetime.now())[0:10]),
                )
                for fw in attack_frameworks:
                    domain = f"{fw.lower()}-attack"
                    navlayer_json = os.path.join(
                        navlayer_output_directory,
                        f"{group_id}_{group_name}-{domain}-layer.json",
                    )
                    if not os.path.exists(navlayer_json):
                        if not os.path.exists(navlayer_output_directory):
                            os.makedirs(navlayer_output_directory)
                            print(
                                "     -> Obtaining ATT&CK Navigator Layers for \033[1;33mThreat Actors\033[1;m related to identified \033[1;32mTechniques\033[1;m..."
                            )
                        try:
                            group_navlayer = _fetch(
                                f"https://attack.mitre.org/groups/{group_id}/{group_id}-{domain}-layer.json",
                                timeout=10,
                            )
                            if group_navlayer.status_code == 200:
                                with open(navlayer_json, "wb") as navlayer_file:
                                    navlayer_file.write(group_navlayer.content)
                        except Exception as e:
                            print(
                                f"    Warning: Could not download nav layer for {group_name} ({fw}): {e}"
                            )

            # Build valid procedure
            # Format expected by extract.py:
            # [0]group_id || [1]group_name || [2]technique_id || [3]technique_name ||
            # [4]usage(relationship desc) || [5]- || [6]group_description(terms) ||
            # [7]technique_description || [8]technique_detection ||
            # [9]technique_platforms || [10]technique_data_sources ||
            # [11]technique_tactics || [12]framework
            # extract.py appends [13]evidence_dict (JSON)
            # Sanitize free-text fields to avoid corrupting the || delimiter
            _usage = technique_usage.replace("||", " ")
            _gdesc = group_description.replace("||", " ")
            _tdesc = technique_description.replace("||", " ")
            _tdet = technique_detection.replace("||", " ")
            _ttactics = technique_tactics.replace("||", " ")
            valid_procedure = f"{group_id}||{group_name}||{technique_id}||{technique_name}||{_usage}||-||{_gdesc}||{_tdesc}||{_tdet}||{technique_platforms}||{technique_data_sources}||{_ttactics}||{technique_framework}"

            # Apply string filter: skip procedures that don't mention any search term.
            # Checked against: group name, group description, technique name, usage,
            # technique description, and technique detection.
            # terms == ['.'] means "no filter" — include everything.
            if str(terms) != "['.']" and terms:
                _proc_text = " ".join(
                    [
                        group_name,
                        _gdesc,
                        technique_id,
                        technique_name,
                        _usage,
                        _tdesc,
                        _tdet,
                    ]
                ).lower()
                if not any(t.strip().lower() in _proc_text for t in terms):
                    continue

            valid_procedures.append(valid_procedure)

            # Track techniques
            techniques_in_scope.append(f"{technique_id}||{technique_name}")
            groups_techniques_in_scope.append(
                f"{group_name}||{technique_id}||{technique_name}||{technique_tactics}"
            )
            groups_in_scope.append(group_name)
    consolidated_procedures = sorted(list(set(valid_procedures)))
    counted_techniques = Counter(techniques_in_scope)
    sorted_techniques = sorted(
        counted_techniques.items(), key=lambda x: x[1], reverse=True
    )
    sorted_threat_actors_techniques_in_scope = list(set(groups_techniques_in_scope))
    technique_combos = []
    for technique in counted_techniques.most_common():
        technique_count = technique[1]
        if ": " in technique[0]:
            parent_technique = technique[0].split(": ")[0]
            sub_technique = technique[0].split(": ")[1]
        else:
            parent_technique = technique[0]
            sub_technique = "-"
        technique_combo = [parent_technique, sub_technique, technique_count]
        technique_combos.append(technique_combo)
    # Build STIX citation lookup if -R is enabled (before extraction loop)
    # Build citation lookup: collect ALL external_references from STIX relationships
    # Keyed by citation source_name → {url, description} for direct lookup
    _citation_url_lookup = {}  # source_name → {"url": ..., "description": ...}
    _mitre_ref_numbers = {}  # (group_name_lower, source_name) → MITRE [N] number
    _all_citation_refs = []
    _seen_citations = (
        set()
    )  # (group, citation_name) — prevents same group re-fetching same URL
    _seen_global_urls = (
        set()
    )  # URL — prevents same URL appearing in _all_citation_refs twice
    if collect_citations:
        for _fw, _ad in all_attack_data.items():
            _sp = getattr(_ad, "stix_filepath", None) or getattr(_ad, "src", None)
            if not _sp:
                continue
            try:
                import json as _json

                with open(_sp) as _f:
                    _bundle = _json.load(_f)
                # Build citation URL lookup from ALL STIX objects
                # (relationships, techniques, groups, software, campaigns)
                for _obj in _bundle.get("objects", []):
                    for _ref in _obj.get("external_references", []):
                        _sn = _ref.get("source_name", "")
                        if (
                            _sn
                            and _sn != "mitre-attack"
                            and _sn not in _citation_url_lookup
                        ):
                            _citation_url_lookup[_sn] = {
                                "url": _ref.get("url", ""),
                                "description": _ref.get("description", ""),
                            }
                # Build MITRE reference numbers from group objects
                for _obj in _bundle.get("objects", []):
                    if _obj.get("type") != "intrusion-set":
                        continue
                    _gname = _obj.get("name", "").strip().lower()
                    _num = 0
                    for _ref in _obj.get("external_references", []):
                        _sn = _ref.get("source_name", "")
                        if _sn == "mitre-attack":
                            continue
                        _num += 1
                        _mitre_ref_numbers[(_gname, _sn)] = _num
            except Exception:
                continue
        if _citation_url_lookup:
            print(
                f"    -> {len(_citation_url_lookup)} unique citation sources indexed for collection\n"
            )
            # Propagate SSL verification state to citation collector
            if _ssl_verify_failed:
                import src.citation_collector as _cc

                _cc.SSL_VERIFY = False

    # Sort by (group, technique_name) so procedures for the same group+technique are contiguous
    consolidated_procedures = sorted(
        consolidated_procedures,
        key=lambda p: (
            p.split("||")[1].strip().lower(),
            p.split("||")[3].strip().lower() if len(p.split("||")) > 3 else "",
        ),
    )

    # Pre-count unique (group, citation) pairs for progress bar
    _total_cit_pairs = 0
    if collect_citations and _citation_url_lookup:
        _pre_seen = set()
        for _p in consolidated_procedures:
            _pp = _p.split("||")
            _pg = _pp[1].strip().lower() if len(_pp) > 1 else ""
            _all_text = _pp[4] if len(_pp) > 4 else ""
            if len(_pp) > 7:
                _all_text += " " + _pp[7]
            if len(_pp) > 8:
                _all_text += " " + _pp[8]
            for _cn in re.findall(r"\(Citation:\s*([^)]+)\)", _all_text):
                _key = (_pg, _cn.strip())
                if _key not in _pre_seen:
                    _pre_seen.add(_key)
                    _total_cit_pairs += 1

    last_group_name = None
    _total_procedures = len(consolidated_procedures)

    # Pre-run ETA estimate, confirmation, and --dry-run scope preview
    _sep_pf = "\033[90m" + "─" * 45 + "\033[0m"
    _cached_count = 0
    _uncached_count = 0

    if collect_citations and _total_cit_pairs > 0:
        from src.citation_collector import _cache_key, CACHE_DIR

        _checked_urls = set()
        for _p in consolidated_procedures:
            _pp = _p.split("||")
            _all_text = _pp[4] if len(_pp) > 4 else ""
            if len(_pp) > 7:
                _all_text += " " + _pp[7]
            if len(_pp) > 8:
                _all_text += " " + _pp[8]
            for _cn in re.findall(r"\(Citation:\s*([^)]+)\)", _all_text):
                _cn = _cn.strip()
                _ref_data = _citation_url_lookup.get(_cn, {})
                _url = _ref_data.get("url", "")
                if not _url or _url in _checked_urls:
                    continue
                _checked_urls.add(_url)
                _cpath = CACHE_DIR / f"{_cache_key(_url)}.json"
                if _cpath.exists():
                    _cached_count += 1
                else:
                    _uncached_count += 1

    # Estimate total time
    _est_cached_s = _cached_count * 0.002
    _est_uncached_s = _uncached_count * 2.0 / max(1, citation_workers)
    _est_proc_s = _total_procedures * 0.005
    _est_total = _est_cached_s + _est_uncached_s + _est_proc_s
    if _est_total >= 3600:
        _est_str = f"{int(_est_total // 3600)}h {int((_est_total % 3600) // 60)}m"
    elif _est_total >= 60:
        _est_str = f"{int(_est_total // 60)}m {int(_est_total % 60):02d}s"
    elif _est_total >= 1:
        _est_str = f"{int(_est_total)}s"
    else:
        _est_str = "< 1s"

    _matched_groups = len(set(groups_in_scope))

    # Build flags summary for dry-run
    _active_flags = []
    if evidence_report:
        _active_flags.append("-E  Evidence Report (XLSX)")
    if collect_citations:
        _active_flags.append("-C  Citation enrichment")
    if columns and not preset:
        _active_flags.append(f"-c  Columns: {columns[:60]}")

    if dry_run or (collect_citations and _total_cit_pairs > 0):
        _plan_label = "Dry-run scope preview" if dry_run else "Pre-fetch plan"
        print(f"    \033[1m{_plan_label}\033[0m")
        print(f"    {_sep_pf}")
        print(f"    🌐 Framework:      {'  '.join(attack_frameworks)}")
        _groups_filtered = [g.strip() for g in groups if g.strip() and g.strip() != "."]
        if _groups_filtered:
            _groups_display = ", ".join(g.replace("_", " ") for g in _groups_filtered)
            print(f"    👥  Groups:      {_matched_groups:>6,}  matched  ({_groups_display})")
        else:
            print(f"    👥  Groups:      {_matched_groups:>6,}  matched")
        print(f"    🩻  Procedures:  {_total_procedures:>6,}")
        if collect_citations:
            _total_cit = _cached_count + _uncached_count
            print(f"    ✍️  Citations:   {_total_cit:>6,}")
            if _uncached_count:
                print(f"        🔍 {_uncached_count:,} to fetch")
            if _cached_count:
                print(f"        💾 {_cached_count:,} cached")
            print(f"    👷  Workers:    {citation_workers:>6,}")
        if _active_flags:
            print(f"    🚩  Flags:")
            for _fl in _active_flags:
                print(f"        {_fl}")
        if collect_citations:
            print(f"    🕰️  Est. time:  {_est_str:>7}")
        print(f"    {_sep_pf}")

        if dry_run:
            print()
            return

        if not auto_confirm:
            try:
                _resp = input("    Continue? [Y/n] ").strip().lower()
                if _resp in ("n", "no"):
                    print("\n    Aborted.\n")
                    return
            except (EOFError, KeyboardInterrupt):
                print("\n    Aborted.\n")
                return
            print()

    # Pre-fetch all uncached citations in one parallel batch
    if collect_citations and _citation_url_lookup:
        from src.citation_collector import _cache_key, _read_cache, CACHE_DIR
        from src.citation_collector import collect_references_parallel

        _prefetch_batch = []
        _prefetch_seen_urls = set()
        for _p in consolidated_procedures:
            _pp = _p.split("||")
            _all_text = _pp[4] if len(_pp) > 4 else ""
            if len(_pp) > 7:
                _all_text += " " + _pp[7]
            if len(_pp) > 8:
                _all_text += " " + _pp[8]
            for _cn in re.findall(r"\(Citation:\s*([^)]+)\)", _all_text):
                _cn = _cn.strip()
                _ref_data = _citation_url_lookup.get(_cn, {})
                _url = _ref_data.get("url", "")
                if not _url or _url in _prefetch_seen_urls:
                    continue
                _prefetch_seen_urls.add(_url)
                _cpath = CACHE_DIR / f"{_cache_key(_url)}.json"
                if not _cpath.exists():
                    _prefetch_batch.append(
                        {
                            "citation_name": _cn,
                            "url": _url,
                            "description": _ref_data.get("description", ""),
                        }
                    )

        if _prefetch_batch:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            from src.citation_collector import collect_reference_content

            _pf_total = len(_prefetch_batch)
            _pf_done = 0
            _pf_ok = 0
            _pf_start = time.time()
            _pf_results = []

            print(
                f"   -> Pre-fetching {_pf_total:,} uncached citations with {citation_workers} workers..."
            )

            def _pf_fetch(cit):
                return collect_reference_content([cit], "", "")

            with ThreadPoolExecutor(max_workers=citation_workers) as _pf_pool:
                _pf_futures = {
                    _pf_pool.submit(_pf_fetch, c): c for c in _prefetch_batch
                }
                for _pf_future in as_completed(_pf_futures):
                    _pf_done += 1
                    try:
                        _pf_result = _pf_future.result()
                        _pf_results.extend(_pf_result)
                        for _r in _pf_result:
                            if _r.get("method") not in (
                                "stix_metadata",
                                "no_content",
                                "",
                                "failed",
                            ):
                                _pf_ok += 1
                    except Exception:
                        pass

                    # Live progress line
                    _pf_elapsed = time.time() - _pf_start
                    _pf_remaining = _pf_total - _pf_done
                    if _pf_done > 1:
                        _pf_avg = _pf_elapsed / _pf_done
                        _pf_eta = _pf_avg * _pf_remaining
                        if _pf_eta >= 60:
                            _pf_eta_str = (
                                f"{int(_pf_eta // 60)}m {int(_pf_eta % 60):02d}s"
                            )
                        else:
                            _pf_eta_str = f"{int(_pf_eta)}s"
                    else:
                        _pf_eta_str = "..."
                    _pf_pct = _pf_done / _pf_total * 100
                    sys.stdout.write(
                        f"\r   -> Pre-fetch: {_pf_done:,}/{_pf_total:,} ({_pf_pct:.1f}%)  "
                        f"ETA: {_pf_eta_str}  "
                        f"({_pf_ok:,} fetched)   "
                    )
                    sys.stdout.flush()

            _pf_elapsed = time.time() - _pf_start
            if _pf_elapsed >= 60:
                _pf_elapsed_str = (
                    f"{int(_pf_elapsed // 60)}m {int(_pf_elapsed % 60):02d}s"
                )
            else:
                _pf_elapsed_str = f"{int(_pf_elapsed)}s"
            sys.stdout.write(
                f"\r\033[2K   ✅ \033[1mPre-fetch complete:\033[0m {_pf_ok:,}/{_pf_total:,} fetched in {_pf_elapsed_str}\n\n"
            )
            sys.stdout.flush()

    _pb_extract = _ProgressBar()
    _cit_num = 0  # running citation counter, resets per group
    _printed_citation_urls = set()  # URLs already printed in terminal, skip duplicates
    _rate_limited_count = 0  # 429 counter
    _active_workers = citation_workers  # adaptive worker count
    _max_workers = citation_workers
    _procs_since_last_429 = 0  # procedures since last rate limit

    for _proc_idx, each_procedure in enumerate(consolidated_procedures, 1):
        _proc_parts = each_procedure.split("||")
        current_group_name = _proc_parts[1]
        if (
            last_group_name
            and current_group_name.strip().lower() != last_group_name.strip().lower()
        ):
            _cit_num = 0
        last_group_name = current_group_name
        _pb_extract.update(
            _proc_idx,
            _total_procedures,
            len(_all_citation_refs),
            _total_cit_pairs,
            current_group_name,
            _rate_limited_count,
            _active_workers,
        )
        (
            technique_findings,
            previous_findings,
        ) = extract_indicators(
            each_procedure,
            terms,
            evidence_found,
            "",
            previous_findings,
            truncate,
            quiet,
        )
        threat_actor_technique_id_name_findings = []

        # constructing sub-technique pairing
        for technique_found in technique_findings:
            threat_actor_found = technique_found.split("||")[1]
            technique_id_found = technique_found.split("||")[2]
            technique_name_found = technique_found.split("||")[3]
            if "." in technique_id_found:
                parent_technique_found = "{}||{}".format(
                    technique_id_found,
                    str(sorted_techniques)
                    .split("{}||".format(technique_id_found))[1]
                    .split("{}".format(technique_name_found))[0][0:-2],
                )
                technique_id_name_found = "{}: {}".format(
                    parent_technique_found, technique_name_found
                )
            else:
                technique_id_name_found = "{}||{}".format(
                    technique_id_found, technique_name_found
                )
            threat_actor_technique_id_name_found = "{}||{}".format(
                threat_actor_found, technique_id_name_found
            )
            threat_actor_technique_id_name_findings.append(
                threat_actor_technique_id_name_found
            )

        # Collect citations silently into buffer (if -C enabled)
        if collect_citations and _citation_url_lookup:
            _parts = each_procedure.split("||")
            _raw_proc = _parts[4] if len(_parts) > 4 else ""
            _group = _parts[1] if len(_parts) > 1 else ""
            _tid = _parts[2] if len(_parts) > 2 else ""
            _tname = _parts[3] if len(_parts) > 3 else ""

            # Collect citations from procedure text, technique description, and detection guidance
            _all_text = _raw_proc
            if len(_parts) > 7:
                _all_text += " " + _parts[7]  # technique_description
            if len(_parts) > 8:
                _all_text += " " + _parts[8]  # technique_detection
            _cit_names = list(
                dict.fromkeys(re.findall(r"\(Citation:\s*([^)]+)\)", _all_text))
            )
            _new_cits = []
            if _cit_names:
                from src.citation_collector import collect_references_parallel

                # Build batch of unseen citations for this procedure
                _batch = []
                for _cn in _cit_names:
                    _cn = _cn.strip()
                    _display_key = (_group.strip().lower(), _cn)
                    if _display_key in _seen_citations:
                        continue
                    _seen_citations.add(_display_key)

                    _ref_data = _citation_url_lookup.get(_cn, {})
                    _batch.append(
                        {
                            "citation_name": _cn,
                            "url": _ref_data.get("url", ""),
                            "description": _ref_data.get("description", ""),
                        }
                    )

                # Fetch all citations for this procedure in parallel
                if _batch:
                    # Extract MITRE-documented indicators for this technique to
                    # guide relevance scoring (BM25/semantic query terms)
                    _proc_indicators: list = []
                    _ev_json_now = _parts[13] if len(_parts) > 13 else "{}"
                    try:
                        _ev_now = json.loads(_ev_json_now)
                        for _ev_vals in _ev_now.values():
                            if isinstance(_ev_vals, list):
                                for _ev_v in _ev_vals:
                                    if isinstance(_ev_v, str):
                                        _proc_indicators.append(_ev_v)
                                    elif isinstance(_ev_v, dict):
                                        _proc_indicators.extend(
                                            str(k) for k in _ev_v.keys()
                                        )
                    except (json.JSONDecodeError, TypeError):
                        pass

                    _fetched = collect_references_parallel(
                        _batch,
                        _tname,
                        _tid,
                        indicators=_proc_indicators or None,
                        max_workers=_active_workers,
                    )
                    _batch_429 = 0
                    for _ref in _fetched:
                        _ref["group"] = _group
                        _ref["technique_id"] = _tid
                        _ref["technique_name"] = _tname
                        # Deduplicate globally by URL — same source cited by many
                        # groups/techniques should only appear once in _all_citation_refs
                        _ref_url = _ref.get("url", "") or _ref.get("citation_name", "")
                        if _ref_url not in _seen_global_urls:
                            _seen_global_urls.add(_ref_url)
                            _all_citation_refs.append(_ref)
                        _new_cits.append(_ref)
                        for _att in _ref.get("attempts", []):
                            if "429" in str(_att):
                                _rate_limited_count += 1
                                _batch_429 += 1

                    # Adaptive throttling: reduce workers on 429, recover when stable
                    if _batch_429 > 0:
                        _active_workers = max(1, _active_workers // 2)
                        _procs_since_last_429 = 0
                    else:
                        _procs_since_last_429 += 1
                        if (
                            _procs_since_last_429 >= 50
                            and _active_workers < _max_workers
                        ):
                            _active_workers = min(_max_workers, _active_workers + 2)
                            _procs_since_last_429 = 0

            # Print citations for ALL techniques (even when no native indicators)
            if _new_cits:
                from src.citation_collector import (
                    extract_indicators_from_text,
                    filter_indicators_by_platform,
                    _INDICATOR_EMOJI,
                )
                from src.exclusions import filter_indicators as _filter_exclusions

                _tech_platforms = (
                    [p.strip() for p in _parts[9].split(",") if p.strip()]
                    if len(_parts) > 9
                    else []
                )

                # Build set of existing indicators for dedup
                _existing_indicators = set()
                _ev_json = _parts[13] if len(_parts) > 13 else "{}"
                try:
                    _ev = json.loads(_ev_json)
                    for _vals in _ev.values():
                        if isinstance(_vals, list):
                            for _v in _vals:
                                if isinstance(_v, dict):
                                    _existing_indicators.update(
                                        str(k).lower() for k in _v.keys()
                                    )
                                else:
                                    _existing_indicators.add(str(_v).lower())
                except (json.JSONDecodeError, TypeError):
                    pass

                _indent = "      "
                for _ref in _new_cits:
                    _ref_print_url = _ref.get("url", "") or _ref.get(
                        "citation_name", ""
                    )
                    # Skip terminal print for URLs already shown under another group/technique
                    if _ref_print_url in _printed_citation_urls:
                        continue
                    _printed_citation_urls.add(_ref_print_url)
                    _cit_num += 1
                    _cn = _ref.get("citation_name", "")
                    _num_str = f"[{_cit_num}]"
                    _method = _ref.get("method", "unknown")
                    if _method == "cached":
                        _icon = "\033[36m\U0001f4be\033[0m"  # cached locally
                    elif _ref.get("extracted_content") and _method not in (
                        "stix_metadata",
                        "no_content",
                    ):
                        _icon = "\033[32m\u2705\033[0m"  # real content fetched
                    elif _method == "stix_metadata":
                        _icon = "\033[33m\u26a0\ufe0f\033[0m "  # metadata only
                    else:
                        _icon = "\033[31m\u274c\033[0m"  # no content
                    _url = _ref.get("url", "")
                    # no_content entries are written to citations_failed.yaml — skip terminal noise
                    if _method == "no_content":
                        continue
                    try:
                        _tw = os.get_terminal_size().columns
                    except OSError:
                        _tw = 120
                    _method_short = _method[:14].ljust(14)
                    # Title: use as much space as available, leaving room for method + icon
                    _title_max = max(20, _tw - 6 - 5 - 4 - 14 - 3 - 5)
                    _name = _cn[:_title_max]
                    _url_indent = " " * (6 + 5 + 1)  # align under title
                    _url_part = (
                        f"\n{_url_indent}\033[90m[{_url}]\033[0m" if _url else ""
                    )
                    print(
                        f"{_indent}\033[90m{_num_str:>5}\033[0m \033[36m{_name}\033[0m \033[90m\u2192\033[0m \033[33m{_method_short}\033[0m {_icon}{_url_part}"
                    )

                    # Extract indicators from fetched content
                    _content = _ref.get("extracted_content", "")
                    if _content and _method not in ("stix_metadata", "no_content", ""):
                        _extracted = extract_indicators_from_text(_content)
                        if _extracted:
                            # Drop cmd indicators from wrong platform
                            # (e.g. Linux cmds from a Windows-only technique).
                            # Unknown tools are always kept. Revert: set
                            # PLATFORM_FILTER_ENABLED=False in citation_collector.py
                            _extracted = filter_indicators_by_platform(
                                _extracted, _tech_platforms
                            )
                            # Apply exclusion list
                            _extracted, _excluded = _filter_exclusions(_extracted)

                            # Filter out indicators already in MITRESaw's native extraction
                            _new_indicators = {}
                            for _etype, _evals in _extracted.items():
                                _novel = [
                                    v
                                    for v in _evals
                                    if v.lower() not in _existing_indicators
                                ]
                                if _novel:
                                    _new_indicators[_etype] = _novel
                                    _existing_indicators.update(
                                        v.lower() for v in _novel
                                    )

                            # Enrich any CVEs found in citation content
                            if "cve" in _new_indicators:
                                from src.tools.map_bespoke_logs import (
                                    enrich_cves_for_evidence,
                                )

                                _new_indicators["cve"] = enrich_cves_for_evidence(
                                    _new_indicators["cve"]
                                )

                            # Print new indicators with emojis
                            if _new_indicators:
                                for _etype, _evals in _new_indicators.items():
                                    _emoji = _INDICATOR_EMOJI.get(_etype, "")
                                    if _etype == "cve":
                                        # _evals is list of enriched dicts
                                        # [{cve_id: "product|desc|indicators|poc|kev"}]
                                        _cve_parts = []
                                        for _cve_entry in _evals:
                                            for _cve_id, _cve_val in _cve_entry.items():
                                                _vparts = (
                                                    _cve_val.split("|")
                                                    if _cve_val
                                                    else []
                                                )
                                                _ind = (
                                                    _vparts[2]
                                                    if len(_vparts) > 2
                                                    else ""
                                                )
                                                _cve_parts.append(
                                                    f"`{_cve_id}`"
                                                    + (f": {_ind}" if _ind else "")
                                                )
                                        _vals_str = ", ".join(_cve_parts[:4])
                                    else:
                                        _vals_str = ", ".join(
                                            f"`{v}`" for v in _evals[:8]
                                        )
                                    print(
                                        f"{_indent}       {_emoji} \033[33m{_vals_str}\033[0m"
                                    )

                                # Store extracted indicators on the ref for XLSX enrichment
                                _ref["extracted_indicators"] = _new_indicators

    threat_actor_technique_id_name_findings = list(
        set(threat_actor_technique_id_name_findings)
    )
    _with_content = 0
    _failed = []
    _failed_yaml = None
    if collect_citations and _all_citation_refs:
        _with_content = sum(
            1
            for r in _all_citation_refs
            if r.get("extracted_content")
            and r.get("method") not in ("stix_metadata", "no_content", "")
        )
        # Deduplicate by URL — the same source can be cited by many procedures
        _seen_failed_urls = set()
        _failed = []
        for r in _all_citation_refs:
            if r.get("method") in ("stix_metadata", "no_content", ""):
                _furl = r.get("url", "") or r.get("citation_name", "")
                if _furl not in _seen_failed_urls:
                    _seen_failed_urls.add(_furl)
                    _failed.append(r)
    _pb_extract.done(_total_procedures, len(_all_citation_refs))

    # Write failed citations report (YAML)
    if _failed:
        _failed_yaml = os.path.join(mitresaw_root_date, "citations_failed.yaml")
        _yaml_data = []
        for _r in _failed:
            _yaml_data.append(
                {
                    "citation_name": _r.get("citation_name", ""),
                    "url": _r.get("url", ""),
                    "method": _r.get("method", ""),
                    "attempts": _r.get("attempts", []),
                    "group": _r.get("group", ""),
                    "technique_id": _r.get("technique_id", ""),
                }
            )
        with open(_failed_yaml, "w") as _f:
            yaml.dump(_yaml_data, _f, default_flow_style=False, sort_keys=False)
    all_evidence.append(technique_findings)
    consolidated_techniques = all_evidence[0]

    # ---------------------------------------------------------------------------
    # Cross-technique redistribution (Fix 1)
    # Build MITRE indicator index from procedure rows, then redistribute
    # full-document indicators to techniques that didn't cite the URL directly.
    # ---------------------------------------------------------------------------
    if collect_citations and _all_citation_refs:
        from src.citation_collector import redistribute_citation_indicators

        _mitre_ind_index: dict = {}
        for _ct in consolidated_techniques:
            _ct_parts = _ct.split("||")
            if len(_ct_parts) <= 13:
                continue
            _ct_g = _ct_parts[1]
            _ct_tid = _ct_parts[2]
            # Skip citation-injected entries — use only MITRE-authored ones
            _ct_usage = _ct_parts[4] if len(_ct_parts) > 4 else ""
            if "Indicators extracted from citation:" in _ct_usage:
                continue
            try:
                _ct_ev = json.loads(_ct_parts[13])
                _ct_inds: set = set()
                for _ev_vals in _ct_ev.values():
                    if isinstance(_ev_vals, list):
                        for _v in _ev_vals:
                            if isinstance(_v, str):
                                _ct_inds.add(_v.lower())
                            elif isinstance(_v, dict):
                                _ct_inds.update(str(k).lower() for k in _v.keys())
                if _ct_inds:
                    _mitre_ind_index.setdefault((_ct_g, _ct_tid), set()).update(
                        _ct_inds
                    )
            except (json.JSONDecodeError, TypeError, IndexError):
                pass

        _redistributed_refs = redistribute_citation_indicators(
            _all_citation_refs, _mitre_ind_index
        )
        if _redistributed_refs:
            # Backfill technique_name from consolidated_techniques where missing
            _tid_to_name: dict = {}
            for _ct in consolidated_techniques:
                _ct_parts = _ct.split("||")
                if len(_ct_parts) > 3 and _ct_parts[2] and _ct_parts[3]:
                    _tid_to_name[_ct_parts[2]] = _ct_parts[3]
            for _rr in _redistributed_refs:
                if not _rr.get("technique_name"):
                    _rr["technique_name"] = _tid_to_name.get(
                        _rr.get("technique_id", ""), ""
                    )
            _all_citation_refs.extend(_redistributed_refs)

    # Inject citation-extracted indicators as additional entries
    _injected = 0
    if collect_citations and _all_citation_refs:
        for _ref in _all_citation_refs:
            _ext_ind = _ref.get("extracted_indicators", {})
            if not _ext_ind:
                continue
            _g = _ref.get("group", "")
            _tid = _ref.get("technique_id", "")
            _tname = _ref.get("technique_name", "")
            _cn = _ref.get("citation_name", "")
            _url = _ref.get("url", "")
            # Build a consolidated_techniques entry in the same ||-delimited format
            # [0]group_id [1]group_name [2]technique_id [3]technique_name
            # [4]usage [5]- [6]group_desc [7]tech_desc [8]tech_detection
            # [9]tech_platforms [10]tech_datasources [11]tech_tactics [12]framework [13]evidence_json
            _evidence_json = json.dumps(_ext_ind)
            _usage = f"Indicators extracted from citation: {_cn}"
            if _url:
                _usage += f" ({_url})"
            # Find an existing entry to copy metadata from
            _template = None
            for _ct in consolidated_techniques:
                _ct_parts = _ct.split("||")
                if len(_ct_parts) > 3 and _ct_parts[1] == _g and _ct_parts[2] == _tid:
                    _template = _ct_parts
                    break
            if _template and len(_template) > 12:
                _new_entry = (
                    f"{_template[0]}||{_g}||{_tid}||{_tname}||"
                    f"{_usage}||-||{_template[6]}||{_template[7]}||{_template[8]}||"
                    f"{_template[9]}||{_template[10]}||{_template[11]}||{_template[12]}||"
                    f"{_evidence_json}"
                )
            else:
                _new_entry = (
                    f"||{_g}||{_tid}||{_tname}||"
                    f"{_usage}||-||||||||||"
                    f"{_evidence_json}"
                )
            consolidated_techniques.append(_new_entry)
            _injected += 1
    # Print citation breakdown
    if collect_citations and _all_citation_refs:
        _cit_total = len(_all_citation_refs)
        _n_failed = len(_failed)
        # Right-align the fraction columns
        _cw = len(f"{_cit_total:,}")  # width of the total (widest number)
        print(f"\n   Citations:")
        print(
            f"     ✅ With content: {_with_content:>{_cw},}/{_cit_total:,}  ({_with_content / _cit_total * 100:.1f}%)"
        )
        if _injected:
            print(
                f"     🔍 Newly added:  {_injected:>{_cw},}/{_cit_total:,}  ({_injected / _cit_total * 100:.1f}%)"
            )
        if _n_failed:
            print(
                f"     ❌ Failed:       {_n_failed:>{_cw},}/{_cit_total:,}  ({_n_failed / _cit_total * 100:.1f}%)"
            )

    # Report CVEs with no actionable intelligence
    from src.tools.map_bespoke_logs import report_cve_summary

    print()
    report_cve_summary()

    if len(consolidated_techniques) > 0:

        # outputting relevant queries
        query_pairings, mapped_log_sources = build_matrix(
            mitresaw_output_directory,
            consolidated_techniques,
            sorted_threat_actors_techniques_in_scope,
            threat_actor_technique_id_name_findings,
        )
        """if queries:
            print()
            print(
                "    -> Compiling queries based on \033[1;31midentifiers\033[1;m based on {}".format(
                    terms_insert
                )
            )"""
        # outputting csv file for ingestion into other tools
        write_csv_summary(
            consolidated_techniques,
            mitresaw_output_directory,
            mitre_files,
            queries,
            query_pairings,
            log_sources,
        )
        # Export main CSV files to JSON/XML if requested
        if export_format != "csv":
            for csv_name in ["ThreatActors_Techniques"]:
                csv_path = os.path.join(mitresaw_output_directory, f"{csv_name}.csv")
                if os.path.exists(csv_path):
                    df = pandas.read_csv(csv_path, on_bad_lines="warn")
                    if export_format == "json":
                        out_path = os.path.join(
                            mitresaw_output_directory, f"{csv_name}.json"
                        )
                        df.to_json(out_path, orient="records", indent=2)
                    elif export_format == "xml":
                        out_path = os.path.join(
                            mitresaw_output_directory, f"{csv_name}.xml"
                        )
                        df.to_xml(out_path)
                    print(f"      {export_format.upper()} written to {out_path}")
        # Generate filtered export if --columns is specified
        if columns:
            valid_columns = [
                "group_sw_id",
                "group_sw_name",
                "group_sw_description",
                "technique_id",
                "technique_name",
                "technique_description",
                "tactic",
                "platforms",
                "framework",
                "procedure_example",
                "evidence",
                "detectable_via",
                "keywords",
            ]
            requested_columns = [c.strip() for c in columns.split(",")]
            invalid = [c for c in requested_columns if c not in valid_columns]
            if invalid:
                print(f"\n    Error: Invalid column(s): {', '.join(invalid)}")
                print(f"    Valid columns: {', '.join(valid_columns)}")
            else:
                csv_path = os.path.join(
                    mitresaw_output_directory, "ThreatActors_Techniques.csv"
                )
                df = pandas.read_csv(csv_path, on_bad_lines="warn")

                if "keywords" in requested_columns:
                    keyword_map = {}
                    for gid, info in group_info_data.items():
                        keyword_map[gid] = match_keywords(info["description"])
                    df["keywords"] = df["group_sw_id"].map(keyword_map).fillna("")

                df = df[requested_columns].drop_duplicates()
                if preset:
                    filtered_base_name = "mitre_procedures"
                    filtered_dir = mitresaw_root_date
                else:
                    filtered_base_name = "mitre_procedures"
                    filtered_dir = mitresaw_output_directory
                if export_format == "json":
                    filtered_path = os.path.join(
                        filtered_dir, f"{filtered_base_name}.json"
                    )
                    df.to_json(filtered_path, orient="records", indent=2)
                elif export_format == "xml":
                    filtered_path = os.path.join(
                        filtered_dir, f"{filtered_base_name}.xml"
                    )
                    df.to_xml(filtered_path)
                else:
                    filtered_path = os.path.join(
                        filtered_dir, f"{filtered_base_name}.csv"
                    )
                    df.to_csv(filtered_path, index=False)
                if not evidence_report:
                    print(f"      Filtered export written to {filtered_path}")

        mitresaw_techniques = re.findall(
            r"\|\|(T\d{3}[\d\.]+)\|\|", str(consolidated_techniques)
        )
        mitresaw_techniques = list(set(mitresaw_techniques))
        mitresaw_techniques_insert = str(mitresaw_techniques)[2:-2].replace(
            "', '",
            '", "comment": "", "score": 1, "color": "#66b1ff", "showSubtechniques": false}}, {{"techniqueID": "',
        )

        # Generate ATT&CK Navigator layer per framework
        for fw in attack_frameworks:
            domain = f"{fw.lower()}-attack"
            mitresaw_navlayer = '{{"description": "{} techniques used by various Threat Actors, produced by MITRESaw", "name": "{}", "domain": "{}", "versions": {{"layer": "4.4", "attack": "15", "navigator": "4.8.1"}}, "techniques": [{{"techniqueID": "{}", "comment": "", "score": 1, "color": "#66b1ff", "showSubtechniques": false}}], "gradient": {{"colors": ["#ffffff", "#66b1ff"], "minValue": 0, "maxValue": 1}}, "legendItems": [{{"label": "identified from MITRESaw analysis", "color": "#66b1ff"}}]}}\n'.format(
                fw,
                mitresaw_output_directory.split("/")[2][11:],
                domain,
                mitresaw_techniques_insert,
            )
            with open(
                os.path.join(mitresaw_output_directory, f"{domain}-layer.json"), "w"
            ) as mitresaw_navlayer_json:
                mitresaw_navlayer_json.write(
                    mitresaw_navlayer.replace("{{", "{").replace("}}", "}")
                )
        build_queries(queries, mitresaw_output_directory, query_pairings)

        # Generate evidence report if requested
        if evidence_report:
            csv_path = os.path.join(
                mitresaw_output_directory, "ThreatActors_Techniques.csv"
            )
            if os.path.exists(csv_path):
                df = pandas.read_csv(csv_path, on_bad_lines="warn")
                result_rows = df.to_dict(orient="records")
                from src.evidence_report import generate_evidence_report
                from datetime import datetime as _dt

                # If no filters provided, put alongside mitre_procedures.csv
                no_filters = (
                    str(operating_platforms) == "['.']"
                    and str(search_terms) == "['.']"
                    and str(provided_groups) == "['.']"
                )
                _er_dir = (
                    mitresaw_root_date if no_filters else mitresaw_output_directory
                )
                _er_path = os.path.join(
                    _er_dir,
                    "mitre_procedures.xlsx",
                )
                generate_evidence_report(
                    rows=result_rows,
                    output_path=_er_path,
                    framework=",".join(attack_frameworks),
                    platforms_arg=",".join(str(p) for p in operating_platforms),
                    searchterms_arg=",".join(str(s) for s in search_terms),
                    threatgroups_arg=",".join(str(g) for g in provided_groups),
                )

                # Append Reference Detail sheet if citations were collected
                if collect_citations and _all_citation_refs:
                    _write_reference_sheet(_er_path, _all_citation_refs)

                # Move CSV alongside evidence report with matching name
                import shutil

                _csv_dest = os.path.join(_er_dir, "mitre_procedures.csv")
                shutil.move(csv_path, _csv_dest)
                print(f"\n   Outputs written to: {_er_dir}/")
                print(f"     🏛️  mitre_procedures.csv")
                print(f"     📎 mitre_procedures.xlsx")
                if _failed_yaml:
                    print(f"     🍠 citations_failed.yaml")

    else:
        print("\n   No evidence could be found which match the provided criteria.")
    print()
