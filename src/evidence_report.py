"""
MITRESaw Evidence Report Generator

Generates a styled XLSX evidence report with one row per atomic indicator
extracted from MITRE ATT&CK procedure examples.
"""

import json
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# XML 1.0 forbids certain control characters in cell values. openpyxl writes
# them verbatim, producing an XLSX that Excel rejects with a "found a problem
# with some content" dialog. Strip them before assigning any string cell value.
# Allowed: \x09 (tab), \x0A (LF), \x0D (CR), \x20+ (printable).
_RE_XML_ILLEGAL = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F\uFFFE\uFFFF]")


def _safe(val):
    """Strip XML-illegal characters from string values; pass other types through."""
    if isinstance(val, str):
        return _RE_XML_ILLEGAL.sub("", val)
    return val




# ---------------------------------------------------------------------------
# Port protocol mapping
# ---------------------------------------------------------------------------

_PORT_PROTOCOL = {
    "80": "TCP/{} (HTTP)",
    "8080": "TCP/{} (HTTP)",
    "8443": "TCP/{} (HTTPS)",
    "8888": "TCP/{} (HTTP)",
    "443": "TCP/{} (TLS)",
    "53": "UDP/TCP/{} (DNS)",
    "22": "TCP/{} (SSH)",
    "3389": "TCP/{} (RDP)",
    "445": "TCP/{} (SMB)",
    "139": "TCP/{} (NetBIOS)",
}


def _format_port(port_str: str) -> str:
    """Format a port string with protocol prefix."""
    port = port_str.strip()
    fmt = _PORT_PROTOCOL.get(port)
    if fmt:
        return fmt.format(port)
    return f"TCP/{port}"


# ---------------------------------------------------------------------------
# Detection context strings
# ---------------------------------------------------------------------------

_DETECTION_CONTEXT = {
    "cmd": (
        "Process Creation \u2014 Sysmon EID 1 / Windows Security EID 4688"
        " (requires command-line auditing enabled)"
    ),
    "reg": (
        "Registry modification \u2014 Sysmon EID 12/13/14 / Windows"
        " Security EID 4657 (requires object access auditing)"
    ),
    "cve": (
        "Exploit telemetry \u2014 check CISA KEV for active exploitation"
        " status; review NVD for PoC availability; patch status is primary control"
    ),
    "ports": (
        "Network traffic \u2014 firewall/proxy logs, Zeek conn.log,"
        " Sysmon EID 3 (network connection)"
    ),
    "paths": (
        "File creation/modification \u2014 Sysmon EID 11 (FileCreate),"
        " EID 23 (FileDelete) / EDR file telemetry"
    ),
    "filepath": (
        "File creation/modification \u2014 Sysmon EID 11 (FileCreate),"
        " EID 23 (FileDelete) / EDR file telemetry"
    ),
    "software": (
        "Process name / image load \u2014 Sysmon EID 1 (process),"
        " EID 7 (image load); check GitHub for tool-specific CLI usage"
    ),
    "event_ids": (
        "This IS a Windows event ID \u2014 ensure the corresponding log"
        " channel is enabled and ingested into your SIEM"
    ),
    "evt": (
        "This IS a Windows event ID \u2014 ensure the corresponding log"
        " channel is enabled and ingested into your SIEM"
    ),
    "none": "No extractable indicators \u2014 review procedure text manually",
}

# Source type labels in display order
_SOURCE_ORDER = ["Procedure", "Technique", "Citation"]


def _format_source_type(sources: set) -> str:
    """Format a set of source labels into an ordered pipe-delimited string."""
    ordered = [s for s in _SOURCE_ORDER if s in sources]
    return " | ".join(ordered) if ordered else "MITRE ATT&CK"


# ---------------------------------------------------------------------------
# Group colour/background mapping
# ---------------------------------------------------------------------------

def _group_bg(name: str) -> str:
    """Return hex background colour for a group row."""
    n = name.lower()
    if "oilrig" in n or "apt34" in n:
        return "0F2035"
    if "apt33" in n or "peach sandstorm" in n:
        return "0F250F"
    if "muddywater" in n or "seedworm" in n:
        return "1F0A2A"
    if "magic hound" in n or "apt35" in n:
        return "2A150A"
    if "apt39" in n or "chafer" in n:
        return "0A1535"
    if "fox kitten" in n or "pioneer kitten" in n:
        return "1A0A2A"
    return "0F1C2E"


def _group_accent(name: str) -> str:
    """Return hex accent colour for group name font."""
    n = name.lower()
    if "oilrig" in n or "apt34" in n:
        return "38BDF8"
    if "apt33" in n or "peach sandstorm" in n:
        return "4ADE80"
    if "muddywater" in n or "seedworm" in n:
        return "C084FC"
    if "magic hound" in n or "apt35" in n:
        return "FB923C"
    if "apt39" in n or "chafer" in n:
        return "60A5FA"
    if "fox kitten" in n or "pioneer kitten" in n:
        return "E879F9"
    return "E0F2FE"


# ---------------------------------------------------------------------------
# Procedure text cleanup
# ---------------------------------------------------------------------------

_RE_MD_LINK = re.compile(r"\[([^\]]+)\]\((https?://[^\)]+)\)")
_RE_CITATION = re.compile(r"\(Citation:[^\)]*\)")


def _md_link_to_id(m: re.Match) -> str:
    """Replace a markdown link with 'Text (ID)' using the last path segment."""
    label = m.group(1)
    url = m.group(2).rstrip("/")
    identifier = url.rsplit("/", 1)[-1]
    return f"{label} ({identifier})"


def _clean_procedure_text(text: str) -> str:
    """Clean MITRE procedure text for display.

    - Convert markdown links [Text](URL) → Text (identifier)
      e.g. [Axiom](https://attack.mitre.org/groups/G0001) → Axiom (G0001)
    - Remove all (Citation: ...) references
    - Collapse extra whitespace left behind
    """
    if not text:
        return text
    text = _RE_MD_LINK.sub(_md_link_to_id, text)
    text = _RE_CITATION.sub("", text)
    text = re.sub(r"  +", " ", text).strip()
    return text


# ---------------------------------------------------------------------------
# URL helpers
# ---------------------------------------------------------------------------

_RE_URL = re.compile(r"https?://\S{10,}")


def _extract_url(procedure_text: str, technique_id: str) -> str:
    """Extract URL from procedure text or construct ATT&CK technique URL."""
    m = _RE_URL.search(procedure_text or "")
    if m:
        url = m.group(0).rstrip(".,;)")
        return url
    # Construct from technique_id: T1059.001 → T1059/001
    tid = technique_id.replace(".", "/")
    return f"https://attack.mitre.org/techniques/{tid}/"


def _nav_layer_url(group_id: str) -> str:
    """Construct ATT&CK Navigator layer URL from group ID."""
    if not group_id or not group_id.startswith("G"):
        return "N/A"
    return f"https://attack.mitre.org/groups/{group_id}/{group_id}-enterprise-layer.json"


# ---------------------------------------------------------------------------
# Contextual evidence builder
# ---------------------------------------------------------------------------

def _build_detection(indicator_type: str, detectable_via: str) -> str:
    """Build detection guidance string."""
    det_parts = []
    det = _DETECTION_CONTEXT.get(indicator_type, "")
    if det:
        det_parts.append(det.lstrip("\n"))
    if detectable_via and str(detectable_via).strip() and str(detectable_via) != "nan":
        det_parts.append(f"ATT&CK Data Source(s): {detectable_via}")
    return "\n".join(det_parts)


# ---------------------------------------------------------------------------
# XLSX styling constants
# ---------------------------------------------------------------------------

_BG_NAVY = "0D1B2A"
_BG_DARK = "0F1C2E"
_BG_ALT = "0A1220"
_BORDER_CLR = "1E3A5F"

_THIN_BORDER = Border(
    left=Side(style="thin", color=_BORDER_CLR),
    right=Side(style="thin", color=_BORDER_CLR),
    top=Side(style="thin", color=_BORDER_CLR),
    bottom=Side(style="thin", color=_BORDER_CLR),
)

_COL_WIDTHS = [50, 16, 14, 28, 18, 14, 14, 22, 55, 50, 40, 45, 38]

_HEADER_LABELS = [
    "Evidential Element",
    "Threat Group",
    "Technique ID",
    "Technique Name",
    "Tactic",
    "Platforms",
    "Framework",
    "Source Type",
    "Procedure Example",
    "Detection Guidance",
    "Log Sources",
    "Reference URL",
    "Navigation Layer URL",
]

_NUM_COLS = len(_HEADER_LABELS)  # 11
_LAST_COL_LETTER = get_column_letter(_NUM_COLS)  # "K"


# ---------------------------------------------------------------------------
# generate_evidence_report()
# ---------------------------------------------------------------------------

def generate_evidence_report(
    rows: list,
    output_path: str,
    framework: str = "Enterprise",
    platforms_arg: str = ".",
    searchterms_arg: str = ".",
    threatgroups_arg: str = ".",
) -> None:
    """Generate a styled XLSX evidence report from MITRESaw result rows.

    Parameters
    ----------
    rows : list[dict]
        Result dicts with keys matching ThreatActors_Techniques.csv columns:
        group_sw_id, group_sw_name, technique_id, technique_name, tactic,
        procedure_example, evidence, detectable_via.
    output_path : str
        Path for the output XLSX file.
    framework, platforms_arg, searchterms_arg, threatgroups_arg : str
        Metadata for the report title/subtitle.
    """

    # Map from evidence dict keys to canonical indicator types
    # The evidence JSON uses "evt" and "filepath" internally
    evidence_key_map = {
        "cmd": "cmd",
        "reg": "reg",
        "cve": "cve",
        "ports": "ports",
        "filepath": "paths",
        "software": "software",
        "evt": "event_ids",
    }

    # ------------------------------------------------------------------
    # ATOMISE rows → one output row per indicator
    # Keyed dict preserves insertion order and enables multi-source merging.
    # ------------------------------------------------------------------
    atomised_map = {}  # dedup_key → entry dict

    for row in rows:
        group_id = str(row.get("group_sw_id", "") or "")
        group_name = _clean_procedure_text(str(row.get("group_sw_name", "") or ""))
        technique_id = str(row.get("technique_id", "") or "")
        technique_name = _clean_procedure_text(str(row.get("technique_name", "") or ""))
        tactic = str(row.get("tactic", "") or "")
        platforms = str(row.get("platforms", "") or "")
        framework = str(row.get("framework", "") or "")
        procedure_text = str(row.get("procedure_example", "") or "")
        procedure_display = _clean_procedure_text(procedure_text)
        tech_desc = str(row.get("technique_description", "") or "")
        detectable_via = str(row.get("detectable_via", "") or "")
        evidence_raw = row.get("evidence", "{}")

        # Parse evidence JSON
        if isinstance(evidence_raw, str):
            try:
                evidence = json.loads(evidence_raw)
            except (json.JSONDecodeError, TypeError):
                evidence = {}
        elif isinstance(evidence_raw, dict):
            evidence = evidence_raw
        else:
            evidence = {}

        is_citation_row = procedure_text.startswith("Indicators extracted from citation:")
        has_any = False

        for ev_key, indicator_type in evidence_key_map.items():
            values = evidence.get(ev_key, [])
            if not values:
                continue

            has_any = True
            for val in values:
                # CVE entries are dicts: {"CVE-ID": "description|..."}
                if isinstance(val, dict):
                    indicator_value = list(val.keys())[0] if val else ""
                else:
                    indicator_value = str(val)

                if not indicator_value:
                    continue

                # Format display value
                if indicator_type == "ports":
                    display_value = _format_port(indicator_value)
                elif indicator_type == "event_ids":
                    display_value = f"Windows Event ID {indicator_value}"
                else:
                    display_value = indicator_value

                dedup_key = (group_name.lower(), technique_id.lower(), display_value.lower())

                # Determine source labels for this occurrence
                if is_citation_row:
                    new_sources = {"Citation"}
                else:
                    new_sources = set()
                    iv_lower = indicator_value.lower()
                    if iv_lower in procedure_text.lower():
                        new_sources.add("Procedure")
                    if iv_lower in tech_desc.lower():
                        new_sources.add("Technique")
                    if not new_sources:
                        new_sources.add("Procedure")  # fallback

                if dedup_key in atomised_map:
                    # Merge source types into the existing entry
                    existing = atomised_map[dedup_key]
                    existing["_sources"] |= new_sources
                    existing["source_type"] = _format_source_type(existing["_sources"])
                    continue

                # Reference URL and nav URL
                ref_url = _extract_url(procedure_text, technique_id)
                nav_url = _nav_layer_url(group_id)

                atomised_map[dedup_key] = {
                    "evidential_element": display_value,
                    "threat_group": group_name,
                    "procedure_example": procedure_display,
                    "technique_id": technique_id,
                    "technique_name": technique_name,
                    "tactic": tactic,
                    "platforms": platforms if platforms != "nan" else "",
                    "framework": framework if framework != "nan" else "",
                    "detection_guidance": _build_detection(indicator_type, detectable_via),
                    "log_sources": detectable_via if detectable_via != "nan" else "",
                    "reference_url": ref_url,
                    "nav_layer_url": nav_url,
                    "source_type": _format_source_type(new_sources),
                    "_sources": new_sources,
                }

        # If no evidence at all, emit placeholder row
        if not has_any:
            dedup_key = (group_name.lower(), technique_id.lower(), "(no extractable indicators)")
            if dedup_key not in atomised_map:
                atomised_map[dedup_key] = {
                    "evidential_element": "(no extractable indicators)",
                    "threat_group": group_name,
                    "procedure_example": procedure_display,
                    "technique_id": technique_id,
                    "technique_name": technique_name,
                    "tactic": tactic,
                    "platforms": platforms if platforms != "nan" else "",
                    "framework": framework if framework != "nan" else "",
                    "detection_guidance": _DETECTION_CONTEXT["none"],
                    "log_sources": detectable_via if detectable_via != "nan" else "",
                    "reference_url": _extract_url(procedure_text, technique_id),
                    "nav_layer_url": _nav_layer_url(group_id),
                    "source_type": "MITRE ATT&CK",
                    "_sources": set(),
                }

    # Strip internal tracking key and flatten to list
    atomised = []
    for entry in atomised_map.values():
        entry.pop("_sources", None)
        atomised.append(entry)

    # ------------------------------------------------------------------
    # Build XLSX workbook
    # ------------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Evidence Report"
    ws.sheet_view.showGridLines = False

    # Fonts
    font_title = Font(name="Calibri", size=16, bold=True, color="0EA5E9")
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="7FB3D3")
    font_header = Font(name="Calibri", size=12, bold=True, color="E0F2FE")
    font_col1  = Font(name="Courier New", size=10, bold=True, color="2DD4BF")  # Evidential Element
    font_col3  = Font(name="Courier New", size=10, bold=True, color="22C55E") # Technique ID
    font_col4  = Font(name="Calibri",    size=10, color="E0F2FE")             # Technique Name
    font_col5  = Font(name="Calibri",    size=10, color="FACC15")             # Tactic
    font_col6  = Font(name="Calibri",    size=10, color="94A3B8")             # Platforms
    font_col7  = Font(name="Calibri",    size=10, color="94A3B8")             # Framework
    font_col8  = Font(name="Calibri",    size=10, color="F97316")             # Source Type
    font_col9  = Font(name="Calibri",    size=10, color="CBD5E1")             # Procedure Example
    font_col10 = Font(name="Courier New", size=10, color="CBD5E1")            # Detection Guidance
    font_col11 = Font(name="Courier New", size=10, color="CBD5E1")            # Log Sources
    font_col12 = Font(name="Calibri",    size=10, color="0EA5E9")             # Reference URL
    font_col13 = Font(name="Calibri",    size=10, color="A78BFA")             # Nav Layer URL

    fill_navy = PatternFill(start_color=_BG_NAVY, end_color=_BG_NAVY, fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Column widths
    for i, w in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ------------------------------------------------------------------
    # Row 1 — Title banner
    # ------------------------------------------------------------------
    unique_groups = len(set(r["threat_group"] for r in atomised)) if atomised else 0
    total_indicators = len(atomised)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    title_text = (
        f"MITRESaw Evidence Report  |  {framework}  |  "
        f"Groups: {unique_groups}  |  Indicators: {total_indicators}  |  "
        f"Generated: {now_str}"
    )
    ws.merge_cells(f"A1:{_LAST_COL_LETTER}1")
    cell = ws["A1"]
    cell.value = title_text
    cell.font = font_title
    cell.fill = fill_navy
    cell.alignment = align_center
    cell.border = _THIN_BORDER
    ws.row_dimensions[1].height = 28

    # ------------------------------------------------------------------
    # Row 2 — Subtitle
    # ------------------------------------------------------------------
    subtitle_text = (
        f"Platforms: {platforms_arg}  |  Search Terms: {searchterms_arg}  |  "
        f"Threat Groups: {threatgroups_arg}  |  Source: MITRE ATT&CK STIX via MITRESaw"
    )
    ws.merge_cells(f"A2:{_LAST_COL_LETTER}2")
    cell = ws["A2"]
    cell.value = subtitle_text
    cell.font = font_subtitle
    cell.fill = fill_navy
    cell.alignment = align_center
    cell.border = _THIN_BORDER
    ws.row_dimensions[2].height = 16

    # ------------------------------------------------------------------
    # Row 3 — Headers
    # ------------------------------------------------------------------
    for col_idx, label in enumerate(_HEADER_LABELS, 1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = _THIN_BORDER
    ws.row_dimensions[3].height = 40

    # ------------------------------------------------------------------
    # Data rows (starting row 4)
    # ------------------------------------------------------------------
    col_fonts = {
        1:  font_col1,
        2:  None,       # per-group accent
        3:  font_col3,
        4:  font_col4,
        5:  font_col5,
        6:  font_col6,
        7:  font_col7,
        8:  font_col8,
        9:  font_col9,
        10: font_col10,
        11: font_col11,
        12: font_col12,
        13: font_col13,
    }

    for row_idx, item in enumerate(atomised, 4):
        group_name = item["threat_group"]
        bg_hex = _group_bg(group_name)

        # Alternate tint for even data rows (0-indexed: row_idx 4 is first data row)
        data_row_num = row_idx - 4
        if data_row_num % 2 == 1:
            bg_hex = _BG_ALT

        row_fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")

        values = [
            item["evidential_element"],     # 1
            group_name,                      # 2
            item["technique_id"],            # 3
            item["technique_name"],          # 4
            item["tactic"],                  # 5
            item.get("platforms", ""),       # 6
            item.get("framework", ""),       # 7
            item["source_type"],             # 8
            item["procedure_example"],       # 9
            item["detection_guidance"],      # 10
            item.get("log_sources", ""),     # 11
            item["reference_url"],           # 12
            item["nav_layer_url"],           # 13
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_safe(val))
            cell.fill = row_fill
            cell.alignment = align_left
            cell.border = _THIN_BORDER

            if col_idx == 2:
                accent = _group_accent(group_name)
                cell.font = Font(name="Calibri", size=10, bold=True, color=accent)
            else:
                cell.font = col_fonts[col_idx]

            # Hyperlinks for URL columns
            if col_idx == 12 and val and val.startswith("http"):
                cell.hyperlink = val
            if col_idx == 13 and val and val != "N/A" and val.startswith("http"):
                cell.hyperlink = val

        ws.row_dimensions[row_idx].height = 70

    # Freeze panes at A4
    ws.freeze_panes = "A4"

    # Auto-filter on row 3
    last_data_row = 3 + len(atomised)
    if atomised:
        ws.auto_filter.ref = f"A3:{_LAST_COL_LETTER}{last_data_row}"

    # Apply border to merged cells in rows 1-2
    for col_idx in range(2, _NUM_COLS + 1):
        for r in (1, 2):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = _THIN_BORDER

    # ------------------------------------------------------------------
    # Sheet 2 — Group Summary
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Group Summary")
    ws2.sheet_view.showGridLines = False

    gs_headers = [
        "Group Name", "Technique Count", "Indicator Count",
        "Tactic Coverage", "Top Tactic",
    ]
    for col_idx, label in enumerate(gs_headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=label)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = _THIN_BORDER
    ws2.row_dimensions[3].height = 40

    # Column widths for Group Summary
    gs_widths = [30, 18, 18, 50, 25]
    for i, w in enumerate(gs_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Compute group stats
    group_stats = {}
    for item in atomised:
        g = item["threat_group"]
        if g not in group_stats:
            group_stats[g] = {
                "techniques": set(),
                "indicators": 0,
                "tactics": {},
            }
        group_stats[g]["techniques"].add(item["technique_id"])
        group_stats[g]["indicators"] += 1
        tactic = item["tactic"]
        group_stats[g]["tactics"][tactic] = group_stats[g]["tactics"].get(tactic, 0) + 1

    gs_row = 4
    for g_name, stats in sorted(group_stats.items()):
        tactics_sorted = sorted(stats["tactics"].items(), key=lambda x: x[1], reverse=True)
        top_tactic = tactics_sorted[0][0] if tactics_sorted else ""
        tactic_coverage = ", ".join(sorted(stats["tactics"].keys()))

        values = [
            g_name,
            len(stats["techniques"]),
            stats["indicators"],
            tactic_coverage,
            top_tactic,
        ]
        bg = _group_bg(g_name)
        row_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=gs_row, column=col_idx, value=_safe(val))
            cell.font = Font(name="Calibri", size=10, color="E0F2FE")
            cell.fill = row_fill
            cell.alignment = align_left
            cell.border = _THIN_BORDER
        gs_row += 1

    # ------------------------------------------------------------------
    # Sheet 3 — Tactic Pivot
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("Tactic Pivot")
    ws3.sheet_view.showGridLines = False

    tp_headers = [
        "Tactic", "Indicator Count", "% of Total", "Example Technique IDs",
    ]
    for col_idx, label in enumerate(tp_headers, 1):
        cell = ws3.cell(row=3, column=col_idx, value=label)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = _THIN_BORDER
    ws3.row_dimensions[3].height = 40

    tp_widths = [30, 18, 14, 55]
    for i, w in enumerate(tp_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Compute tactic stats
    tactic_stats = {}
    for item in atomised:
        t = item["tactic"]
        if t not in tactic_stats:
            tactic_stats[t] = {
                "count": 0,
                "techniques": set(),
            }
        tactic_stats[t]["count"] += 1
        tactic_stats[t]["techniques"].add(item["technique_id"])

    total_count = sum(s["count"] for s in tactic_stats.values())
    tp_row = 4
    last_tp_row = tp_row + len(tactic_stats) - 1

    for tactic, stats in sorted(tactic_stats.items(), key=lambda x: x[1]["count"], reverse=True):
        example_tids = " | ".join(sorted(stats["techniques"])[:5])
        pct = stats["count"] / total_count if total_count else 0

        values = [
            tactic,
            stats["count"],
            pct,
            example_tids,
        ]
        bg = _BG_DARK if (tp_row - 4) % 2 == 0 else _BG_ALT
        row_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        for col_idx, val in enumerate(values, 1):
            cell = ws3.cell(row=tp_row, column=col_idx, value=_safe(val))
            cell.font = Font(name="Calibri", size=10, color="E0F2FE")
            cell.fill = row_fill
            cell.alignment = align_left
            cell.border = _THIN_BORDER
            if col_idx == 3:
                cell.number_format = "0.0%"
        tp_row += 1

    # ------------------------------------------------------------------
    # Sheet 4 — Technique Matrix (only if 2+ groups)
    # ------------------------------------------------------------------
    unique_groups_list = sorted(set(item["threat_group"] for item in atomised))
    if len(unique_groups_list) >= 2:
        ws4 = wb.create_sheet("Technique Matrix")
        ws4.sheet_view.showGridLines = False

        # Build group→techniques mapping
        group_techs = {}
        tech_names = {}
        tech_tactics = {}
        for item in atomised:
            g = item["threat_group"]
            tid = item["technique_id"]
            group_techs.setdefault(g, set()).add(tid)
            if tid not in tech_names:
                tech_names[tid] = item["technique_name"]
            if tid not in tech_tactics:
                tech_tactics[tid] = item["tactic"]

        # Count how many groups use each technique, sort descending
        all_tids = set()
        for tids in group_techs.values():
            all_tids.update(tids)
        tech_group_count = {}
        for tid in all_tids:
            tech_group_count[tid] = sum(
                1 for g in unique_groups_list if tid in group_techs.get(g, set())
            )
        sorted_tids = sorted(
            all_tids, key=lambda t: (-tech_group_count[t], t)
        )

        # Title row
        num_matrix_cols = 4 + len(unique_groups_list)  # Technique ID, Name, Tactic, Count, groups...
        last_matrix_col = get_column_letter(num_matrix_cols)
        ws4.merge_cells(f"A1:{last_matrix_col}1")
        t = ws4["A1"]
        t.value = (
            f"Technique Intersection Matrix  |  {len(sorted_tids)} Techniques  |  "
            f"{len(unique_groups_list)} Groups  |  Sorted by group coverage (descending)"
        )
        t.font = font_title
        t.fill = fill_navy
        t.alignment = align_center
        t.border = _THIN_BORDER
        ws4.row_dimensions[1].height = 28

        # Headers — row 2
        matrix_headers = ["Technique ID", "Technique Name", "Tactic", "Group Count"]
        matrix_headers.extend(unique_groups_list)
        for col_idx, label in enumerate(matrix_headers, 1):
            cell = ws4.cell(row=2, column=col_idx, value=label)
            cell.font = font_header
            cell.fill = fill_navy
            cell.alignment = Alignment(
                horizontal="center", vertical="center",
                wrap_text=True, text_rotation=90 if col_idx > 4 else 0,
            )
            cell.border = _THIN_BORDER
        ws4.row_dimensions[2].height = 120 if len(unique_groups_list) > 5 else 60

        # Column widths
        ws4.column_dimensions["A"].width = 14
        ws4.column_dimensions["B"].width = 30
        ws4.column_dimensions["C"].width = 22
        ws4.column_dimensions["D"].width = 12
        for i in range(len(unique_groups_list)):
            ws4.column_dimensions[get_column_letter(5 + i)].width = 4

        # Font for 1/0 cells
        font_one = Font(name="Calibri", size=10, bold=True, color="22C55E")
        font_zero = Font(name="Calibri", size=10, color="334155")
        fill_hit = PatternFill(start_color="0A2A0A", end_color="0A2A0A", fill_type="solid")

        # Data rows
        for row_idx, tid in enumerate(sorted_tids, 3):
            count = tech_group_count[tid]
            bg = _BG_DARK if (row_idx - 3) % 2 == 0 else _BG_ALT
            row_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

            # Technique ID
            c = ws4.cell(row=row_idx, column=1, value=tid)
            c.font = Font(name="Courier New", size=10, bold=True, color="22C55E")
            c.fill = row_fill
            c.alignment = align_left
            c.border = _THIN_BORDER

            # Technique Name
            c = ws4.cell(row=row_idx, column=2, value=tech_names.get(tid, ""))
            c.font = Font(name="Calibri", size=10, color="E0F2FE")
            c.fill = row_fill
            c.alignment = align_left
            c.border = _THIN_BORDER

            # Tactic
            c = ws4.cell(row=row_idx, column=3, value=tech_tactics.get(tid, ""))
            c.font = Font(name="Calibri", size=10, color="FACC15")
            c.fill = row_fill
            c.alignment = align_left
            c.border = _THIN_BORDER

            # Group Count
            c = ws4.cell(row=row_idx, column=4, value=count)
            c.font = Font(name="Calibri", size=10, bold=True, color="0EA5E9")
            c.fill = row_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _THIN_BORDER

            # Group columns — 1 or empty
            for gi, gname in enumerate(unique_groups_list):
                has = tid in group_techs.get(gname, set())
                c = ws4.cell(row=row_idx, column=5 + gi, value=1 if has else None)
                if has:
                    c.font = font_one
                    c.fill = fill_hit
                else:
                    c.font = font_zero
                    c.fill = row_fill
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = _THIN_BORDER

            ws4.row_dimensions[row_idx].height = 18

        ws4.freeze_panes = "E3"
        last_data = 2 + len(sorted_tids)
        if sorted_tids:
            ws4.auto_filter.ref = f"A2:{last_matrix_col}{last_data}"

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------
    wb.save(output_path)
